from __future__ import annotations

from datetime import UTC, datetime
from hashlib import sha256
import json
from pathlib import Path
import shutil
import threading

import duckdb
from openpyxl import Workbook
import pytest
import mrs3.performance_v2_retest as retest_module

from mrs3.performance_v2_retest import (
    RetestStatus,
    build_retest_manifest,
    mark_retest_from_audit,
    retest_status,
)
from mrs3.performance_v2_store import PerformanceV2StoreError, initialize_performance_v2
from mrs3.panel_strategy_batch import validate_strategy_manifest
from mrs3.performance_v2_input import adapt_strategy_identity
from mrs3.performance_v2_import import PerformanceV2ImportRequest, import_performance_v2
from mrs3.performance_v2_store import PerformanceV2Config, performance_v2_database_path


def _strategy(
    connection: duckdb.DuckDBPyConnection,
    *,
    name: str,
    lifecycle_status: str = "ACTIVE",
    current_result_id: int | None = 1,
) -> int:
    strategy_id = connection.execute(
        """
        insert into strategies (
            strategy_name, symbol, side, timeframe, close_ma_len, order_count,
            analysis_run_id, candidate_identity, lifecycle_status, current_result_id,
            created_at_utc, updated_at_utc
        ) values (?, 'BTCUSDT', 'LONG', '1h', 20, 1, 'run-a', ?, ?, ?, now(), now())
        returning strategy_id
        """,
        [name, name, lifecycle_status, current_result_id],
    ).fetchone()[0]
    return int(strategy_id)


def _write_workbook(path: Path, high: list[object], review: list[object], *, header: object = "Strategy ID") -> Path:
    workbook = Workbook()
    high_sheet = workbook.active
    high_sheet.title = "HIGH"
    review_sheet = workbook.create_sheet("REVIEW")
    for sheet, values in ((high_sheet, high), (review_sheet, review)):
        sheet.append([header, "note"])
        for value in values:
            sheet.append([value, "ok"] if not isinstance(value, tuple) else list(value))
    workbook.create_sheet("EXTRA")
    workbook.save(path)
    return path


def test_retest_status_counts_only_active_rows_with_results() -> None:
    with duckdb.connect(":memory:") as connection:
        initialize_performance_v2(connection)
        active = _strategy(connection, name="active")
        discarded = _strategy(connection, name="discarded", lifecycle_status="DISCARDED")
        without_result = _strategy(connection, name="without-result", current_result_id=None)
        for strategy_id in (active, discarded, without_result):
            connection.execute(
                "insert into strategy_tags (strategy_id, tag, source, source_ref, updated_at_utc) values (?, 'RETEST', 'TEST', 'fixture', now())",
                [strategy_id],
            )

        assert retest_status(connection) == RetestStatus(active_count=1)


def test_retest_and_rejected_tags_coexist() -> None:
    with duckdb.connect(":memory:") as connection:
        initialize_performance_v2(connection)
        strategy_id = _strategy(connection, name="both-tags")
        connection.execute(
            "insert into strategy_tags (strategy_id, tag, source, source_ref, updated_at_utc) values (?, 'REJECTED', 'TEST', 'rejected', now()), (?, 'RETEST', 'TEST', 'retest', now())",
            [strategy_id, strategy_id],
        )

        assert connection.execute(
            "select tag, source_ref from strategy_tags where strategy_id = ? order by tag", [strategy_id]
        ).fetchall() == [("REJECTED", "rejected"), ("RETEST", "retest")]
        assert retest_status(connection) == RetestStatus(active_count=1)


def test_audit_seed_preserves_existing_rejected_tag_and_strategy_state(tmp_path: Path) -> None:
    with duckdb.connect(":memory:") as connection:
        initialize_performance_v2(connection)
        strategy_id = _strategy(connection, name="preserve-rejected")
        old_timestamp = datetime(2026, 1, 1, tzinfo=UTC)
        connection.execute(
            "insert into strategy_tags (strategy_id, tag, source, source_ref, updated_at_utc) values (?, 'REJECTED', 'SELECTION_REVIEW', 'review-old', ?)",
            [strategy_id, old_timestamp],
        )
        before_strategy = connection.execute(
            "select lifecycle_status, current_result_id from strategies where strategy_id = ?", [strategy_id]
        ).fetchone()
        workbook = _write_workbook(tmp_path / "preserve.xlsx", [strategy_id], [])

        assert mark_retest_from_audit(connection, workbook) == 1
        assert connection.execute(
            "select source, source_ref, updated_at_utc from strategy_tags where strategy_id = ? and tag = 'REJECTED'",
            [strategy_id],
        ).fetchone() == ("SELECTION_REVIEW", "review-old", old_timestamp)
        assert connection.execute(
            "select lifecycle_status, current_result_id from strategies where strategy_id = ?", [strategy_id]
        ).fetchone() == before_strategy
        assert connection.execute(
            "select source, source_ref from strategy_tags where strategy_id = ? and tag = 'RETEST'", [strategy_id]
        ).fetchone() == ("PERIOD_INTEGRITY_AUDIT", "preserve.xlsx")


def test_mark_retest_from_audit_is_idempotent_and_reads_both_sheets(tmp_path: Path) -> None:
    with duckdb.connect(":memory:") as connection:
        initialize_performance_v2(connection)
        first = _strategy(connection, name="first")
        second = _strategy(connection, name="second")
        third = _strategy(connection, name="third")
        workbook = _write_workbook(tmp_path / "audit.xlsx", [first], [second, third])

        assert mark_retest_from_audit(connection, workbook) == 3
        assert retest_status(connection) == RetestStatus(active_count=3)
        assert mark_retest_from_audit(connection, workbook) == 3
        assert connection.execute(
            "select strategy_id, tag, source, source_ref from strategy_tags order by strategy_id"
        ).fetchall() == [
            (first, "RETEST", "PERIOD_INTEGRITY_AUDIT", "audit.xlsx"),
            (second, "RETEST", "PERIOD_INTEGRITY_AUDIT", "audit.xlsx"),
            (third, "RETEST", "PERIOD_INTEGRITY_AUDIT", "audit.xlsx"),
        ]


def test_mark_retest_accepts_header_only_sheets_without_writes(tmp_path: Path) -> None:
    with duckdb.connect(":memory:") as connection:
        initialize_performance_v2(connection)
        workbook = _write_workbook(tmp_path / "empty.xlsx", [], [])

        assert mark_retest_from_audit(connection, workbook) == 0
        assert connection.execute("select count(*) from strategy_tags where tag = 'RETEST'").fetchone() == (0,)


def test_mark_retest_rejects_missing_sheet_and_header_without_writes(tmp_path: Path) -> None:
    workbook = Workbook()
    workbook.active.title = "HIGH"
    workbook.active.append(["Strategy ID", 1])
    path = tmp_path / "missing.xlsx"
    workbook.save(path)

    with duckdb.connect(":memory:") as connection:
        initialize_performance_v2(connection)
        with pytest.raises(PerformanceV2StoreError):
            mark_retest_from_audit(connection, path)
        assert connection.execute("select count(*) from strategy_tags where tag = 'RETEST'").fetchone() == (0,)

    path = _write_workbook(tmp_path / "header.xlsx", [1], [2], header="Strategy ID ")
    with duckdb.connect(":memory:") as connection:
        initialize_performance_v2(connection)
        with pytest.raises(PerformanceV2StoreError):
            mark_retest_from_audit(connection, path)
        assert connection.execute("select count(*) from strategy_tags where tag = 'RETEST'").fetchone() == (0,)


@pytest.mark.parametrize("value", [1.5, "1", True, "=1"])
def test_mark_retest_rejects_non_integer_strategy_ids_without_writes(tmp_path: Path, value: object) -> None:
    path = _write_workbook(tmp_path / "invalid.xlsx", [value], [])
    with duckdb.connect(":memory:") as connection:
        initialize_performance_v2(connection)
        _strategy(connection, name="one")
        with pytest.raises(PerformanceV2StoreError):
            mark_retest_from_audit(connection, path)
        assert connection.execute("select count(*) from strategy_tags where tag = 'RETEST'").fetchone() == (0,)


def test_mark_retest_rejects_duplicate_ids_without_writes(tmp_path: Path) -> None:
    path = _write_workbook(tmp_path / "duplicate.xlsx", [1], [1])
    with duckdb.connect(":memory:") as connection:
        initialize_performance_v2(connection)
        _strategy(connection, name="one")
        with pytest.raises(PerformanceV2StoreError):
            mark_retest_from_audit(connection, path)
        assert connection.execute("select count(*) from strategy_tags where tag = 'RETEST'").fetchone() == (0,)


def test_mark_retest_rejects_unknown_ids_without_writes(tmp_path: Path) -> None:
    path = _write_workbook(tmp_path / "unknown.xlsx", [999999], [])
    with duckdb.connect(":memory:") as connection:
        initialize_performance_v2(connection)
        _strategy(connection, name="one")
        with pytest.raises(PerformanceV2StoreError):
            mark_retest_from_audit(connection, path)
        assert connection.execute("select count(*) from strategy_tags where tag = 'RETEST'").fetchone() == (0,)


def test_mark_retest_rejects_nonblank_row_with_missing_target(tmp_path: Path) -> None:
    path = _write_workbook(tmp_path / "missing-id.xlsx", [("", "not blank")], [])
    with duckdb.connect(":memory:") as connection:
        initialize_performance_v2(connection)
        with pytest.raises(PerformanceV2StoreError):
            mark_retest_from_audit(connection, path)
        assert connection.execute("select count(*) from strategy_tags where tag = 'RETEST'").fetchone() == (0,)


def test_build_retest_manifest_renders_typed_mixed_run_strategies(tmp_path: Path) -> None:
    template = Path("templates/strategies/retest-mrs3/base.json").resolve()
    with duckdb.connect(":memory:") as connection:
        initialize_performance_v2(connection)
        first = _strategy(connection, name="stored-long")
        second = connection.execute(
            """
            insert into strategies (
                strategy_name, symbol, side, timeframe, close_ma_len, order_count,
                analysis_run_id, candidate_identity, lifecycle_status, current_result_id,
                created_at_utc, updated_at_utc
            ) values ('stored-short', 'ETHUSDT', 'SHORT', '4h', 30, 2, 'run-b', 'candidate-b', 'ACTIVE', 2, now(), now())
            returning strategy_id
            """
        ).fetchone()[0]
        connection.execute(
            "insert into analysis_plateaus values ('run-a', 'P1', 5, 100), ('run-b', 'P2', 4, 80), ('run-b', 'P3', 6, 120)"
        )
        connection.execute(
            "insert into strategy_orders values (?, 1, 10, .99, 100, .7, 'run-a', 'P1', 15), (?, 1, 11, 1.01, 100, .6, 'run-b', 'P2', 20), (?, 2, 12, 1.02, 200, .4, 'run-b', 'P3', 25)",
            [first, second, second],
        )
        connection.execute(
            """
            insert into strategy_results (
                strategy_id, report_start_utc, report_end_utc, exchange, commission_rate,
                initial_balance, final_balance, imported_at_utc
            ) values (?, '2026-01-01', '2026-01-10', 'Bybit', .0004, 100, 101, now()),
                     (?, '2026-01-01', '2026-01-10', 'Bybit', .0004, 100, 101, now())
            """,
            [first, second],
        )
        connection.execute("update strategies set current_result_id = (select result_id from strategy_results where strategy_id = ? ) where strategy_id = ?", [first, first])
        connection.execute("update strategies set current_result_id = (select result_id from strategy_results where strategy_id = ? ) where strategy_id = ?", [second, second])
        connection.execute(
            "insert into strategy_tags values (?, 'RETEST', 'TEST', 'fixture', now()), (?, 'RETEST', 'TEST', 'fixture', now())",
            [first, second],
        )

        batch = build_retest_manifest(
            connection,
            {"LONG": template, "SHORT": template},
            tmp_path / "output",
        )

    manifest = json.loads(batch.manifest_path.read_text(encoding="utf-8"))
    assert batch.strategy_count == 2
    assert manifest["strategy_analysis_run_ids"] == {
        "stored-long.json": "run-a",
        "stored-short.json": "run-b",
    }
    validated = validate_strategy_manifest(batch.manifest_path)
    assert validated.strategy_source == (tmp_path / "output" / "strategies").resolve()
    diagnostics = manifest["candidate_diagnostics"]
    for name, expected in {
        "stored-long": ("LONG", "BTCUSDT", 20, [(10, 100, "P1", .7)]),
        "stored-short": ("SHORT", "ETHUSDT", 30, [(11, 100, "P2", .6), (12, 200, "P3", .4)]),
    }.items():
        strategy = json.loads((validated.strategy_source / f"{name}.json").read_text(encoding="utf-8"))
        identity = adapt_strategy_identity(
            strategy,
            strategy_name=name,
            order_plateau_diagnostics=diagnostics[
                next(key for key, names in manifest["candidate_identity_to_strategy_names"].items() if name in names)
            ],
        )
        assert identity.side == expected[0]
        assert identity.symbol == expected[1]
        assert identity.close_ma_len == expected[2]
        assert [(item.open_ma_len, item.shift_bp, item.plateau_id, float(item.lot_x)) for item in identity.orders] == expected[3]


def _single_order_retest_db(path: Path, names: tuple[str, ...] = ("stored",)) -> tuple[duckdb.DuckDBPyConnection, tuple[int, ...]]:
    connection = duckdb.connect(str(path))
    initialize_performance_v2(connection)
    ids: list[int] = []
    for index, name in enumerate(names, start=1):
        strategy_id = _strategy(connection, name=name)
        connection.execute(
            "insert into analysis_plateaus values (?, 'P1', 4, 80)", [f"run-{index}"]
        )
        connection.execute(
            "insert into strategy_orders values (?, 1, 7, .995, 50, 1, ?, 'P1', 20)",
            [strategy_id, f"run-{index}"],
        )
        connection.execute(
            "insert into strategy_results (strategy_id, report_start_utc, report_end_utc, exchange, commission_rate, initial_balance, final_balance, imported_at_utc) values (?, '2026-01-01', '2026-01-09', 'Bybit', .0004, 100, 101, now())",
            [strategy_id],
        )
        result_id = connection.execute("select result_id from strategy_results where strategy_id = ?", [strategy_id]).fetchone()[0]
        connection.execute("update strategies set analysis_run_id = ?, current_result_id = ? where strategy_id = ?", [f"run-{index}", result_id, strategy_id])
        connection.execute(
            "insert into strategy_tags values (?, 'RETEST', 'TEST', 'fixture', now())", [strategy_id]
        )
        ids.append(strategy_id)
    return connection, tuple(ids)


def test_retest_rejects_casefolded_filename_collision_before_output_write(tmp_path: Path) -> None:
    connection, _ = _single_order_retest_db(tmp_path / "db.duckdb", ("Same", "same"))
    output = tmp_path / "output"
    try:
        with pytest.raises(PerformanceV2StoreError, match="case-insensitively"):
            build_retest_manifest(
                connection,
                {"LONG": Path("templates/strategies/retest-mrs3/base.json")},
                output,
            )
    finally:
        connection.close()
    assert not (output / "strategies").exists()
    assert not (output / "strategy_manifest.json").exists()


def test_retest_publication_rollback_preserves_previous_batch(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    connection, _ = _single_order_retest_db(tmp_path / "db.duckdb")
    output = tmp_path / "output"
    template = Path("templates/strategies/retest-mrs3/base.json")
    first = build_retest_manifest(connection, {"LONG": template}, output)
    old_strategy = (first.strategies_path / "stored.json").read_bytes()
    old_manifest = first.manifest_path.read_bytes()
    original_rename = retest_module.Path.rename
    injected = False

    def fail_final_manifest_rename(source: Path, destination: Path) -> Path:
        nonlocal injected
        result = original_rename(source, destination)
        if not injected and source.parent.name.startswith(".retest-stage-") and destination == output / "strategies":
            injected = True
            raise OSError("injected between final renames")
        return result

    monkeypatch.setattr(retest_module.Path, "rename", fail_final_manifest_rename)
    try:
        with pytest.raises(OSError, match="injected between final renames"):
            build_retest_manifest(connection, {"LONG": template}, output)
    finally:
        connection.close()
    assert injected
    assert (output / "strategies" / "stored.json").read_bytes() == old_strategy
    assert (output / "strategy_manifest.json").read_bytes() == old_manifest


def test_retest_publication_rollback_restores_after_old_strategy_move_failure(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    connection, _ = _single_order_retest_db(tmp_path / "db.duckdb")
    output = tmp_path / "output"
    template = Path("templates/strategies/retest-mrs3/base.json")
    first = build_retest_manifest(connection, {"LONG": template}, output)
    old_strategy = (first.strategies_path / "stored.json").read_bytes()
    old_manifest = first.manifest_path.read_bytes()
    original_rename = retest_module.Path.rename
    injected = False

    def fail_old_strategy_move(source: Path, destination: Path) -> Path:
        nonlocal injected
        result = original_rename(source, destination)
        if not injected and source == output / "strategies" and destination.parent.name.startswith(".retest-backup-"):
            injected = True
            raise OSError("injected old strategy move failure")
        return result

    monkeypatch.setattr(retest_module.Path, "rename", fail_old_strategy_move)
    try:
        with pytest.raises(OSError, match="injected old strategy move failure"):
            build_retest_manifest(connection, {"LONG": template}, output)
    finally:
        connection.close()
    assert injected
    assert (output / "strategies" / "stored.json").read_bytes() == old_strategy
    assert (output / "strategy_manifest.json").read_bytes() == old_manifest


def test_retest_rollback_failure_keeps_backup_and_both_failure_causes(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    connection, _ = _single_order_retest_db(tmp_path / "db.duckdb")
    output = tmp_path / "output"
    template = Path("templates/strategies/retest-mrs3/base.json")
    build_retest_manifest(connection, {"LONG": template}, output)
    original_rename = retest_module.Path.rename
    original_rmtree = retest_module.shutil.rmtree

    def fail_publish(source: Path, destination: Path) -> Path:
        result = original_rename(source, destination)
        if source.parent.name.startswith(".retest-stage-") and destination == output / "strategies":
            raise OSError("injected publish failure")
        return result

    def fail_rollback(path: Path, *args: object, **kwargs: object) -> None:
        if path == output / "strategies":
            raise OSError("injected rollback failure")
        original_rmtree(path, *args, **kwargs)

    monkeypatch.setattr(retest_module.Path, "rename", fail_publish)
    monkeypatch.setattr(retest_module.shutil, "rmtree", fail_rollback)
    try:
        with pytest.raises(PerformanceV2StoreError, match="injected publish failure.*injected rollback failure") as error:
            build_retest_manifest(connection, {"LONG": template}, output)
    finally:
        connection.close()
    backup_paths = list(output.glob(".retest-backup-*"))
    assert len(backup_paths) == 1
    assert str(backup_paths[0]) in str(error.value)
    assert (backup_paths[0] / "strategies" / "stored.json").is_file()


def test_retest_publication_replaces_existing_output_successfully(tmp_path: Path) -> None:
    connection, _ = _single_order_retest_db(tmp_path / "db.duckdb")
    output = tmp_path / "output"
    template = Path("templates/strategies/retest-mrs3/base.json")
    try:
        first = build_retest_manifest(connection, {"LONG": template}, output)
        first_run_id = json.loads(first.manifest_path.read_text(encoding="utf-8"))["analysis_run_id"]
        second = build_retest_manifest(connection, {"LONG": template}, output)
        second_run_id = json.loads(second.manifest_path.read_text(encoding="utf-8"))["analysis_run_id"]
    finally:
        connection.close()
    assert first.manifest_path == second.manifest_path
    assert first_run_id != second_run_id
    assert sorted(path.name for path in output.iterdir()) == ["strategies", "strategy_manifest.json"]


def test_retest_publication_has_process_local_single_writer_guard(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    connection, _ = _single_order_retest_db(tmp_path / "db.duckdb")
    output = tmp_path / "output"
    template = Path("templates/strategies/retest-mrs3/base.json")
    first = build_retest_manifest(connection, {"LONG": template}, output)
    connection.close()
    generated = [json.loads((first.strategies_path / "stored.json").read_text(encoding="utf-8"))]
    manifest = json.loads(first.manifest_path.read_text(encoding="utf-8"))
    entered = threading.Event()
    release = threading.Event()
    original_reconcile = retest_module._reconcile_retest_publication

    def hold_reconcile(path: Path) -> None:
        entered.set()
        assert release.wait(2)
        original_reconcile(path)

    monkeypatch.setattr(retest_module, "_reconcile_retest_publication", hold_reconcile)
    errors: list[BaseException] = []

    def publish() -> None:
        try:
            retest_module._publish_retest(output, generated, manifest)
        except BaseException as error:
            errors.append(error)

    worker = threading.Thread(target=publish)
    worker.start()
    assert entered.wait(2)
    acquired = retest_module._RETEST_PUBLICATION_LOCK.acquire(blocking=False)
    if acquired:
        retest_module._RETEST_PUBLICATION_LOCK.release()
    assert not acquired
    release.set()
    worker.join(5)
    assert not worker.is_alive()
    assert not errors


def test_retest_publication_has_cross_process_writer_guard(tmp_path: Path) -> None:
    output = tmp_path / "output"

    with retest_module._RetestPublicationLock(output):
        with pytest.raises(PerformanceV2StoreError, match="output is busy"):
            with retest_module._RetestPublicationLock(output):
                pass


def test_retest_staged_validation_failure_keeps_existing_output(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    connection, _ = _single_order_retest_db(tmp_path / "db.duckdb")
    output = tmp_path / "output"
    template = Path("templates/strategies/retest-mrs3/base.json")
    first = build_retest_manifest(connection, {"LONG": template}, output)
    old_strategy = (first.strategies_path / "stored.json").read_bytes()
    old_manifest = first.manifest_path.read_bytes()
    original_validate = retest_module.validate_strategy_manifest

    def fail_staged_validation(path: Path):
        if path.parent.name.startswith(".retest-stage-"):
            raise PerformanceV2StoreError("injected staged validation failure")
        return original_validate(path)

    monkeypatch.setattr(retest_module, "validate_strategy_manifest", fail_staged_validation)
    try:
        with pytest.raises(PerformanceV2StoreError, match="injected staged validation failure"):
            build_retest_manifest(connection, {"LONG": template}, output)
    finally:
        connection.close()
    assert (output / "strategies" / "stored.json").read_bytes() == old_strategy
    assert (output / "strategy_manifest.json").read_bytes() == old_manifest


def test_retest_binds_manifest_hash_to_serialized_staged_strategy(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    connection, _ = _single_order_retest_db(tmp_path / "db.duckdb")
    output = tmp_path / "output"
    original_publish = retest_module._publish_strategies

    def mutate_staged_strategy(stage: Path, variants, generated):
        result = original_publish(stage, variants, generated)
        strategy_path = stage / "strategies" / "stored.json"
        strategy = json.loads(strategy_path.read_text(encoding="utf-8"))
        strategy["mrs3"]["ma_long"][0]["len"] = 999  # type: ignore[index]
        strategy_path.write_text(json.dumps(strategy), encoding="utf-8")
        return result

    monkeypatch.setattr(retest_module, "_publish_strategies", mutate_staged_strategy)
    try:
        with pytest.raises(PerformanceV2StoreError, match="staged strategy hash mismatch"):
            build_retest_manifest(
                connection,
                {"LONG": Path("templates/strategies/retest-mrs3/base.json")},
                output,
            )
    finally:
        connection.close()
    assert not (output / "strategies").exists()
    assert not (output / "strategy_manifest.json").exists()


def test_retest_rejects_extra_staged_strategy_file_before_manifest_validation(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    connection, _ = _single_order_retest_db(tmp_path / "db.duckdb")
    output = tmp_path / "output"
    original_publish = retest_module._publish_strategies

    def add_extra_staged_file(stage: Path, variants, generated):
        result = original_publish(stage, variants, generated)
        (stage / "strategies" / "extra.json").write_text("{}", encoding="utf-8")
        return result

    monkeypatch.setattr(retest_module, "_publish_strategies", add_extra_staged_file)
    try:
        with pytest.raises(PerformanceV2StoreError, match="files do not match manifest hash map"):
            build_retest_manifest(
                connection,
                {"LONG": Path("templates/strategies/retest-mrs3/base.json")},
                output,
            )
    finally:
        connection.close()
    assert not (output / "strategies").exists()
    assert not (output / "strategy_manifest.json").exists()


def test_retest_recovers_stale_backup_before_republication(tmp_path: Path) -> None:
    connection, _ = _single_order_retest_db(tmp_path / "db.duckdb")
    output = tmp_path / "output"
    template = Path("templates/strategies/retest-mrs3/base.json")
    try:
        first = build_retest_manifest(connection, {"LONG": template}, output)
        stale_backup = output / ".retest-backup-stale"
        stale_backup.mkdir()
        first.strategies_path.rename(stale_backup / "strategies")
        first.manifest_path.rename(stale_backup / "strategy_manifest.json")
        second = build_retest_manifest(connection, {"LONG": template}, output)
    finally:
        connection.close()
    validate_strategy_manifest(second.manifest_path)
    assert not stale_backup.exists()
    assert sorted(path.name for path in output.iterdir()) == ["strategies", "strategy_manifest.json"]


def test_retest_recovers_stale_stage_before_republication(tmp_path: Path) -> None:
    connection, _ = _single_order_retest_db(tmp_path / "db.duckdb")
    output = tmp_path / "output"
    template = Path("templates/strategies/retest-mrs3/base.json")
    try:
        first = build_retest_manifest(connection, {"LONG": template}, output)
        stale_stage = output / ".retest-stage-stale"
        stale_stage.mkdir()
        first.strategies_path.rename(stale_stage / "strategies")
        first.manifest_path.rename(stale_stage / "strategy_manifest.json")
        second = build_retest_manifest(connection, {"LONG": template}, output)
    finally:
        connection.close()
    validate_strategy_manifest(second.manifest_path)
    assert not stale_stage.exists()
    assert sorted(path.name for path in output.iterdir()) == ["strategies", "strategy_manifest.json"]


def test_retest_rejects_ambiguous_stale_publication_candidates(tmp_path: Path) -> None:
    connection, _ = _single_order_retest_db(tmp_path / "db.duckdb")
    output = tmp_path / "output"
    template = Path("templates/strategies/retest-mrs3/base.json")
    try:
        first = build_retest_manifest(connection, {"LONG": template}, output)
        for suffix in ("one", "two"):
            backup = output / f".retest-backup-{suffix}"
            shutil.copytree(first.strategies_path, backup / "strategies")
            shutil.copyfile(first.manifest_path, backup / "strategy_manifest.json")
        with pytest.raises(PerformanceV2StoreError, match="ambiguous RETEST publication"):
            build_retest_manifest(connection, {"LONG": template}, output)
    finally:
        connection.close()
    validate_strategy_manifest(first.manifest_path)
    assert (output / "strategies" / "stored.json").is_file()


def test_retest_rejects_null_base_point_trades_before_rendering(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    connection, (strategy_id,) = _single_order_retest_db(tmp_path / "db.duckdb")

    class _ResultWithNullBase:
        def __init__(self, result):
            self._result = result

        def fetchall(self):
            rows = self._result.fetchall()
            return [row[:8] + (None,) + row[9:] for row in rows]

    class _ConnectionWithNullBase:
        def execute(self, sql, parameters=None):
            result = connection.execute(sql) if parameters is None else connection.execute(sql, parameters)
            if "select o.strategy_id" in sql:
                return _ResultWithNullBase(result)
            return result

    monkeypatch.setattr(retest_module, "require_performance_v2", lambda _connection: None)
    try:
        with pytest.raises(PerformanceV2StoreError, match="stored"):
            build_retest_manifest(
                _ConnectionWithNullBase(),
                {"LONG": Path("templates/strategies/retest-mrs3/base.json")},
                tmp_path / "output",
            )
    finally:
        connection.close()
    assert not (tmp_path / "output").exists()


@pytest.mark.parametrize(
    ("field", "value"),
    [("strategy_name", None), ("symbol", None), ("timeframe", None), ("close_ma_len", None),
     ("analysis_run_id", None), ("candidate_identity", 1), ("order_analysis_run_id", None)],
)
def test_retest_rejects_untyped_identity_fields_before_output(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch, field: str, value: object
) -> None:
    connection, _ = _single_order_retest_db(tmp_path / "db.duckdb")

    class _MutatedResult:
        def __init__(self, result, index: int, replacement: object):
            self._result = result
            self._index = index
            self._replacement = replacement

        def fetchall(self):
            rows = []
            for row in self._result.fetchall():
                values = list(row)
                values[self._index] = self._replacement
                rows.append(tuple(values))
            return rows

    field_indexes = {
        "strategy_name": ("select s.strategy_id, s.strategy_name", 1),
        "symbol": ("select s.strategy_id, s.strategy_name", 2),
        "timeframe": ("select s.strategy_id, s.strategy_name", 4),
        "close_ma_len": ("select s.strategy_id, s.strategy_name", 5),
        "analysis_run_id": ("select s.strategy_id, s.strategy_name", 7),
        "candidate_identity": ("select s.strategy_id, s.strategy_name", 8),
        "order_analysis_run_id": ("select o.strategy_id", 6),
    }
    marker, index = field_indexes[field]

    class _MutatingConnection:
        def execute(self, sql, parameters=None):
            result = connection.execute(sql) if parameters is None else connection.execute(sql, parameters)
            if marker in sql:
                return _MutatedResult(result, index, value)
            return result

    monkeypatch.setattr(retest_module, "require_performance_v2", lambda _connection: None)
    try:
        with pytest.raises(PerformanceV2StoreError):
            build_retest_manifest(
                _MutatingConnection(),
                {"LONG": Path("templates/strategies/retest-mrs3/base.json")},
                tmp_path / "output",
            )
    finally:
        connection.close()
    assert not (tmp_path / "output").exists()


def _mixed_ingest_retest_db(path: Path) -> duckdb.DuckDBPyConnection:
    connection = duckdb.connect(str(path))
    initialize_performance_v2(connection)
    for index, (name, run_id, plateau_id, lot) in enumerate(
        (("mixed-first", "source-run-a", "P1", "0.7"), ("mixed-second", "source-run-b", "P2", "0.3")),
        start=1,
    ):
        strategy_id = connection.execute(
            """
            insert into strategies (
                strategy_name, symbol, side, timeframe, close_ma_len, order_count,
                analysis_run_id, candidate_identity, lifecycle_status, current_result_id,
                created_at_utc, updated_at_utc
            ) values (?, 'ONUSDT', 'LONG', '1h', 3, 1, ?, ?, 'ACTIVE', null, now(), now())
            returning strategy_id
            """,
            [name, run_id, f"candidate-{index}"],
        ).fetchone()[0]
        connection.execute(
            "insert into analysis_plateaus values (?, ?, 4, ?)", [run_id, plateau_id, 80 + index]
        )
        connection.execute(
            "insert into strategy_orders values (?, 1, 7, .995, 50, ?, ?, ?, ?)",
            [strategy_id, lot, run_id, plateau_id, 20 + index],
        )
        result_id = connection.execute(
            """
            insert into strategy_results (
                strategy_id, report_start_utc, report_end_utc, exchange, commission_rate,
                initial_balance, final_balance, imported_at_utc
            ) values (?, '2026-01-01', '2026-01-09', 'Bybit', .0004, 100, 101, now())
            returning result_id
            """,
            [strategy_id],
        ).fetchone()[0]
        connection.execute("update strategies set current_result_id = ? where strategy_id = ?", [result_id, strategy_id])
        connection.execute(
            "insert into strategy_tags values (?, 'RETEST', 'TEST', 'fixture', now())", [strategy_id]
        )
    return connection


def test_retest_manifest_round_trips_through_v2_store_ingest_with_mixed_runs(tmp_path: Path) -> None:
    source = _mixed_ingest_retest_db(tmp_path / "source.duckdb")
    output = tmp_path / "output"
    template = Path("templates/strategies/retest-mrs3/base.json")
    try:
        batch = build_retest_manifest(source, {"LONG": template}, output)
    finally:
        source.close()

    inbox = tmp_path / "inbox"
    strategies = inbox / "strategies"
    reports = tmp_path / "reports"
    strategies.mkdir(parents=True)
    reports.mkdir()
    normal = json.loads(batch.manifest_path.read_text(encoding="utf-8"))
    strategy_hashes: dict[str, str] = {}
    entries: list[dict[str, object]] = []
    fixture = Path(__file__).parent / "fixtures" / "performance" / "report_current_v2.html"
    for index, filename in enumerate(sorted(normal["strategy_json_sha256"]), start=1):
        source_path = batch.strategies_path / filename
        target_path = strategies / filename
        shutil.copyfile(source_path, target_path)
        canonical_hash = normal["strategy_json_sha256"][filename]
        strategy_hashes[filename] = canonical_hash
        name = target_path.stem
        report_path = reports / f"{name}.html"
        report_bytes = fixture.read_bytes() + f"\n<!-- {name} -->\n".encode("utf-8")
        report_path.write_bytes(report_bytes)
        report_hash = sha256(report_bytes).hexdigest()
        strategy_bytes = target_path.read_bytes()
        candidate = next(key for key, names in normal["candidate_identity_to_strategy_names"].items() if name in names)
        entries.append(
            {
                "manifest_entry_id": f"{index:032x}",
                "strategy_name": name,
                "strategy_version_id": canonical_hash,
                "strategy_path": str(target_path),
                "report_path": str(report_path),
                "wizard_run_id": "retest-round-trip",
                "exchange_name": "Bybit",
                "source_strategy_sha256": sha256(strategy_bytes).hexdigest(),
                "source_report_sha256": report_hash,
                "candidate_identity": candidate,
            }
        )
    inbox_manifest = {
        "schema_version": 1,
        "batch_id": batch.run_id,
        "expected_strategy_names": [entry["strategy_name"] for entry in entries],
        "tester_config_sha256": "t" * 64,
        "commission_contract": {
            "MakerFee": "0.0002",
            "TakerFee": "0.0004",
            "SlippagePercent": "0.01",
            "FundingRate": "0.0001",
            "FundingIntervalHours": "8",
        },
        "commission_contract_id": "c" * 64,
        "run_mode": "FAST",
        "entries": entries,
        "v6_provenance": {
            "analysis_run_id": batch.run_id,
            "generation_manifest_sha256": normal["generation_manifest_sha256"],
            "strategy_json_sha256": strategy_hashes,
            "strategy_analysis_run_ids": normal["strategy_analysis_run_ids"],
            "candidate_identity_to_strategy_names": normal["candidate_identity_to_strategy_names"],
            "candidate_diagnostics": normal["candidate_diagnostics"],
        },
    }
    (inbox / "inbox_manifest.json").write_text(json.dumps(inbox_manifest), encoding="utf-8")
    config = PerformanceV2Config(tmp_path / "v2-store", workers=1)
    target = performance_v2_database_path(config)
    target.parent.mkdir(parents=True)
    dates_path = tmp_path / "Input" / "dates.xlsx"
    dates_path.parent.mkdir()
    workbook = Workbook()
    workbook.active.append(["ONUSDT", datetime(2025, 12, 25)])
    workbook.save(dates_path)
    with duckdb.connect(str(target)) as connection:
        initialize_performance_v2(connection)

    result = import_performance_v2(
        PerformanceV2ImportRequest(inbox, reports, config, listing_dates_path=Path("Input/dates.xlsx"))
    )
    assert result.imported_count == 2
    with duckdb.connect(str(target), read_only=True) as connection:
        rows = connection.execute(
            """
            select s.strategy_name, s.analysis_run_id, o.analysis_run_id, p.analysis_run_id
            from strategies s
            join strategy_orders o on o.strategy_id = s.strategy_id
            join analysis_plateaus p on p.analysis_run_id = o.analysis_run_id and p.plateau_id = o.plateau_id
            order by s.strategy_name
            """
        ).fetchall()
        assert rows == [
            ("mixed-first", "source-run-a", "source-run-a", "source-run-a"),
            ("mixed-second", "source-run-b", "source-run-b", "source-run-b"),
        ]
        plateaus_before = connection.execute(
            "select analysis_run_id, plateau_id, plateau_point_count, plateau_total_trades "
            "from analysis_plateaus order by analysis_run_id, plateau_id"
        ).fetchall()

    repeated = import_performance_v2(
        PerformanceV2ImportRequest(inbox, reports, config, listing_dates_path=Path("Input/dates.xlsx"))
    )
    assert repeated.imported_count == 0
    assert repeated.skipped_count == 2
    with duckdb.connect(str(target), read_only=True) as connection:
        assert connection.execute(
            """
            select s.strategy_name, s.analysis_run_id, o.analysis_run_id, p.analysis_run_id
            from strategies s
            join strategy_orders o on o.strategy_id = s.strategy_id
            join analysis_plateaus p on p.analysis_run_id = o.analysis_run_id and p.plateau_id = o.plateau_id
            order by s.strategy_name
            """
        ).fetchall() == rows
        assert connection.execute(
            "select analysis_run_id, plateau_id, plateau_point_count, plateau_total_trades "
            "from analysis_plateaus order by analysis_run_id, plateau_id"
        ).fetchall() == plateaus_before
