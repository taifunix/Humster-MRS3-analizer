from __future__ import annotations

from dataclasses import replace
from hashlib import sha256
import csv
import json
from datetime import datetime, timedelta, timezone
from decimal import Decimal
from pathlib import Path

import duckdb
import pytest
from openpyxl import Workbook, load_workbook
import mrs3.performance_v2_import as import_module

from mrs3.performance_v2_import import (
    PerformanceV2ImportError,
    PerformanceV2ImportRequest,
    PerformanceV2LockedError,
    import_performance_v2,
)
from mrs3.performance_v2_html import parse_current_performance_v2_html
from mrs3.performance_v2_input import read_performance_v2_inbox
from mrs3.performance_v2_store import (
    PerformanceV2Config,
    PerformanceV2StoreError,
    PerformanceV2WriterLock,
    initialize_performance_v2,
    performance_v2_database_path,
)


FIXTURE = Path(__file__).parent / "fixtures" / "performance" / "report_current_v2.html"


def _strategy(name: str, *, orders: int = 1, close_ma: int = 3) -> dict[str, object]:
    return {
        "name": name,
        "exchange": {"name": "Bybit", "use_upnl": True},
        "basic": {
            "strategy": "mrs3",
            "symbol": "ONUSDT",
            "time_frame": "1h",
            "use_long": True,
            "use_short": False,
        },
        "mrs3": {
            "ma_long": [
                {"id": i, "len": 6 + i, "multiplier": 1 - i / 1000, "lot_x": 1 / orders}
                for i in range(1, orders + 1)
            ],
            "ma_short": [],
            "ma_close_long": {"len": close_ma},
            "ma_close_short": {"len": close_ma},
        },
    }


def _canonical_strategy_hash(strategy: dict[str, object]) -> str:
    return sha256(
        json.dumps(strategy, ensure_ascii=False, sort_keys=True, separators=(",", ":")).encode()
    ).hexdigest()


def _inbox(tmp_path: Path, names: tuple[str, ...] = ("alpha",), *, orders: int = 1) -> tuple[Path, Path, bytes]:
    inbox = tmp_path / "inbox"
    strategy_root = inbox / "strategies"
    report_root = tmp_path / "reports"
    strategy_root.mkdir(parents=True)
    report_root.mkdir()
    report = FIXTURE.read_bytes()
    diagnostics: dict[str, object] = {}
    entries: list[dict[str, object]] = []
    for index, name in enumerate(names, start=1):
        strategy = _strategy(name, orders=orders, close_ma=3 + index - 1)
        strategy_bytes = json.dumps(strategy, separators=(",", ":")).encode()
        strategy_path = strategy_root / f"{name}.json"
        strategy_path.write_bytes(strategy_bytes)
        report_path = report_root / f"{name}.html"
        report_path.write_bytes(report)
        diagnostics[f"candidate-{index}"] = {
            "order_count": orders,
            "orders": [
                {
                    "order_id": order_id,
                    "plateau_id": f"P{order_id}",
                    "plateau_point_count": 4,
                    "base_point_trades": 20,
                    "plateau_total_trades": 80,
                }
                for order_id in range(1, orders + 1)
            ],
        }
        entries.append(
            {
                "manifest_entry_id": f"{index:032x}",
                "strategy_name": name,
                "strategy_version_id": _canonical_strategy_hash(strategy),
                "strategy_path": str(strategy_path),
                "report_path": str(report_path),
                "wizard_run_id": "run-1",
                "exchange_name": "Bybit",
                "source_strategy_sha256": sha256(strategy_bytes).hexdigest(),
                "source_report_sha256": sha256(report).hexdigest(),
            }
        )
    manifest = {
        "schema_version": 1,
        "batch_id": "v2-test",
        "expected_strategy_names": list(names),
        "tester_config_sha256": "t" * 64,
        "commission_contract": {
            "MakerFee": "0.0002",
            "TakerFee": "0.0004",
            "SlippagePercent": "0.01",
            "FundingRate": "0.0001",
            "FundingIntervalHours": "8",
        },
        "commission_contract_id": "c" * 64,
        "run_mode": "FAST",
        "entries": entries,
        "v6_provenance": {
            "analysis_run_id": "a" * 64,
            "generation_manifest_sha256": "g" * 64,
            "strategy_json_sha256": {f"{name}.json": entries[i]["strategy_version_id"] for i, name in enumerate(names)},
            "candidate_identity_to_strategy_names": {
                f"candidate-{i}": [name] for i, name in enumerate(names, start=1)
            },
            "candidate_diagnostics": diagnostics,
        },
    }
    (inbox / "inbox_manifest.json").write_text(json.dumps(manifest), encoding="utf-8")
    snapshot = sha256(
        b"".join(
            path.relative_to(inbox).as_posix().encode() + b"\0" + path.read_bytes()
            for path in sorted(inbox.rglob("*"))
            if path.is_file()
        )
    ).digest()
    return inbox, report_root, snapshot


def _request(
    tmp_path: Path,
    *,
    names: tuple[str, ...] = ("alpha",),
    orders: int = 1,
    mode: str = "ADD",
    initialize_db: bool = True,
) -> tuple[PerformanceV2ImportRequest, bytes]:
    inbox, report_root, snapshot = _inbox(tmp_path, names, orders=orders)
    config = PerformanceV2Config(tmp_path / "v2", workers=4)
    dates_path = tmp_path / "Input" / "dates.xlsx"
    dates_path.parent.mkdir()
    workbook = Workbook()
    workbook.active.append(["ONUSDT", datetime(2025, 12, 25)])
    workbook.save(dates_path)
    request = PerformanceV2ImportRequest(
        inbox, report_root, config, mode=mode, listing_dates_path=Path("Input/dates.xlsx")
    )
    if initialize_db:
        _db(request)
    return request, snapshot


def _db(request: PerformanceV2ImportRequest) -> Path:
    target = performance_v2_database_path(request.config)
    target.parent.mkdir(parents=True, exist_ok=True)
    with duckdb.connect(str(target)) as connection:
        initialize_performance_v2(connection)
    return target


def _rewrite_report(request: PerformanceV2ImportRequest, replacement: bytes) -> None:
    report_path = request.report_root / "alpha.html"
    report_path.write_bytes(replacement)
    manifest_path = request.inbox / "inbox_manifest.json"
    manifest = json.loads(manifest_path.read_text())
    manifest["entries"][0]["source_report_sha256"] = sha256(replacement).hexdigest()
    manifest_path.write_text(json.dumps(manifest))


def test_append_rows_uses_duckdb_native_dataframe_append() -> None:
    class AppendOnlyConnection:
        def __init__(self) -> None:
            self.connection = duckdb.connect(":memory:")
            self.connection.execute("create table rows_to_append (id integer, amount decimal(38, 12))")

        def append(self, table: str, frame) -> None:
            self.connection.append(table, frame)

        def executemany(self, *_args: object) -> None:
            raise AssertionError("large report rows must not use executemany")

    connection = AppendOnlyConnection()
    import_module._append_rows(
        connection, "rows_to_append", ("id", "amount"), [(1, Decimal("1.25")), (2, Decimal("2.50"))]
    )

    assert connection.connection.execute("select * from rows_to_append order by id").fetchall() == [
        (1, Decimal("1.250000000000")),
        (2, Decimal("2.500000000000")),
    ]


def test_append_rows_rounds_long_decimal_before_dataframe_type_inference() -> None:
    class CaptureConnection:
        frame = None

        def append(self, _table: str, frame) -> None:
            self.frame = frame

    connection = CaptureConnection()

    import_module._append_rows(
        connection,
        "rows_to_append",
        ("id", "amount"),
        [(1, Decimal("-29.769149208741522230595327812"))],
    )

    assert connection.frame.iloc[0]["amount"] == "-29.769149208742"

    target = duckdb.connect(":memory:")
    target.execute("create table rows_to_append (id integer, amount decimal(38, 12))")
    import_module._append_rows(
        target,
        "rows_to_append",
        ("id", "amount"),
        [(1, Decimal("-29.769149208741522230595327812"))],
    )
    assert target.execute("select amount from rows_to_append").fetchone() == (
        Decimal("-29.769149208742"),
    )


def test_add_publishes_multiple_strategies_and_one_current_result_each(tmp_path: Path) -> None:
    request, snapshot = _request(tmp_path, names=("alpha", "beta"), orders=2)
    inbox_bytes = (request.inbox / "inbox_manifest.json").read_bytes()

    result = import_performance_v2(request)

    assert result.imported_count == 2
    assert result.skipped_count == 0
    assert result.rejected_count == 0
    target = performance_v2_database_path(request.config)
    with duckdb.connect(str(target), read_only=True) as connection:
        assert connection.execute("select count(*) from strategies").fetchone() == (2,)
        assert connection.execute("select count(*) from strategy_orders").fetchone() == (4,)
        assert connection.execute("select count(*) from strategy_results").fetchone() == (2,)
        assert connection.execute("select count(*) from strategies where current_result_id is not null").fetchone() == (2,)
    assert (request.inbox / "inbox_manifest.json").read_bytes() == inbox_bytes
    assert snapshot


def test_import_reports_parse_progress_for_each_completed_report(tmp_path: Path) -> None:
    request, _ = _request(tmp_path, names=("alpha", "beta"))
    events: list[tuple[str, int, int]] = []

    import_performance_v2(
        request,
        progress=lambda stage, completed, total: events.append((stage, completed, total)),
    )

    assert events[0] == ("PARSING", 0, 2)
    assert [completed for stage, completed, total in events if stage == "PARSING"] == [0, 1, 2]
    assert events[-1] == ("PUBLISHING", 2, 2)


def test_add_accepts_tester_report_order_ids_outside_mrs3_order_slots(tmp_path: Path) -> None:
    request, _ = _request(tmp_path)
    report = FIXTURE.read_bytes().replace(b"<td>1</td><td>opened</td>", b"<td>2</td><td>opened</td>", 1)
    report = report.replace(b"<td>1</td><td>closed</td>", b"<td>2</td><td>closed</td>", 1)
    _rewrite_report(request, report)

    assert import_performance_v2(request).imported_count == 1


def test_identical_current_payload_is_skipped_and_changed_add_is_rejected(tmp_path: Path) -> None:
    request, _ = _request(tmp_path)
    first = import_performance_v2(request)
    assert first.imported_count == 1

    second = import_performance_v2(request)
    assert second.imported_count == 0
    assert second.skipped_count == 1

    changed_strategy = json.loads((request.inbox / "strategies" / "alpha.json").read_text())
    changed_strategy["mrs3"]["ma_long"][0]["len"] = 99  # type: ignore[index]
    strategy_path = request.inbox / "strategies" / "alpha.json"
    strategy_bytes = json.dumps(changed_strategy, separators=(",", ":")).encode()
    strategy_path.write_bytes(strategy_bytes)
    manifest_path = request.inbox / "inbox_manifest.json"
    manifest = json.loads(manifest_path.read_text())
    manifest["entries"][0]["source_strategy_sha256"] = sha256(strategy_bytes).hexdigest()
    manifest["entries"][0]["strategy_version_id"] = _canonical_strategy_hash(changed_strategy)
    manifest["v6_provenance"]["strategy_json_sha256"]["alpha.json"] = manifest["entries"][0]["strategy_version_id"]
    manifest_path.write_text(json.dumps(manifest))
    with pytest.raises(PerformanceV2ImportError, match="existing"):
        import_performance_v2(request)


def test_replace_requires_mapping_and_rolls_back_on_typed_mismatch(tmp_path: Path) -> None:
    request, _ = _request(tmp_path)
    import_performance_v2(request)
    target = performance_v2_database_path(request.config)
    with duckdb.connect(str(target), read_only=True) as connection:
        strategy_id = connection.execute("select strategy_id from strategies where strategy_name = 'alpha'").fetchone()[0]
        old_result = connection.execute("select current_result_id from strategies where strategy_id = ?", [strategy_id]).fetchone()[0]

    with pytest.raises(PerformanceV2ImportError, match="mapping"):
        import_performance_v2(PerformanceV2ImportRequest(request.inbox, request.report_root, request.config, mode="REPLACE"))

    strategy_path = request.inbox / "strategies" / "alpha.json"
    changed = json.loads(strategy_path.read_text())
    changed["mrs3"]["ma_close_long"]["len"] = 999  # type: ignore[index]
    changed_bytes = json.dumps(changed, separators=(",", ":")).encode()
    strategy_path.write_bytes(changed_bytes)
    manifest_path = request.inbox / "inbox_manifest.json"
    manifest = json.loads(manifest_path.read_text())
    manifest["entries"][0]["source_strategy_sha256"] = sha256(changed_bytes).hexdigest()
    manifest["entries"][0]["strategy_version_id"] = _canonical_strategy_hash(changed)
    manifest["v6_provenance"]["strategy_json_sha256"]["alpha.json"] = manifest["entries"][0]["strategy_version_id"]
    manifest_path.write_text(json.dumps(manifest))
    with pytest.raises(PerformanceV2ImportError, match="typed"):
        import_performance_v2(
            PerformanceV2ImportRequest(
                request.inbox, request.report_root, request.config,
                mode="REPLACE", replacement_strategy_ids={"alpha": strategy_id},
                listing_dates_path=request.listing_dates_path,
            )
        )
    with duckdb.connect(str(target), read_only=True) as connection:
        assert connection.execute("select current_result_id from strategies where strategy_id = ?", [strategy_id]).fetchone() == (old_result,)


def test_replace_switches_current_result_and_replaces_only_scoped_children(tmp_path: Path) -> None:
    request, _ = _request(tmp_path)
    import_performance_v2(request)
    target = performance_v2_database_path(request.config)
    with duckdb.connect(str(target)) as connection:
        strategy_id = connection.execute("select strategy_id from strategies where strategy_name = 'alpha'").fetchone()[0]
        old_result = connection.execute("select current_result_id from strategies where strategy_id = ?", [strategy_id]).fetchone()[0]
        old_actions = connection.execute(
            "select * from strategy_actions where result_id = ? order by action_index", [old_result]
        ).fetchall()
        old_equity = connection.execute(
            "select * from strategy_equity where result_id = ? order by sample_index", [old_result]
        ).fetchall()
        window_start = datetime(2026, 1, 1, tzinfo=timezone.utc)
        window_end = datetime(2026, 1, 2, tzinfo=timezone.utc)
        connection.execute(
            """insert into window_metrics (
                result_id, requested_start_utc, requested_end_utc, metrics_version,
                effective_start_utc, effective_end_utc, availability_status, unavailable_reason,
                growth_factor, return_pct, daily_log_return, daily_growth_pct, max_drawdown_pct,
                return_dd_ratio, fees_pct, profit_factor, trade_count, win_rate_pct,
                holding_seconds, time_in_market_pct, calculated_at_utc
            ) values (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            [
                old_result, window_start, window_end, "test", window_start, window_end, "AVAILABLE", None,
                Decimal("1"), Decimal("0"), Decimal("0"), Decimal("0"), Decimal("0"), Decimal("0"),
                Decimal("0"), Decimal("1"), 1, Decimal("100"), Decimal("0"), Decimal("0"), window_end,
            ],
        )
    changed = FIXTURE.read_bytes().replace(
        b"2026-01-01 - 2026-01-09", b"2026-01-01 - 2026-01-10"
    ).replace(b"1009.9", b"1019.9")
    _rewrite_report(request, changed)

    result = import_performance_v2(
        PerformanceV2ImportRequest(
            request.inbox, request.report_root, request.config,
            mode="REPLACE", replacement_strategy_ids={"alpha": strategy_id},
            listing_dates_path=request.listing_dates_path,
        )
    )

    assert result.imported_count == 1
    with duckdb.connect(str(target), read_only=True) as connection:
        new_result = connection.execute("select current_result_id from strategies where strategy_id = ?", [strategy_id]).fetchone()[0]
        assert new_result == old_result
        assert not hasattr(import_module, "_prepare_replace_children")
        assert connection.execute("select count(*) from strategy_results where strategy_id = ?", [strategy_id]).fetchone() == (1,)
        assert connection.execute("select final_balance from strategy_results where result_id = ?", [new_result]).fetchone() == (Decimal("1019.9"),)
        assert len(connection.execute("select * from strategy_actions where result_id = ?", [old_result]).fetchall()) == len(old_actions)
        assert len(connection.execute("select * from strategy_equity where result_id = ?", [old_result]).fetchall()) == len(old_equity)
        assert connection.execute("select count(*) from window_metrics where result_id = ?", [old_result]).fetchone() == (0,)
        assert connection.execute("select report_end_utc from strategy_results where result_id = ?", [old_result]).fetchone()[0].date().isoformat() == "2026-01-10"
    assert (request.inbox / "strategies" / "alpha.json").read_bytes()
    assert (request.inbox / "inbox_manifest.json").read_bytes()


def test_replace_rollback_restores_old_result_after_publish_failure(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    request, _ = _request(tmp_path)
    import_performance_v2(request)
    target = performance_v2_database_path(request.config)
    with duckdb.connect(str(target)) as connection:
        strategy_id, old_result = connection.execute(
            "select strategy_id, current_result_id from strategies where strategy_name = 'alpha'"
        ).fetchone()
        old_actions = connection.execute("select count(*) from strategy_actions where result_id = ?", [old_result]).fetchone()[0]
        connection.execute(
            "insert into strategy_tags values (?, 'RETEST', 'RETEST_WORKFLOW', 'test', now())",
            [strategy_id],
        )
    _rewrite_report(request, FIXTURE.read_bytes().replace(b"1009.9", b"1019.9"))
    original_inbox = {path.relative_to(request.inbox): path.read_bytes() for path in request.inbox.rglob("*") if path.is_file()}
    import mrs3.performance_v2_import as import_module
    append_rows = import_module._append_rows

    def fail_after_new_result(connection, table, columns, rows):
        if table == "strategy_equity":
            raise RuntimeError("injected child failure")
        append_rows(connection, table, columns, rows)

    monkeypatch.setattr(import_module, "_append_rows", fail_after_new_result)

    with pytest.raises(PerformanceV2ImportError, match="transaction failed|injected"):
        import_performance_v2(
            PerformanceV2ImportRequest(
                request.inbox,
                request.report_root,
                request.config,
                mode="REPLACE",
                replacement_strategy_ids={"alpha": strategy_id},
                clear_retest_on_success=True,
                listing_dates_path=request.listing_dates_path,
            )
        )
    with duckdb.connect(str(target), read_only=True) as connection:
        assert connection.execute("select current_result_id from strategies where strategy_id = ?", [strategy_id]).fetchone() == (old_result,)
        assert connection.execute("select count(*) from strategy_actions where result_id = ?", [old_result]).fetchone() == (old_actions,)
        assert connection.execute(
            "select count(*) from strategy_tags where strategy_id = ? and tag = 'RETEST'", [strategy_id]
        ).fetchone() == (1,)
    assert list(request.config.database_root.glob("performance_v2_failures_*.csv"))
    assert {path.relative_to(request.inbox): path.read_bytes() for path in request.inbox.rglob("*") if path.is_file()} == original_inbox


def test_retest_replace_rejects_a_shorter_effective_period_without_mutation(tmp_path: Path) -> None:
    request, _ = _request(tmp_path)
    assert import_performance_v2(request).imported_count == 1
    target = performance_v2_database_path(request.config)
    with duckdb.connect(str(target), read_only=True) as connection:
        strategy_id, result_id = connection.execute(
            "select strategy_id, current_result_id from strategies where strategy_name = 'alpha'"
        ).fetchone()
        actions = connection.execute(
            "select count(*) from strategy_actions where result_id = ?", [result_id]
        ).fetchone()[0]
        equity = connection.execute(
            "select count(*) from strategy_equity where result_id = ?", [result_id]
        ).fetchone()[0]
    shorter = FIXTURE.read_bytes().replace(b"2026-01-01 - 2026-01-09", b"2026-01-01 - 2026-01-04")
    assert parse_current_performance_v2_html(shorter, request.config).metrics["Report range"] == "2026-01-01 - 2026-01-04"
    _rewrite_report(request, shorter)

    with pytest.raises(PerformanceV2ImportError, match="shorter effective period"):
        import_performance_v2(
            PerformanceV2ImportRequest(
                request.inbox, request.report_root, request.config,
                mode="REPLACE", replacement_strategy_ids={"alpha": strategy_id},
                listing_dates_path=request.listing_dates_path,
            )
        )
    with duckdb.connect(str(target), read_only=True) as connection:
        assert connection.execute(
            "select current_result_id from strategies where strategy_id = ?", [strategy_id]
        ).fetchone() == (result_id,)
        assert connection.execute("select count(*) from strategy_actions where result_id = ?", [result_id]).fetchone() == (actions,)
        assert connection.execute("select count(*) from strategy_equity where result_id = ?", [result_id]).fetchone() == (equity,)


def test_lock_conflict_does_not_read_inbox_or_create_staging(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    request, _ = _request(tmp_path, initialize_db=False)
    target = _db(request)
    held = duckdb.connect(str(target))
    import mrs3.performance_v2_import as import_module
    monkeypatch.setattr(import_module.duckdb, "connect", lambda *args, **kwargs: (_ for _ in ()).throw(duckdb.IOException("lock")))
    monkeypatch.setattr(import_module, "read_performance_v2_inbox", lambda *args, **kwargs: pytest.fail("inbox was read while locked"))
    try:
        with pytest.raises(PerformanceV2LockedError):
            import_performance_v2(request)
    finally:
        held.close()
    assert not (request.config.database_root / ".staging").exists()


def test_writer_lock_fails_closed_before_opening_database(tmp_path: Path) -> None:
    request, _ = _request(tmp_path)

    with PerformanceV2WriterLock(request.config.database_root):
        with pytest.raises(PerformanceV2LockedError):
            import_performance_v2(request)


def test_writer_lock_does_not_create_missing_database_root(tmp_path: Path) -> None:
    missing_root = tmp_path / "missing-v2"

    with pytest.raises(PerformanceV2StoreError, match="root does not exist"):
        with PerformanceV2WriterLock(missing_root):
            pass

    assert not missing_root.exists()


def test_missing_target_fails_before_connect_and_does_not_create_root(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    request, _ = _request(tmp_path, initialize_db=False)
    import mrs3.performance_v2_import as import_module
    monkeypatch.setattr(import_module.duckdb, "connect", lambda *args, **kwargs: pytest.fail("connect called"))

    with pytest.raises(PerformanceV2ImportError, match="target does not exist"):
        import_performance_v2(request)

    target = performance_v2_database_path(request.config)
    assert not target.exists()
    assert not request.config.database_root.exists()
    assert not (request.config.database_root / "import_audit.v2.json").exists()


def test_empty_target_fails_schema_gate_without_staging_or_audit(tmp_path: Path) -> None:
    request, _ = _request(tmp_path, initialize_db=False)
    target = performance_v2_database_path(request.config)
    target.parent.mkdir(parents=True)
    target.touch()

    with pytest.raises(PerformanceV2ImportError, match="schema version 2"):
        import_performance_v2(request)

    assert target.exists()
    assert not (request.config.database_root / ".staging").exists()
    assert not (request.config.database_root / "import_audit.v2.json").exists()


def test_warmup_does_not_publish_an_open_only_report(tmp_path: Path) -> None:
    request, _ = _request(tmp_path)
    prepared = read_performance_v2_inbox(request.inbox, request.report_root)
    parsed = parse_current_performance_v2_html(FIXTURE.read_bytes(), request.config)
    parsed = replace(
        parsed,
        actions=(replace(parsed.actions[0], timestamp_utc=datetime(2026, 1, 7, tzinfo=timezone.utc)),),
    )

    result, failure = import_module._warmup_report(
        prepared.entries[0], parsed, {"ONUSDT": datetime(2026, 1, 1, tzinfo=timezone.utc)}
    )

    assert result is None
    assert failure is not None and failure["reason"] == "NO_EFFECTIVE_TRADE"


def test_no_warmup_drops_open_at_end_lifecycle_and_counts_it(tmp_path: Path) -> None:
    request, _ = _request(tmp_path)
    prepared = read_performance_v2_inbox(request.inbox, request.report_root)
    parsed = parse_current_performance_v2_html(FIXTURE.read_bytes(), request.config)
    open_action = replace(
        parsed.actions[0], action_index=2, timestamp_utc=datetime(2026, 1, 4, tzinfo=timezone.utc),
    )
    parsed = replace(parsed, actions=parsed.actions + (open_action,))

    result, failure = import_module._warmup_report(
        prepared.entries[0], parsed, {"ONUSDT": datetime(2025, 12, 25, tzinfo=timezone.utc)}
    )

    assert failure is None and result is not None
    assert len(result.actions) == 2
    assert result.metrics["Total Trades"] == "1"
    assert result.excluded_trade_count == 1


def test_no_warmup_open_only_report_is_not_published(tmp_path: Path) -> None:
    request, _ = _request(tmp_path)
    prepared = read_performance_v2_inbox(request.inbox, request.report_root)
    parsed = parse_current_performance_v2_html(FIXTURE.read_bytes(), request.config)
    parsed = replace(parsed, actions=(replace(parsed.actions[0], timestamp_utc=datetime(2026, 1, 4, tzinfo=timezone.utc)),))

    result, failure = import_module._warmup_report(
        prepared.entries[0], parsed, {"ONUSDT": datetime(2025, 12, 25, tzinfo=timezone.utc)}
    )

    assert result is None
    assert failure is not None and failure["reason"] == "NO_EFFECTIVE_TRADE"
    assert failure["excluded_trade_count"] == 1


def test_warmup_excludes_crossing_trade_pnl_and_fees(tmp_path: Path) -> None:
    request, _ = _request(tmp_path)
    prepared = read_performance_v2_inbox(request.inbox, request.report_root)
    parsed = parse_current_performance_v2_html(FIXTURE.read_bytes(), request.config)
    actions = (
        replace(parsed.actions[0], timestamp_utc=datetime(2026, 1, 2, tzinfo=timezone.utc), fee=Decimal("0.5"), balance=Decimal("999.5")),
        replace(parsed.actions[1], timestamp_utc=datetime(2026, 1, 7, tzinfo=timezone.utc), fee=Decimal("0.5"), pnl=Decimal("10"), balance=Decimal("1009")),
        replace(parsed.actions[0], timestamp_utc=datetime(2026, 1, 8, tzinfo=timezone.utc), fee=Decimal("1"), balance=Decimal("1008")),
        replace(parsed.actions[1], timestamp_utc=datetime(2026, 1, 9, tzinfo=timezone.utc), fee=Decimal("1"), pnl=Decimal("4"), balance=Decimal("1011")),
    )
    metrics = dict(parsed.metrics)
    metrics["Report range"] = "2026-01-01 - 2026-01-10"
    parsed = replace(parsed, metrics=metrics, actions=actions, wallet_series=(), equity_series=())

    result, failure = import_module._warmup_report(
        prepared.entries[0], parsed, {"ONUSDT": datetime(2026, 1, 3, tzinfo=timezone.utc)}
    )

    assert failure is None and result is not None
    assert len(result.actions) == 2
    assert result.wallet_series == () and result.equity_series == ()
    assert result.inventory.wallet_sample_count == result.inventory.equity_sample_count == 0
    assert result.metrics["Initial balance"] == "1000.0"
    assert result.metrics["Max Drawdown"] == "N/A"
    assert result.metrics["Total Trades"] == "1"
    assert result.metrics["Total PnL"] == "2"
    assert result.metrics["Total fees"] == "2"
    assert result.reported_start_utc == datetime(2026, 1, 1, tzinfo=timezone.utc)
    assert result.reported_end_utc == datetime(2026, 1, 10, tzinfo=timezone.utc)
    assert result.listing_date_utc == datetime(2026, 1, 3, tzinfo=timezone.utc)
    assert result.effective_start_utc == datetime(2026, 1, 8, tzinfo=timezone.utc)
    assert result.effective_end_utc == result.reported_end_utc
    assert result.warmup_hours == 120
    assert result.excluded_trade_count == 1
    assert result.exclusion_reason is None


def test_warmup_drawdown_peak_starts_at_baseline_for_first_sample_below_it(tmp_path: Path) -> None:
    request, _ = _request(tmp_path)
    prepared = read_performance_v2_inbox(request.inbox, request.report_root)
    parsed = parse_current_performance_v2_html(FIXTURE.read_bytes(), request.config)
    parsed = replace(
        parsed,
        actions=(
            replace(parsed.actions[0], timestamp_utc=datetime(2026, 1, 8, tzinfo=timezone.utc), fee=Decimal("1"), balance=Decimal("999")),
            replace(parsed.actions[1], timestamp_utc=datetime(2026, 1, 9, tzinfo=timezone.utc), pnl=Decimal("0"), fee=Decimal("0"), balance=Decimal("999")),
        ),
        wallet_series=((datetime(2026, 1, 8, tzinfo=timezone.utc), Decimal("999")),),
        equity_series=(),
    )

    result, failure = import_module._warmup_report(
        prepared.entries[0], parsed, {"ONUSDT": datetime(2026, 1, 3, tzinfo=timezone.utc)}
    )

    assert failure is None and result is not None
    assert result.wallet_series == ((datetime(2026, 1, 8, tzinfo=timezone.utc), Decimal("999")),)
    assert result.inventory.wallet_sample_count == 1
    assert result.metrics["Max Drawdown"] == "1"


def test_warmup_pins_signed_gross_loss_for_a_losing_trade(tmp_path: Path) -> None:
    request, _ = _request(tmp_path)
    prepared = read_performance_v2_inbox(request.inbox, request.report_root)
    parsed = parse_current_performance_v2_html(FIXTURE.read_bytes(), request.config)
    parsed = replace(
        parsed,
        actions=(
            replace(parsed.actions[0], timestamp_utc=datetime(2026, 1, 8, tzinfo=timezone.utc), fee=Decimal("0"), pnl=Decimal("0"), balance=Decimal("1000")),
            replace(parsed.actions[1], timestamp_utc=datetime(2026, 1, 9, tzinfo=timezone.utc), fee=Decimal("0"), pnl=Decimal("-2"), balance=Decimal("998")),
        ),
        wallet_series=(),
        equity_series=(),
    )

    result, failure = import_module._warmup_report(
        prepared.entries[0], parsed, {"ONUSDT": datetime(2026, 1, 3, tzinfo=timezone.utc)}
    )

    assert failure is None and result is not None
    # Gross loss is a positive magnitude, matching the window calculator.
    assert result.metrics["Gross loss"] == "2"


def test_result_values_persist_full_precision_effective_provenance(tmp_path: Path) -> None:
    request, _ = _request(tmp_path)
    prepared = read_performance_v2_inbox(request.inbox, request.report_root)
    parsed = parse_current_performance_v2_html(FIXTURE.read_bytes(), request.config)
    effective_start = datetime(2026, 1, 1, 12, tzinfo=timezone.utc)
    effective_end = datetime(2026, 1, 9, 18, 30, tzinfo=timezone.utc)
    parsed = replace(parsed, effective_start_utc=effective_start, effective_end_utc=effective_end)

    values = import_module._result_values(
        prepared.entries[0], parsed, {"TakerFee": "0.0004"}, datetime(2026, 1, 10, tzinfo=timezone.utc)
    )

    assert values[0:2] == [effective_start, effective_end]


def test_warmup_does_not_publish_when_the_only_trade_crosses_warmup(tmp_path: Path) -> None:
    request, _ = _request(tmp_path)
    prepared = read_performance_v2_inbox(request.inbox, request.report_root)
    parsed = parse_current_performance_v2_html(FIXTURE.read_bytes(), request.config)
    parsed = replace(
        parsed,
        actions=(
            replace(parsed.actions[0], timestamp_utc=datetime(2026, 1, 2, tzinfo=timezone.utc)),
            replace(parsed.actions[1], timestamp_utc=datetime(2026, 1, 8, tzinfo=timezone.utc)),
        ),
        wallet_series=(),
        equity_series=(),
    )

    result, failure = import_module._warmup_report(
        prepared.entries[0], parsed, {"ONUSDT": datetime(2026, 1, 3, tzinfo=timezone.utc)}
    )

    assert result is None
    assert failure is not None
    assert failure["reason"] == "NO_EFFECTIVE_TRADE"
    assert failure["excluded_trade_count"] == 1


def test_warmup_rejects_unknown_action_without_dropping_it(tmp_path: Path) -> None:
    request, _ = _request(tmp_path)
    prepared = read_performance_v2_inbox(request.inbox, request.report_root)
    parsed = parse_current_performance_v2_html(FIXTURE.read_bytes(), request.config)
    parsed = replace(parsed, actions=(replace(parsed.actions[0], action="mystery"), parsed.actions[1]))

    result, failure = import_module._warmup_report(
        prepared.entries[0], parsed, {"ONUSDT": datetime(2025, 12, 25, tzinfo=timezone.utc)}
    )

    assert result is None
    assert failure is not None and failure["reason"] == "UNKNOWN_ACTION"


def test_import_persists_warmup_provenance(tmp_path: Path) -> None:
    request, _ = _request(tmp_path)
    dates = tmp_path / "Input" / "dates.xlsx"
    dates.parent.mkdir(exist_ok=True)
    workbook = Workbook()
    workbook.active.append(["ONUSDT", datetime(2025, 12, 25)])
    workbook.save(dates)
    request = PerformanceV2ImportRequest(
        request.inbox, request.report_root, request.config, listing_dates_path=Path("Input/dates.xlsx")
    )

    assert import_performance_v2(request).imported_count == 1
    target = performance_v2_database_path(request.config)
    with duckdb.connect(str(target), read_only=True) as connection:
        row = connection.execute(
            """select report_start_utc, report_end_utc, reported_start_utc, reported_end_utc,
                      listing_date_utc, listing_date_raw, listing_date_source,
                      effective_start_utc, effective_end_utc, warmup_hours,
                      excluded_trade_count, exclusion_reason
                 from strategy_results"""
        ).fetchone()
    assert row[0].date().isoformat() == "2026-01-01"
    assert row[1].date().isoformat() == "2026-01-09"
    assert row[2].date().isoformat() == "2026-01-01"
    assert row[3].date().isoformat() == "2026-01-09"
    assert row[4].date().isoformat() == "2025-12-25"
    assert row[5] and row[6] == "configured_listing_dates_path"
    assert row[7].date().isoformat() == "2026-01-01"
    assert row[8].date().isoformat() == "2026-01-09"
    assert row[9:] == (120, 0, None)
def test_warmup_normalizes_listing_timezone_and_keeps_inclusive_report_end(tmp_path: Path) -> None:
    request, _ = _request(tmp_path)
    prepared = read_performance_v2_inbox(request.inbox, request.report_root)
    parsed = parse_current_performance_v2_html(FIXTURE.read_bytes(), request.config)
    metrics = dict(parsed.metrics)
    metrics["Report range"] = "2026-01-01 - 2026-01-10"
    parsed = replace(
        parsed,
        metrics=metrics,
        actions=(
            replace(parsed.actions[0], timestamp_utc=datetime(2026, 1, 6, 10, tzinfo=timezone.utc)),
            replace(parsed.actions[1], timestamp_utc=datetime(2026, 1, 10, tzinfo=timezone.utc)),
        ),
        wallet_series=(),
        equity_series=(),
    )

    result, failure = import_module._warmup_report(
        prepared.entries[0],
        parsed,
        {"ONUSDT": datetime(2026, 1, 1, 12, tzinfo=timezone(timedelta(hours=2)))},
    )

    assert failure is None and result is not None
    assert len(result.actions) == 2
    assert result.actions[-1].timestamp_utc == datetime(2026, 1, 10, tzinfo=timezone.utc)


def test_warmup_drops_a_trade_closed_after_inclusive_report_end(tmp_path: Path) -> None:
    request, _ = _request(tmp_path)
    prepared = read_performance_v2_inbox(request.inbox, request.report_root)
    parsed = parse_current_performance_v2_html(FIXTURE.read_bytes(), request.config)
    metrics = dict(parsed.metrics)
    metrics["Report range"] = "2026-01-01 - 2026-01-10"
    parsed = replace(
        parsed,
        metrics=metrics,
        actions=(
            replace(parsed.actions[0], timestamp_utc=datetime(2026, 1, 9, 10, tzinfo=timezone.utc)),
            replace(parsed.actions[1], timestamp_utc=datetime(2026, 1, 10, 0, 1, tzinfo=timezone.utc)),
        ),
        wallet_series=(),
        equity_series=(),
    )

    result, failure = import_module._warmup_report(
        prepared.entries[0], parsed, {"ONUSDT": datetime(2026, 1, 4, tzinfo=timezone.utc)}
    )

    assert result is None
    assert failure is not None and failure["reason"] == "NO_EFFECTIVE_TRADE"


@pytest.mark.parametrize(
    ("field", "reason"),
    [("actions", "ACTIONS_OUT_OF_ORDER"), ("wallet_series", "WALLET_OUT_OF_ORDER"), ("equity_series", "EQUITY_OUT_OF_ORDER")],
)
def test_warmup_rejects_out_of_order_source_rows(tmp_path: Path, field: str, reason: str) -> None:
    request, _ = _request(tmp_path)
    prepared = read_performance_v2_inbox(request.inbox, request.report_root)
    parsed = parse_current_performance_v2_html(FIXTURE.read_bytes(), request.config)
    values = {
        "actions": tuple(reversed(parsed.actions)),
        "wallet_series": tuple(reversed(parsed.wallet_series)),
        "equity_series": tuple(reversed(parsed.equity_series)),
    }
    parsed = replace(parsed, **{field: values[field]})

    result, failure = import_module._warmup_report(
        prepared.entries[0], parsed, {"ONUSDT": datetime(2025, 12, 25, tzinfo=timezone.utc)}
    )

    assert result is None
    assert failure is not None and failure["reason"] == reason


def test_warmup_rejects_orphan_increased_action(tmp_path: Path) -> None:
    request, _ = _request(tmp_path)
    prepared = read_performance_v2_inbox(request.inbox, request.report_root)
    parsed = parse_current_performance_v2_html(FIXTURE.read_bytes(), request.config)
    parsed = replace(parsed, actions=(replace(parsed.actions[0], action="increased"), parsed.actions[1]))

    result, failure = import_module._warmup_report(
        prepared.entries[0], parsed, {"ONUSDT": datetime(2025, 12, 25, tzinfo=timezone.utc)}
    )

    assert result is None
    assert failure is not None and failure["reason"] == "INVALID_ACTION_STATE"


def test_single_mode_without_listing_dates_fails_closed(tmp_path: Path) -> None:
    request, _ = _request(tmp_path)
    request = PerformanceV2ImportRequest(request.inbox, request.report_root, request.config)
    manifest_path = request.inbox / "inbox_manifest.json"
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    manifest.update({"run_mode": "SINGLE_MODE", "test_start": "2026-01-01", "test_end": "2026-01-09"})
    manifest_path.write_text(json.dumps(manifest), encoding="utf-8")
    prepared = read_performance_v2_inbox(request.inbox, request.report_root)
    parsed = parse_current_performance_v2_html(FIXTURE.read_bytes(), request.config)

    filtered, failures = import_module._prepare_listing_ranges(request, prepared, (parsed,))

    assert filtered == (None,)
    assert failures and failures[0]["reason"] == "LISTING_MISSING"


def test_all_invalid_reports_fail_without_empty_in_clause(tmp_path: Path) -> None:
    request, _ = _request(tmp_path, names=("alpha", "beta"))
    manifest_path = request.inbox / "inbox_manifest.json"
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    for index, name in enumerate(("alpha", "beta")):
        report_path = request.report_root / f"{name}.html"
        report_path.write_text("<html>invalid</html>", encoding="utf-8")
        manifest["entries"][index]["source_report_sha256"] = sha256(report_path.read_bytes()).hexdigest()
    manifest_path.write_text(json.dumps(manifest), encoding="utf-8")

    result = import_performance_v2(request)

    assert result.status == "FAILED"
    assert (result.imported_count, result.skipped_count, result.rejected_count) == (0, 0, 2)
    assert result.failure_count == 2
    with duckdb.connect(str(performance_v2_database_path(request.config)), read_only=True) as connection:
        assert connection.execute("select count(*) from strategies").fetchone() == (0,)
        for table in (
            "strategy_results",
            "strategy_orders",
            "strategy_actions",
            "strategy_equity",
            "window_metrics",
            "analysis_plateaus",
        ):
            assert connection.execute(f"select count(*) from {table}").fetchone() == (0,)
        assert connection.execute("select status from import_runs").fetchone() == ("FAILED",)
        assert connection.execute("select distinct status from import_files").fetchall() == [("REJECTED:INVALID_REPORT",)]
    failure_csv = result.failure_report_path
    assert failure_csv is not None
    failure_text = failure_csv.read_text(encoding="utf-8")
    assert failure_text.splitlines()[0].split(",")[:5] == [
        "reason", "strategy_name", "symbol", "import_id", "outcome"
    ]
    failure_rows = list(csv.DictReader(failure_text.splitlines()))
    assert all(row["outcome"] == "FAILED" and row["error"] for row in failure_rows)
    assert result.failure_report_xlsx_path is not None and result.failure_report_xlsx_path.is_file()


def test_failure_report_blanks_none_and_strips_control_characters(tmp_path: Path) -> None:
    config = PerformanceV2Config(tmp_path / "v2")
    csv_path, xlsx_path = import_module._write_failure_reports(
        config,
        [{"reason": "INVALID_REPORT", "strategy_name": "alpha", "symbol": "ONUSDT", "error": None},
         {"reason": "LISTING_MISSING", "strategy_name": "beta", "symbol": "ONUSDT", "error": "bad\npath\x00"}],
        import_id="test", status="FAILED",
    )

    rows = list(csv.DictReader(csv_path.read_text(encoding="utf-8").splitlines()))
    assert rows[0]["error"] == ""
    assert rows[1]["error"] == "bad path"
    assert xlsx_path is not None and xlsx_path.is_file()


def test_clear_retest_on_success_uses_replaced_strategy_ids_only(tmp_path: Path) -> None:
    request, _ = _request(tmp_path, names=("alpha", "beta"))
    target = performance_v2_database_path(request.config)
    with duckdb.connect(str(target)) as connection:
        # Seed an unrelated strategy so strategy and result IDs are visibly
        # out of phase before the replacement transaction.
        connection.execute(
            """insert into strategies (
                strategy_name, symbol, side, timeframe, close_ma_len, order_count,
                analysis_run_id, candidate_identity, lifecycle_status, current_result_id,
                created_at_utc, updated_at_utc
             ) values ('seed', 'SEEDUSDT', 'LONG', '1h', 3, 1, 'seed-run', 'seed-candidate', 'DISCARDED', null, now(), now())"""
        )
    import_performance_v2(request)
    with duckdb.connect(str(target)) as connection:
        ids = dict(
            connection.execute(
                "select strategy_name, strategy_id from strategies where strategy_name in ('alpha', 'beta')"
            ).fetchall()
        )
        connection.executemany(
            "insert into strategy_tags values (?, 'RETEST', 'RETEST_WORKFLOW', 'test', now())",
            [(ids["alpha"],), (ids["beta"],)],
        )
    prepared = read_performance_v2_inbox(request.inbox, request.report_root)
    parsed = parse_current_performance_v2_html(FIXTURE.read_bytes(), request.config)
    parsed = replace(
        parsed,
        listing_date_utc=datetime(2025, 12, 25, tzinfo=timezone.utc),
        reported_start_utc=datetime(2026, 1, 1, tzinfo=timezone.utc),
        reported_end_utc=datetime(2026, 1, 9, tzinfo=timezone.utc),
        effective_start_utc=datetime(2026, 1, 1, tzinfo=timezone.utc),
        effective_end_utc=datetime(2026, 1, 9, tzinfo=timezone.utc),
        warmup_hours=120,
    )
    replacement = PerformanceV2ImportRequest(
        request.inbox,
        request.report_root,
        request.config,
        mode="REPLACE",
        replacement_strategy_ids=ids,
        clear_retest_on_success=True,
        listing_dates_path=request.listing_dates_path,
    )
    with duckdb.connect(str(target)) as connection:
        import_module._publish(connection, replacement, prepared, (parsed, None), "selective-clear")

    with duckdb.connect(str(target), read_only=True) as connection:
        pairs = connection.execute("select strategy_id, current_result_id from strategies").fetchall()
        assert all(strategy_id != result_id for strategy_id, result_id in pairs)
        assert connection.execute(
            "select count(*) from strategy_tags where strategy_id = ? and tag = 'RETEST'", [ids["alpha"]]
        ).fetchone() == (0,)
        assert connection.execute(
            "select count(*) from strategy_tags where strategy_id = ? and tag = 'RETEST'", [ids["beta"]]
        ).fetchone() == (1,)


def test_import_request_rejects_absolute_listing_dates_path(tmp_path: Path) -> None:
    request, _ = _request(tmp_path)

    with pytest.raises(ValueError, match="relative"):
        PerformanceV2ImportRequest(
            request.inbox,
            request.report_root,
            request.config,
            listing_dates_path=tmp_path / "Input" / "dates.xlsx",
        )


def test_relative_input_listing_dates_path_is_resolved_from_inbox_parent(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    request, _ = _request(tmp_path)
    dates_path = tmp_path / "Input" / "dates.xlsx"
    dates_path.parent.mkdir(exist_ok=True)
    dates_path.touch()
    seen: list[Path] = []
    monkeypatch.setattr(
        import_module,
        "load_listing_dates",
        lambda path: (seen.append(path), {"ONUSDT": datetime(2025, 12, 1, tzinfo=timezone.utc)})[1],
    )
    request = PerformanceV2ImportRequest(
        request.inbox,
        request.report_root,
        request.config,
        listing_dates_path=Path("Input/dates.xlsx"),
    )

    result = import_performance_v2(request)

    assert result.status == "COMMITTED"
    assert result.imported_count == 1
    assert seen == [dates_path]


def test_listing_dates_path_reports_missing_file_distinctly(tmp_path: Path) -> None:
    request, _ = _request(tmp_path)
    request = PerformanceV2ImportRequest(
        request.inbox,
        request.report_root,
        request.config,
        listing_dates_path=Path("missing/dates.xlsx"),
    )

    with pytest.raises(PerformanceV2ImportError, match="not found"):
        import_performance_v2(request)


def test_listing_dates_path_reports_non_regular_file_distinctly(tmp_path: Path) -> None:
    request, _ = _request(tmp_path)
    dates_path = tmp_path / "Input" / "dates.xlsx"
    dates_path.unlink()
    dates_path.mkdir()

    with pytest.raises(PerformanceV2ImportError, match="not a regular file"):
        import_performance_v2(request)


def test_valid_strategy_is_published_when_sibling_report_is_invalid(tmp_path: Path) -> None:
    request, _ = _request(tmp_path, names=("alpha", "beta"))
    invalid = request.report_root / "beta.html"
    invalid.write_text("<html>invalid</html>", encoding="utf-8")
    manifest_path = request.inbox / "inbox_manifest.json"
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    beta_entry = manifest["entries"][1]
    beta_entry["source_report_sha256"] = sha256(invalid.read_bytes()).hexdigest()
    manifest_path.write_text(json.dumps(manifest), encoding="utf-8")

    result = import_performance_v2(request)

    assert result.status == "COMMITTED"
    assert result.imported_count == 1
    assert result.skipped_count == 0
    assert result.rejected_count == 1
    assert result.failure_report_path is not None and result.failure_report_path.is_file()
    assert result.failure_report_xlsx_path is not None and result.failure_report_xlsx_path.is_file()
    with duckdb.connect(str(performance_v2_database_path(request.config)), read_only=True) as connection:
        assert connection.execute("select status from import_runs").fetchone() == ("COMMITTED",)
        assert connection.execute("select strategy_name from strategies order by strategy_name").fetchall() == [("alpha",)]


def test_misaligned_equity_series_rejects_only_that_report(tmp_path: Path) -> None:
    request, _ = _request(tmp_path, names=("alpha", "beta"))
    report_path = request.report_root / "beta.html"
    broken = report_path.read_bytes().replace(
        b'const equitySeries = [[1767225600000,"1000"],[1767229200000,"999.95"],[1767402000000,"1009.9"]];',
        b'const equitySeries = [[1767225600000,"1000"],[1767229200000,"999.95"]];',
    )
    report_path.write_bytes(broken)
    manifest_path = request.inbox / "inbox_manifest.json"
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    manifest["entries"][1]["source_report_sha256"] = sha256(broken).hexdigest()
    manifest_path.write_text(json.dumps(manifest), encoding="utf-8")

    result = import_performance_v2(request)

    assert (result.imported_count, result.rejected_count) == (1, 1)
    assert result.failure_report_path is not None
    assert "wallet/equity sample counts must match" in result.failure_report_path.read_text(encoding="utf-8")


def test_invalid_schema_fails_before_staging_or_audit(tmp_path: Path) -> None:
    request, _ = _request(tmp_path, initialize_db=False)
    target = performance_v2_database_path(request.config)
    target.parent.mkdir(parents=True)
    with duckdb.connect(str(target)) as connection:
        connection.execute("create table not_v2 (value integer)")

    with pytest.raises(PerformanceV2ImportError, match="schema version 2"):
        import_performance_v2(request)

    assert not (request.config.database_root / ".staging").exists()
    assert not (request.config.database_root / "import_audit.v2.json").exists()


def test_absent_target_is_not_reported_as_lock_conflict(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    request, _ = _request(tmp_path, initialize_db=False)
    import mrs3.performance_v2_import as import_module
    monkeypatch.setattr(
        import_module.duckdb,
        "connect",
        lambda *args, **kwargs: (_ for _ in ()).throw(duckdb.IOException("lock")),
    )

    with pytest.raises(PerformanceV2ImportError, match="target does not exist") as error:
        import_performance_v2(request)

    assert not isinstance(error.value, PerformanceV2LockedError)
    assert not request.config.database_root.exists()


def test_add_reuses_active_strategy_for_a_different_name(tmp_path: Path) -> None:
    first, _ = _request(tmp_path)
    assert import_performance_v2(first).imported_count == 1

    second, _ = _request(tmp_path / "second", names=("beta",))
    second = PerformanceV2ImportRequest(
        second.inbox, second.report_root, first.config,
        listing_dates_path=second.listing_dates_path,
    )
    result = import_performance_v2(second)

    assert (result.imported_count, result.skipped_count) == (0, 1)
    with duckdb.connect(str(performance_v2_database_path(first.config)), read_only=True) as connection:
        assert connection.execute("select count(*) from strategies").fetchone() == (1,)
        assert connection.execute("select strategy_name from strategies").fetchone() == ("alpha",)


def test_add_scopes_invalid_active_lookup_to_incoming_typed_base(tmp_path: Path) -> None:
    first, _ = _request(tmp_path)
    assert import_performance_v2(first).imported_count == 1
    target = performance_v2_database_path(first.config)
    with duckdb.connect(str(target)) as connection:
        connection.execute(
            """insert into strategies (
                strategy_name, symbol, side, timeframe, close_ma_len, order_count,
                analysis_run_id, candidate_identity, lifecycle_status, current_result_id,
                created_at_utc, updated_at_utc
             ) values ('unrelated-legacy', 'OTHERUSDT', 'LONG', '1h', 3, 1,
                       'legacy-run', 'legacy-candidate', 'ACTIVE', null, now(), now())"""
        )

    second, _ = _request(tmp_path / "second", names=("beta",))
    second = PerformanceV2ImportRequest(
        second.inbox, second.report_root, first.config,
        listing_dates_path=second.listing_dates_path,
    )
    result = import_performance_v2(second)

    assert (result.imported_count, result.skipped_count) == (0, 1)
    with duckdb.connect(str(target), read_only=True) as connection:
        assert connection.execute("select count(*) from strategies").fetchone() == (2,)


def test_lot_x_is_part_of_the_canonical_key(tmp_path: Path) -> None:
    first, _ = _request(tmp_path)
    second, _ = _request(tmp_path / "second", names=("beta",))
    strategy_path = second.inbox / "strategies" / "beta.json"
    strategy = json.loads(strategy_path.read_text(encoding="utf-8"))
    strategy["mrs3"]["ma_long"][0]["lot_x"] = "0.75"  # type: ignore[index]
    strategy_bytes = json.dumps(strategy, separators=(",", ":")).encode()
    strategy_path.write_bytes(strategy_bytes)
    manifest_path = second.inbox / "inbox_manifest.json"
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    digest = _canonical_strategy_hash(strategy)
    manifest["entries"][0]["source_strategy_sha256"] = sha256(strategy_bytes).hexdigest()
    manifest["entries"][0]["strategy_version_id"] = digest
    manifest["v6_provenance"]["strategy_json_sha256"]["beta.json"] = digest
    manifest_path.write_text(json.dumps(manifest), encoding="utf-8")
    second = PerformanceV2ImportRequest(
        second.inbox, second.report_root, first.config,
        listing_dates_path=second.listing_dates_path,
    )

    assert import_performance_v2(first).imported_count == 1
    assert import_performance_v2(second).imported_count == 1
    with duckdb.connect(str(performance_v2_database_path(first.config)), read_only=True) as connection:
        assert connection.execute("select count(*) from strategies").fetchone() == (2,)
        assert {row[0] for row in connection.execute("select lot_x from strategy_orders").fetchall()} == {
            Decimal("1.000000000000"), Decimal("0.750000000000")
        }


def test_lot_x_key_survives_decimal_round_trip_at_schema_precision(tmp_path: Path) -> None:
    request, _ = _request(tmp_path)
    strategy_path = request.inbox / "strategies" / "alpha.json"
    strategy = json.loads(strategy_path.read_text(encoding="utf-8"))
    strategy["mrs3"]["ma_long"][0]["lot_x"] = "123.456789012345"  # type: ignore[index]
    strategy_bytes = json.dumps(strategy, separators=(",", ":")).encode()
    strategy_path.write_bytes(strategy_bytes)
    manifest_path = request.inbox / "inbox_manifest.json"
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    digest = _canonical_strategy_hash(strategy)
    manifest["entries"][0]["source_strategy_sha256"] = sha256(strategy_bytes).hexdigest()
    manifest["entries"][0]["strategy_version_id"] = digest
    manifest["v6_provenance"]["strategy_json_sha256"]["alpha.json"] = digest
    manifest_path.write_text(json.dumps(manifest), encoding="utf-8")

    assert import_performance_v2(request).imported_count == 1
    prepared = read_performance_v2_inbox(request.inbox, request.report_root)
    with duckdb.connect(str(performance_v2_database_path(request.config)), read_only=True) as connection:
        existing, orders, _results = import_module._load_existing(
            connection,
            ("alpha",),
            typed_prefixes=(import_module._typed_key(prepared.entries[0])[:5],),
        )
    row = existing["alpha"]
    assert import_module._stored_typed_key(row, orders[int(row[1])]) == import_module._typed_key(prepared.entries[0])


def test_add_replaces_canonical_result_only_for_a_strict_interval_superset(tmp_path: Path) -> None:
    first, _ = _request(tmp_path)
    assert import_performance_v2(first).imported_count == 1
    target = performance_v2_database_path(first.config)
    with duckdb.connect(str(target), read_only=True) as connection:
        old_strategy_id, old_result_id, old_name = connection.execute(
            "select strategy_id, current_result_id, strategy_name from strategies"
        ).fetchone()
        old_orders = connection.execute(
            "select order_id, open_ma_len, open_multiplier, shift_bp, lot_x, plateau_id, base_point_trades "
            "from strategy_orders where strategy_id = ? order by order_id", [old_strategy_id]
        ).fetchall()
    second, _ = _request(tmp_path / "second", names=("beta",))
    report_path = second.report_root / "beta.html"
    report = report_path.read_bytes().replace(
        b"2026-01-01 - 2026-01-09", b"2025-12-20 - 2026-01-10"
    )
    report_path.write_bytes(report)
    manifest_path = second.inbox / "inbox_manifest.json"
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    manifest["entries"][0]["source_report_sha256"] = sha256(report).hexdigest()
    manifest_path.write_text(json.dumps(manifest), encoding="utf-8")
    second = PerformanceV2ImportRequest(
        second.inbox, second.report_root, first.config,
        listing_dates_path=second.listing_dates_path,
    )

    result = import_performance_v2(second)

    assert (result.imported_count, result.skipped_count) == (1, 0)
    with duckdb.connect(str(target), read_only=True) as connection:
        new_strategy_id, new_result_id, new_name = connection.execute(
            "select strategy_id, current_result_id, strategy_name from strategies"
        ).fetchone()
        assert (new_strategy_id, new_name) == (old_strategy_id, old_name)
        assert new_result_id == old_result_id
        assert connection.execute(
            "select order_id, open_ma_len, open_multiplier, shift_bp, lot_x, plateau_id, base_point_trades "
            "from strategy_orders where strategy_id = ? order by order_id", [new_strategy_id]
        ).fetchall() == old_orders
        assert connection.execute("select count(*) from strategies").fetchone() == (1,)
        assert connection.execute("select count(*) from strategy_results").fetchone() == (1,)
        assert connection.execute("select report_end_utc from strategy_results").fetchone()[0].date().isoformat() == "2026-01-10"


def test_add_superset_does_not_clear_a_retest_tag(tmp_path: Path) -> None:
    first, _ = _request(tmp_path)
    assert import_performance_v2(first).imported_count == 1
    target = performance_v2_database_path(first.config)
    with duckdb.connect(str(target)) as connection:
        strategy_id = connection.execute("select strategy_id from strategies").fetchone()[0]
        connection.execute(
            "insert into strategy_tags values (?, 'RETEST', 'RETEST_WORKFLOW', 'test', now())",
            [strategy_id],
        )
    second, _ = _request(tmp_path / "second", names=("beta",))
    report_path = second.report_root / "beta.html"
    report = report_path.read_bytes().replace(b"2026-01-01 - 2026-01-09", b"2025-12-20 - 2026-01-10")
    report_path.write_bytes(report)
    manifest_path = second.inbox / "inbox_manifest.json"
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    manifest["entries"][0]["source_report_sha256"] = sha256(report).hexdigest()
    manifest_path.write_text(json.dumps(manifest), encoding="utf-8")

    result = import_performance_v2(PerformanceV2ImportRequest(
        second.inbox, second.report_root, first.config,
        clear_retest_on_success=True, listing_dates_path=second.listing_dates_path,
    ))

    assert result.imported_count == 1
    with duckdb.connect(str(target), read_only=True) as connection:
        assert connection.execute(
            "select count(*) from strategy_tags where strategy_id = ? and tag = 'RETEST'", [strategy_id]
        ).fetchone() == (1,)


def test_typed_key_canonicalizes_order_slots_and_excludes_provenance(tmp_path: Path) -> None:
    request, _ = _request(tmp_path, orders=2)
    prepared = read_performance_v2_inbox(request.inbox, request.report_root)
    entry = prepared.entries[0]
    reversed_orders = tuple(
        replace(order, order_id=3 - order.order_id)
        for order in reversed(entry.identity.orders)
    )
    alternate = replace(
        entry,
        strategy_name="other-name",
        analysis_run_id="other-run",
        candidate_identity="other-candidate",
        identity=replace(entry.identity, strategy_name="other-name", orders=reversed_orders),
    )

    assert import_module._typed_key(entry) == import_module._typed_key(alternate)


def test_interval_relation_requires_a_strict_proper_superset() -> None:
    start = datetime(2026, 1, 1, tzinfo=timezone.utc)
    end = datetime(2026, 1, 10, tzinfo=timezone.utc)
    current = (start, end)

    assert import_module._interval_relation(current, current) == "EQUAL"
    assert import_module._interval_relation((start - timedelta(days=1), end), current) == "SUPERSET"
    assert import_module._interval_relation((start, end + timedelta(days=1)), current) == "SUPERSET"
    assert import_module._interval_relation((start + timedelta(hours=1), end), current) == "SKIP"
    assert import_module._interval_relation((start - timedelta(days=1), end - timedelta(hours=1)), current) == "SKIP"
    assert import_module._interval_relation((None, end), current) == "UNKNOWN"
    assert import_module._interval_relation(current, (start, None)) == "UNKNOWN"


def test_lot_quantization_is_shared_for_float_and_decimal_values() -> None:
    assert import_module._quantized_lot(0.87524499704) == Decimal("0.875244997040")
    assert import_module._quantized_lot(Decimal("0.875244997040000")) == Decimal("0.875244997040")
    assert import_module._quantized_lot(Decimal("-0.0000000000001")) == Decimal("0.000000000000")
    assert import_module._quantized_lot(Decimal("99999999999999999999999999.999999999999")) == Decimal(
        "99999999999999999999999999.999999999999"
    )


def test_comparison_interval_rejects_naive_listing_timestamp(tmp_path: Path) -> None:
    request, _ = _request(tmp_path)
    prepared = read_performance_v2_inbox(request.inbox, request.report_root)
    report = parse_current_performance_v2_html(FIXTURE.read_bytes(), request.config)
    naive = replace(report, listing_date_utc=datetime(2025, 12, 25))
    assert import_module._comparison_interval(naive) is None


def test_add_fails_closed_when_active_typed_configuration_is_unreadable(tmp_path: Path) -> None:
    first, _ = _request(tmp_path)
    assert import_performance_v2(first).imported_count == 1
    target = performance_v2_database_path(first.config)
    with duckdb.connect(str(target)) as connection:
        strategy_id = connection.execute("select strategy_id from strategies").fetchone()[0]
        connection.execute("delete from strategy_orders where strategy_id = ?", [strategy_id])

    second, _ = _request(tmp_path / "second", names=("beta",))
    second = PerformanceV2ImportRequest(
        second.inbox, second.report_root, first.config,
        listing_dates_path=second.listing_dates_path,
    )
    with pytest.raises(PerformanceV2ImportError, match="invalid typed configuration"):
        import_performance_v2(second)


def test_add_fails_closed_when_stored_multiplier_disagrees_with_shift(tmp_path: Path) -> None:
    first, _ = _request(tmp_path)
    assert import_performance_v2(first).imported_count == 1
    target = performance_v2_database_path(first.config)
    with duckdb.connect(str(target)) as connection:
        connection.execute("update strategy_orders set open_multiplier = 0.9")
    second, _ = _request(tmp_path / "second", names=("beta",))

    with pytest.raises(PerformanceV2ImportError, match="invalid typed configuration"):
        import_performance_v2(PerformanceV2ImportRequest(
            second.inbox, second.report_root, first.config,
            listing_dates_path=second.listing_dates_path,
        ))


def test_add_fails_closed_when_active_current_result_is_dangling(tmp_path: Path) -> None:
    first, _ = _request(tmp_path)
    assert import_performance_v2(first).imported_count == 1
    target = performance_v2_database_path(first.config)
    with duckdb.connect(str(target)) as connection:
        strategy_id = connection.execute("select strategy_id from strategies").fetchone()[0]
        connection.execute("update strategies set current_result_id = 999999 where strategy_id = ?", [strategy_id])

    second, _ = _request(tmp_path / "second", names=("beta",))
    second = PerformanceV2ImportRequest(
        second.inbox, second.report_root, first.config,
        listing_dates_path=second.listing_dates_path,
    )
    with pytest.raises(PerformanceV2ImportError, match="no current result"):
        import_performance_v2(second)


@pytest.mark.parametrize(
    "expected",
    ["not-a-mapping", {}, {"symbol": "ONUSDT"}, {"symbol": "ONUSDT", "orders": None}],
)
def test_replace_rejects_malformed_expected_identity(tmp_path: Path, expected: object) -> None:
    request, _ = _request(tmp_path)
    assert import_performance_v2(request).imported_count == 1
    target = performance_v2_database_path(request.config)
    with duckdb.connect(str(target), read_only=True) as connection:
        strategy_id, result_id = connection.execute(
            "select strategy_id, current_result_id from strategies where strategy_name = 'alpha'"
        ).fetchone()
    replacement = PerformanceV2ImportRequest(
        request.inbox, request.report_root, request.config,
        mode="REPLACE", replacement_strategy_ids={"alpha": strategy_id},
        expected_strategy_identities={"alpha": expected},
        listing_dates_path=request.listing_dates_path,
    )
    with pytest.raises(PerformanceV2ImportError, match="typed strategy mismatch"):
        import_performance_v2(replacement)
    with duckdb.connect(str(target), read_only=True) as connection:
        assert connection.execute(
            "select current_result_id from strategies where strategy_id = ?", [strategy_id]
        ).fetchone() == (result_id,)


def test_replace_rejects_complete_expected_identity_with_bad_order(tmp_path: Path) -> None:
    request, _ = _request(tmp_path)
    assert import_performance_v2(request).imported_count == 1
    prepared = read_performance_v2_inbox(request.inbox, request.report_root)
    entry = prepared.entries[0]
    target = performance_v2_database_path(request.config)
    with duckdb.connect(str(target), read_only=True) as connection:
        strategy_id, result_id = connection.execute(
            "select strategy_id, current_result_id from strategies where strategy_name = 'alpha'"
        ).fetchone()
    order = entry.identity.orders[0]
    expected = {
        "symbol": entry.identity.symbol,
        "side": entry.identity.side,
        "timeframe": entry.identity.timeframe,
        "close_ma_len": entry.identity.close_ma_len,
        "order_count": entry.identity.order_count,
        "orders": [{"open_ma_len": order.open_ma_len, "shift_bp": order.shift_bp + 1, "lot_x": str(order.lot_x)}],
    }
    replacement = PerformanceV2ImportRequest(
        request.inbox, request.report_root, request.config,
        mode="REPLACE", replacement_strategy_ids={"alpha": strategy_id},
        expected_strategy_identities={"alpha": expected},
        listing_dates_path=request.listing_dates_path,
    )
    with pytest.raises(PerformanceV2ImportError, match="typed strategy mismatch"):
        import_performance_v2(replacement)
    with duckdb.connect(str(target), read_only=True) as connection:
        assert connection.execute(
            "select current_result_id from strategies where strategy_id = ?", [strategy_id]
        ).fetchone() == (result_id,)


def test_same_batch_equal_typed_entries_keep_manifest_first(tmp_path: Path) -> None:
    request, _ = _request(tmp_path, names=("alpha", "beta"))
    strategy_path = request.inbox / "strategies" / "beta.json"
    strategy = json.loads(strategy_path.read_text(encoding="utf-8"))
    strategy["mrs3"]["ma_close_long"]["len"] = 3  # type: ignore[index]
    strategy["mrs3"]["ma_close_short"]["len"] = 3  # type: ignore[index]
    strategy_bytes = json.dumps(strategy, separators=(",", ":")).encode()
    strategy_path.write_bytes(strategy_bytes)
    manifest_path = request.inbox / "inbox_manifest.json"
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    beta = manifest["entries"][1]
    beta["source_strategy_sha256"] = sha256(strategy_bytes).hexdigest()
    beta["strategy_version_id"] = _canonical_strategy_hash(strategy)
    manifest["v6_provenance"]["strategy_json_sha256"]["beta.json"] = beta["strategy_version_id"]
    manifest_path.write_text(json.dumps(manifest), encoding="utf-8")

    result = import_performance_v2(request)

    assert (result.imported_count, result.skipped_count) == (1, 1)
    with duckdb.connect(str(performance_v2_database_path(request.config)), read_only=True) as connection:
        assert connection.execute("select strategy_name from strategies").fetchone() == ("alpha",)


def test_same_batch_incomparable_typed_intervals_roll_back(tmp_path: Path) -> None:
    request, _ = _request(tmp_path, names=("alpha", "beta"))
    strategy_path = request.inbox / "strategies" / "beta.json"
    strategy = json.loads(strategy_path.read_text(encoding="utf-8"))
    strategy["mrs3"]["ma_close_long"]["len"] = 3  # type: ignore[index]
    strategy["mrs3"]["ma_close_short"]["len"] = 3  # type: ignore[index]
    strategy_bytes = json.dumps(strategy, separators=(",", ":")).encode()
    strategy_path.write_bytes(strategy_bytes)
    manifest_path = request.inbox / "inbox_manifest.json"
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    manifest["entries"][1]["source_strategy_sha256"] = sha256(strategy_bytes).hexdigest()
    manifest["entries"][1]["strategy_version_id"] = _canonical_strategy_hash(strategy)
    manifest["v6_provenance"]["strategy_json_sha256"]["beta.json"] = manifest["entries"][1]["strategy_version_id"]
    manifest_path.write_text(json.dumps(manifest), encoding="utf-8")
    for name, replacement in (("alpha", b"2025-12-20 - 2026-01-09"),
                              ("beta", b"2026-01-01 - 2026-01-20")):
        report_path = request.report_root / f"{name}.html"
        report_path.write_bytes(FIXTURE.read_bytes().replace(b"2026-01-01 - 2026-01-09", replacement))
        manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
        index = 0 if name == "alpha" else 1
        manifest["entries"][index]["source_report_sha256"] = sha256(report_path.read_bytes()).hexdigest()
        manifest_path.write_text(json.dumps(manifest), encoding="utf-8")
    with pytest.raises(PerformanceV2ImportError, match="incomparable"):
        import_performance_v2(request)
    with duckdb.connect(str(performance_v2_database_path(request.config)), read_only=True) as connection:
        assert connection.execute("select count(*) from strategies").fetchone() == (0,)
