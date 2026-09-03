from datetime import UTC, datetime
import json
from io import BytesIO
from pathlib import Path

import duckdb
from openpyxl import load_workbook
import pandas as pd
import pytest

from mrs3.performance_v2_selection import SelectionConfig, parse_selection_request, write_selection_workbook
from mrs3.performance_v2_selection_review import (
    META_SHEET,
    SelectionReviewError,
    apply_prior_rejected,
    canonical_contract,
    import_selection_review,
    latest_effective_finalists,
    new_run_metadata,
    persist_selection_snapshot,
)
from mrs3.panel import PanelController
from mrs3.performance_v2_store import initialize_performance_v2


def _database(tmp_path: Path, *, filename: str = "performance.duckdb") -> duckdb.DuckDBPyConnection:
    connection = duckdb.connect(str(tmp_path / filename))
    initialize_performance_v2(connection)
    now = datetime(2026, 9, 2, tzinfo=UTC)
    for strategy_id in (1, 2):
        connection.execute(
            """insert into strategies values (?, ?, 'BTCUSDT', 'LONG', '1h', 5, 1, 'run', ?, 'ACTIVE', null, ?, ?)""",
            [strategy_id, f"strategy-{strategy_id}", f"candidate-{strategy_id}", now, now],
        )
        connection.execute(
            """insert into strategy_results (
                result_id, strategy_id, report_start_utc, report_end_utc, exchange,
                commission_rate, initial_balance, final_balance, total_pnl, total_pnl_pct,
                max_drawdown, max_drawdown_pct, total_fees, total_trades, imported_at_utc
            ) values (?, ?, ?, ?, 'Bybit', .0004, 100, 110, 10, 10, 5, 5, 1, 10, ?)""",
            [100 + strategy_id, strategy_id, now, now, now],
        )
        connection.execute("update strategies set current_result_id = ? where strategy_id = ?", [100 + strategy_id, strategy_id])
    return connection


def _request():
    return parse_selection_request({"symbol": "BTCUSDT", "side": "LONG", "stages": [
        {"id": "rank_robust_top_n", "enabled": True, "scope": "pair_side", "top_n": 1},
    ]})


def _result() -> pd.DataFrame:
    return pd.DataFrame([
        {"strategy_id": 1, "result_id": 101, "strategy_name": "strategy-1", "symbol": "BTCUSDT", "side": "LONG",
         "timeframe": "1h", "order_count": 1, "close_ma_len": 5, "auto_status": "FINALIST", "finalist": True,
         "final_rank": 1, "final_score": 90.0, "elimination_reason": None, "analog_group_key": '["a"]',
         "auto_analog_of_strategy_id": None, "prior_rejected": False, "eliminated_by_rank_robust_top_n": False},
        {"strategy_id": 2, "result_id": 102, "strategy_name": "strategy-2", "symbol": "BTCUSDT", "side": "LONG",
         "timeframe": "1h", "order_count": 1, "close_ma_len": 5, "auto_status": "ANALOG", "finalist": False,
         "final_rank": None, "final_score": 80.0, "elimination_reason": "ANALOG", "analog_group_key": '["a"]',
         "auto_analog_of_strategy_id": 1, "prior_rejected": False, "eliminated_by_rank_robust_top_n": True},
    ])


def _export(connection: duckdb.DuckDBPyConnection, tmp_path: Path) -> tuple[Path, dict[str, str]]:
    request = _request()
    metadata = new_run_metadata(connection)
    path = write_selection_workbook(_result(), tmp_path / "review.xlsx", request, metadata)
    persist_selection_snapshot(connection, request, SelectionConfig(), _result(), metadata, path.read_bytes())
    return path, metadata


def test_contract_hashes_are_canonical() -> None:
    first = canonical_contract(_request(), SelectionConfig())
    second = canonical_contract(_request(), SelectionConfig())
    assert first == second
    assert len(first[1]) == len(first[3]) == 64


def test_export_persists_exact_snapshot_and_review_contract(tmp_path: Path) -> None:
    connection = _database(tmp_path)
    path, metadata = _export(connection, tmp_path)
    workbook = load_workbook(path)
    headers = [cell.value for cell in workbook["All candidates"][1]]

    assert workbook[META_SHEET].sheet_state == "veryHidden"
    assert {"Result ID", "Auto Status", "User Status", "Auto Rank", "User Rank", "Auto Analog Of ID", "Analog Of ID", "Comment"}.issubset(headers)
    assert connection.execute("select selection_run_id, candidate_count, auto_finalist_count from selection_runs").fetchone() == (metadata["selection_run_id"], 2, 1)
    assert connection.execute("select count(*) from selection_results").fetchone() == (2,)


def test_export_includes_current_retest_tag_and_editable_validation(tmp_path: Path) -> None:
    connection = _database(tmp_path)
    connection.execute(
        "insert into strategy_tags values (1, 'RETEST', 'PERIOD_INTEGRITY_AUDIT', 'audit.xlsx', now())"
    )
    request = _request()
    metadata = new_run_metadata(connection)
    result = apply_prior_rejected(connection, _result())
    path = write_selection_workbook(result, tmp_path / "retest.xlsx", request, metadata)
    sheet = load_workbook(path)["All candidates"]
    headers = [cell.value for cell in sheet[1]]
    values = {sheet.cell(row, headers.index("ID") + 1).value: sheet.cell(row, headers.index("RETEST") + 1).value for row in range(2, sheet.max_row + 1)}

    assert headers[headers.index("User Status") + 1] == "RETEST"
    assert values == {1: "RETEST", 2: None}
    retest_validation = next(validation for validation in sheet.data_validations.dataValidation if validation.formula1 == '"RETEST"')
    assert retest_validation.allow_blank
    assert str(retest_validation.sqref).endswith(f"{sheet.cell(sheet.max_row, headers.index('RETEST') + 1).column_letter}{sheet.max_row}")


def test_production_selection_export_preserves_existing_tags_on_round_trip(tmp_path: Path, monkeypatch) -> None:
    root = tmp_path / "panel"
    database_root = root / "data"
    database_root.mkdir(parents=True)
    connection = _database(database_root, filename="strategy_performance.duckdb")
    connection.execute(
        "insert into strategy_tags values (1, 'REJECTED', 'SELECTION_REVIEW', 'old-review', now()), (1, 'RETEST', 'PERIOD_INTEGRITY_AUDIT', 'audit.xlsx', now())"
    )
    connection.close()
    local_config = root / "config.local.json"
    local_config.write_text(json.dumps({"panel_paths": {"performance_db_root": "legacy"}}), encoding="utf-8")
    (root / "config.performance.json").write_text(
        json.dumps({"unified_performance_v2": {"database_root": "data", "workers": 1}}), encoding="utf-8"
    )
    controller = PanelController(root, local_config)
    import mrs3.panel as panel_module
    monkeypatch.setattr(panel_module, "selection_cache_status", lambda *_args, **_kwargs: {"ready": True})
    monkeypatch.setattr(panel_module, "load_selection_candidates", lambda *_args, **_kwargs: _result())

    _, data = controller.strategies_performance_v2_selection({"symbol": "BTCUSDT", "side": "LONG", "stages": []})
    sheet = load_workbook(BytesIO(data))["All candidates"]
    headers = {cell.value: cell.column for cell in sheet[1]}
    exported = {sheet.cell(row, headers["ID"]).value: sheet.cell(row, headers["RETEST"]).value for row in range(2, sheet.max_row + 1)}
    assert exported == {1: "RETEST", 2: None}

    with duckdb.connect(str(database_root / "strategy_performance.duckdb")) as connection:
        response = import_selection_review(connection, data)
        assert connection.execute(
            "select strategy_id, tag, source, source_ref from strategy_tags order by strategy_id, tag"
        ).fetchall() == [
            (1, "REJECTED", "SELECTION_REVIEW", response["review_import_id"]),
            (1, "RETEST", "SELECTION_REVIEW", response["review_import_id"]),
        ]


def test_review_accepts_blank_trailing_headers(tmp_path: Path) -> None:
    connection = _database(tmp_path)
    path, _ = _export(connection, tmp_path)
    workbook = load_workbook(path)
    sheet = workbook["All candidates"]
    first_blank = sheet.max_column + 1
    sheet.cell(1, first_blank, "")
    sheet.cell(1, first_blank + 1, "")
    padded = tmp_path / "padded.xlsx"
    workbook.save(padded)

    assert import_selection_review(connection, padded.read_bytes())["row_count"] == 2


def test_review_import_is_atomic_and_syncs_rejected_and_retest_tags(tmp_path: Path) -> None:
    connection = _database(tmp_path)
    old_retest_timestamp = datetime(2026, 1, 1, tzinfo=UTC)
    connection.execute(
        "insert into strategy_tags (strategy_id, tag, source, source_ref, updated_at_utc) values (2, 'RETEST', 'PERIOD_INTEGRITY_AUDIT', 'audit-old.xlsx', ?)",
        [old_retest_timestamp],
    )
    path, _ = _export(connection, tmp_path)
    workbook = load_workbook(path)
    sheet = workbook["All candidates"]
    headers = {cell.value: cell.column for cell in sheet[1]}
    sheet.cell(2, headers["User Status"], "RESERVE")
    sheet.cell(2, headers["User Rank"], 2)
    sheet.cell(3, headers["User Status"], "REJECTED")
    sheet.cell(3, headers["Analog Of ID"]).value = None
    edited = tmp_path / "edited.xlsx"
    workbook.save(edited)

    response = import_selection_review(connection, edited.read_bytes())

    assert response["finalist_count"] == 0
    assert connection.execute("select strategy_id, tag from strategy_tags order by tag").fetchall() == [(2, "REJECTED"), (2, "RETEST")]
    rejected = connection.execute(
        "select source, source_ref from strategy_tags where strategy_id = 2 and tag = 'REJECTED'"
    ).fetchone()
    assert rejected == ("SELECTION_REVIEW", response["review_import_id"])
    assert latest_effective_finalists(connection, "BTCUSDT") == (True, set())
    with pytest.raises(SelectionReviewError, match="SELECTION_REVIEW_ALREADY_IMPORTED"):
        import_selection_review(connection, edited.read_bytes())
    assert connection.execute("select count(*) from selection_review_imports").fetchone() == (1,)


def test_retest_sync_overwrites_asserted_and_leaves_absent_ids_untouched(tmp_path: Path) -> None:
    connection = _database(tmp_path)
    now = datetime(2026, 1, 1, tzinfo=UTC)
    connection.execute(
        "insert into strategies values (3, 'strategy-3', 'BTCUSDT', 'LONG', '1h', 5, 1, 'run', 'candidate-3', 'ACTIVE', null, ?, ?)",
        [now, now],
    )
    connection.execute(
        "insert into strategy_tags values (1, 'RETEST', 'PERIOD_INTEGRITY_AUDIT', 'audit-1', ?), (2, 'RETEST', 'PERIOD_INTEGRITY_AUDIT', 'audit-2', ?), (3, 'RETEST', 'RETEST_WORKFLOW', 'job-3', ?)",
        [now, now, now],
    )
    path, _ = _export(connection, tmp_path)
    workbook = load_workbook(path)
    sheet = workbook["All candidates"]
    headers = {cell.value: cell.column for cell in sheet[1]}
    sheet.cell(2, headers["RETEST"], " RETEST ")
    sheet.cell(3, headers["RETEST"], None)
    edited = tmp_path / "sync.xlsx"
    workbook.save(edited)

    response = import_selection_review(connection, edited.read_bytes())

    assert response["row_count"] == 2
    assert connection.execute(
        "select strategy_id, source, source_ref from strategy_tags where tag = 'RETEST' order by strategy_id"
    ).fetchall() == [
        (1, "SELECTION_REVIEW", response["review_import_id"]),
        (2, "PERIOD_INTEGRITY_AUDIT", "audit-2"),
        (3, "RETEST_WORKFLOW", "job-3"),
    ]


@pytest.mark.parametrize("value", ["retest", 1, 1.5, True, datetime(2026, 1, 1)])
def test_invalid_retest_values_are_rejected_before_writes(tmp_path: Path, value: object) -> None:
    connection = _database(tmp_path)
    path, _ = _export(connection, tmp_path)
    workbook = load_workbook(path)
    sheet = workbook["All candidates"]
    headers = {cell.value: cell.column for cell in sheet[1]}
    sheet.cell(2, headers["RETEST"], value)
    invalid = tmp_path / "invalid-retest.xlsx"
    workbook.save(invalid)
    before = connection.execute("select count(*) from selection_review_imports").fetchone()

    with pytest.raises(SelectionReviewError, match="SELECTION_REVIEW_INVALID_RETEST"):
        import_selection_review(connection, invalid.read_bytes())

    assert connection.execute("select count(*) from selection_review_imports").fetchone() == before
    assert connection.execute("select count(*) from selection_review_rows").fetchone() == (0,)
    connection.execute("select 1").fetchone()


def test_old_or_manually_built_workbook_is_not_accepted(tmp_path: Path) -> None:
    connection = _database(tmp_path)
    path = write_selection_workbook(_result(), tmp_path / "old.xlsx", _request())

    with pytest.raises(SelectionReviewError, match="SELECTION_REVIEW_SCHEMA_MISMATCH"):
        import_selection_review(connection, path.read_bytes())


@pytest.mark.parametrize("layout", ["non_adjacent", "blank_between", "missing_header"])
def test_retest_header_layout_is_strict_without_writes(tmp_path: Path, layout: str) -> None:
    connection = _database(tmp_path)
    path, _ = _export(connection, tmp_path)
    workbook = load_workbook(path)
    sheet = workbook["All candidates"]
    headers = {cell.value: cell.column for cell in sheet[1]}
    retest_column = headers["RETEST"]
    if layout == "non_adjacent":
        auto_rank_column = headers["Auto Rank"]
        sheet.cell(1, retest_column).value = "Auto Rank"
        sheet.cell(1, auto_rank_column).value = "RETEST"
    elif layout == "blank_between":
        sheet.insert_cols(retest_column)
        sheet.cell(1, retest_column).value = None
    else:
        for row in range(2, sheet.max_row + 1):
            sheet.cell(row, retest_column).value = "RETEST"
        sheet.cell(1, retest_column).value = None
    invalid = tmp_path / f"invalid-header-{layout}.xlsx"
    workbook.save(invalid)

    with pytest.raises(SelectionReviewError, match="SELECTION_REVIEW_SCHEMA_MISMATCH"):
        import_selection_review(connection, invalid.read_bytes())
    assert connection.execute("select count(*) from selection_review_imports").fetchone() == (0,)
    assert connection.execute("select count(*) from selection_review_rows").fetchone() == (0,)
    assert connection.execute("select count(*) from strategy_tags").fetchone() == (0,)
    assert connection.execute("select 1").fetchone() == (1,)


def test_changed_automatic_status_is_rejected_without_writes(tmp_path: Path) -> None:
    connection = _database(tmp_path)
    path, _ = _export(connection, tmp_path)
    workbook = load_workbook(path)
    sheet = workbook["All candidates"]
    headers = {cell.value: cell.column for cell in sheet[1]}
    sheet.cell(2, headers["Auto Status"], "FILTERED")
    changed = tmp_path / "changed.xlsx"
    workbook.save(changed)

    with pytest.raises(SelectionReviewError, match="SELECTION_REVIEW_AUTOMATIC_FIELDS_CHANGED"):
        import_selection_review(connection, changed.read_bytes())
    assert connection.execute("select count(*) from selection_review_imports").fetchone() == (0,)


def test_formula_anywhere_in_review_workbook_is_rejected(tmp_path: Path) -> None:
    connection = _database(tmp_path)
    path, _ = _export(connection, tmp_path)
    workbook = load_workbook(path)
    workbook["Finalists"]["A2"] = "=1"
    changed = tmp_path / "formula.xlsx"
    workbook.save(changed)

    with pytest.raises(SelectionReviewError, match="SELECTION_REVIEW_INVALID_FILE"):
        import_selection_review(connection, changed.read_bytes())


def test_formula_in_retest_is_rejected_without_writes(tmp_path: Path) -> None:
    connection = _database(tmp_path)
    path, _ = _export(connection, tmp_path)
    workbook = load_workbook(path)
    sheet = workbook["All candidates"]
    headers = {cell.value: cell.column for cell in sheet[1]}
    sheet.cell(2, headers["RETEST"], "=\"RETEST\"")
    changed = tmp_path / "formula-retest.xlsx"
    workbook.save(changed)

    with pytest.raises(SelectionReviewError, match="SELECTION_REVIEW_INVALID_FILE"):
        import_selection_review(connection, changed.read_bytes())
    assert connection.execute("select count(*) from selection_review_imports").fetchone() == (0,)
    assert connection.execute("select count(*) from selection_review_rows").fetchone() == (0,)
    assert connection.execute("select count(*) from strategy_tags").fetchone() == (0,)
    assert connection.execute("select 1").fetchone() == (1,)


def test_first_asserted_retest_has_full_provenance_and_coexists_with_rejected(tmp_path: Path) -> None:
    connection = _database(tmp_path)
    path, _ = _export(connection, tmp_path)
    workbook = load_workbook(path)
    sheet = workbook["All candidates"]
    headers = {cell.value: cell.column for cell in sheet[1]}
    sheet.cell(2, headers["User Status"], "REJECTED")
    sheet.cell(2, headers["User Rank"]).value = None
    sheet.cell(2, headers["Analog Of ID"]).value = None
    sheet.cell(2, headers["RETEST"], "RETEST")
    sheet.cell(3, headers["User Status"], "RESERVE")
    sheet.cell(3, headers["User Rank"], 2)
    sheet.cell(3, headers["Analog Of ID"]).value = None
    edited = tmp_path / "asserted-retest.xlsx"
    workbook.save(edited)

    response = import_selection_review(connection, edited.read_bytes())

    assert connection.execute(
        "select tag, source, source_ref from strategy_tags where strategy_id = 1 order by tag"
    ).fetchall() == [
        ("REJECTED", "SELECTION_REVIEW", response["review_import_id"]),
        ("RETEST", "SELECTION_REVIEW", response["review_import_id"]),
    ]


def test_newer_export_makes_older_workbook_non_latest(tmp_path: Path) -> None:
    connection = _database(tmp_path)
    old_path, _ = _export(connection, tmp_path)
    old_bytes = old_path.read_bytes()
    _export(connection, tmp_path)

    with pytest.raises(SelectionReviewError, match="SELECTION_REVIEW_NOT_LATEST_RUN"):
        import_selection_review(connection, old_bytes)
    assert connection.execute("select count(*) from selection_review_imports").fetchone() == (0,)


def test_stale_result_ids_reject_the_complete_review(tmp_path: Path) -> None:
    connection = _database(tmp_path)
    path, _ = _export(connection, tmp_path)
    connection.execute("update strategies set current_result_id = 999 where strategy_id = 2")

    with pytest.raises(SelectionReviewError) as raised:
        import_selection_review(connection, path.read_bytes())
    assert raised.value.code == "SELECTION_REVIEW_STALE_RESULTS"
    assert raised.value.details == [2]
    assert connection.execute("select count(*) from selection_review_imports").fetchone() == (0,)


def test_review_can_exceed_auto_top_n_and_later_remove_rejected_tag(tmp_path: Path) -> None:
    connection = _database(tmp_path)
    path, _ = _export(connection, tmp_path)
    workbook = load_workbook(path)
    sheet = workbook["All candidates"]
    headers = {cell.value: cell.column for cell in sheet[1]}
    sheet.cell(3, headers["User Status"], "REJECTED")
    sheet.cell(3, headers["Analog Of ID"]).value = None
    first = tmp_path / "first-review.xlsx"
    workbook.save(first)
    import_selection_review(connection, first.read_bytes())
    assert connection.execute("select strategy_id from strategy_tags").fetchall() == [(2,)]

    workbook = load_workbook(first)
    sheet = workbook["All candidates"]
    sheet.cell(3, headers["User Status"], "FINALIST")
    sheet.cell(3, headers["User Rank"], 2)
    second = tmp_path / "second-review.xlsx"
    workbook.save(second)
    imported = import_selection_review(connection, second.read_bytes())

    assert imported["finalist_count"] == 2
    assert connection.execute("select count(*) from selection_review_imports").fetchone() == (2,)
    assert connection.execute("select count(*) from strategy_tags").fetchone() == (0,)
    assert latest_effective_finalists(connection, "BTCUSDT") == (True, {1, 2})


def test_analog_must_target_a_finalist_or_reserve_in_same_run(tmp_path: Path) -> None:
    connection = _database(tmp_path)
    path, _ = _export(connection, tmp_path)
    workbook = load_workbook(path)
    sheet = workbook["All candidates"]
    headers = {cell.value: cell.column for cell in sheet[1]}
    sheet.cell(2, headers["User Status"], "REJECTED")
    sheet.cell(2, headers["User Rank"]).value = None
    invalid = tmp_path / "invalid-analog.xlsx"
    workbook.save(invalid)

    with pytest.raises(SelectionReviewError, match="SELECTION_REVIEW_INVALID_ANALOG"):
        import_selection_review(connection, invalid.read_bytes())
