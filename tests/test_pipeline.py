from __future__ import annotations

from dataclasses import replace
from decimal import Decimal
import hashlib
import json
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook

from mrs3.config import AlgorithmConfig
from mrs3.models import Side
from mrs3.pipeline import ALGORITHM_VERSION, PipelineInput, SelectionInputs, _apply_package_event_unions, _pair_history, _publish_strategies, _write_json_atomic, run_published_pipeline, run_selection
from tests.factories import write_selection_inputs
from tests.test_package_loader import write_real_package


def _digest_files(directory: Path, pattern: str) -> str:
    digest = hashlib.sha256()
    for path in sorted(directory.glob(pattern)):
        digest.update(path.relative_to(directory).as_posix().encode())
        digest.update(path.read_bytes())
    return digest.hexdigest()


def test_strategy_json_writer_preserves_template_key_order(tmp_path: Path) -> None:
    strategy = {
        "name": "A",
        "is_runing": False,
        "exchange": {"name": "Bybit", "account": "main"},
        "basic": {"strategy": "mrs3", "symbol": "ONUSDT"},
        "mrs3": {
            "ma_long": [
                {"id": 1, "side": "buy", "len": 3, "multiplier": 0.996, "lot_x": 1.0}
            ],
            "ma_short": [],
        },
    }
    target = tmp_path / "strategy.json"

    _write_json_atomic(target, strategy)

    written = json.loads(target.read_text(encoding="utf-8"))
    assert list(written) == list(strategy)
    assert list(written["exchange"]) == list(strategy["exchange"])
    assert list(written["basic"]) == list(strategy["basic"])
    assert list(written["mrs3"]) == list(strategy["mrs3"])
    assert list(written["mrs3"]["ma_long"][0]) == list(strategy["mrs3"]["ma_long"][0])


def test_published_pipeline_scopes_listing_dates_and_fails_closed(monkeypatch: pytest.MonkeyPatch) -> None:
    import mrs3.pipeline as pipeline

    points = pd.DataFrame([{"symbol": "AAAUSDT", "side": "LONG"}])
    monkeypatch.setattr(
        pipeline,
        "annotate_eligibility",
        lambda values, _: values.assign(economic_pass=True, event_eligible=True),
    )
    monkeypatch.setattr(pipeline, "annotate_refine", lambda values, _: (values, pd.DataFrame()))
    monkeypatch.setattr(pipeline, "build_plateaus", lambda values, _: (values, pd.DataFrame()))
    monkeypatch.setattr(pipeline, "find_isolated_peaks", lambda *_: pd.DataFrame())
    monkeypatch.setattr(pipeline, "build_close_profiles", lambda _, plateaus, __: (plateaus, pd.DataFrame()))
    monkeypatch.setattr(pipeline, "select_base_one_order", lambda *_: pd.DataFrame())
    monkeypatch.setattr(pipeline, "build_structures", lambda *_: (pd.DataFrame(columns=["status", "plateau_ids"]), pd.DataFrame()))
    monkeypatch.setattr(pipeline, "load_listing_dates", lambda _: {"AAAUSDT": pd.Timestamp("2020-01-01", tz="UTC"), "UNUSED": pd.Timestamp("2020-01-01", tz="UTC")})

    result = run_published_pipeline(PipelineInput("surface", points), Path("dates.csv"), Side.LONG, AlgorithmConfig.defaults())

    assert result.algorithm_version == ALGORITHM_VERSION
    assert result.algorithm_config["listing_dates"] == {"AAAUSDT": "2020-01-01T00:00:00+00:00"}
    monkeypatch.setattr(pipeline, "load_listing_dates", lambda _: {})
    with pytest.raises(ValueError, match="missing listing dates"):
        run_published_pipeline(PipelineInput("surface", points), Path("dates.csv"), Side.LONG, AlgorithmConfig.defaults())
    with pytest.raises(ValueError, match="side does not match"):
        run_published_pipeline(PipelineInput("surface", points), Path("dates.csv"), Side.SHORT, AlgorithmConfig.defaults())


def test_pipeline_builds_two_plateaus_and_validated_json(tmp_path: Path) -> None:
    paths = write_selection_inputs(tmp_path / "inputs")
    config = AlgorithmConfig.from_json(paths["config"])
    inputs = SelectionInputs(
        csv_path=paths["csv"],
        dates_path=paths["dates"],
        template_path=paths["template"],
        side=Side.LONG,
        output_dir=tmp_path / "output",
    )

    result = run_selection(inputs, config)

    assert len(result.plateaus) == 2
    assert set(result.structures["order_count"]) == {2}
    assert len(result.generated_strategies) == 5
    assert result.manifest["ready_json_count"] == 5
    assert result.manifest["base_json_count"] == 1
    assert result.manifest["mrs3_json_count"] == 4
    assert result.manifest["ready_structure_count_by_orders"] == {
        "2ORD": 2,
        "3ORD": 0,
        "4ORD": 0,
    }
    assert result.manifest["ready_json_count"] == (
        result.manifest["base_json_count"]
        + result.manifest["mrs3_json_count"]
    )
    assert result.manifest["mrs3_json_count"] == 2 * result.manifest[
        "ready_structure_count"
    ]
    assert result.manifest["event_mode"] == "legacy_trades_proxy"
    assert set(result.plateaus["plateau_event_count"]) == {"N/A_LEGACY_PROXY"}
    assert set(result.plateaus["plateau_event_ids_hash"]) == {"N/A_LEGACY_PROXY"}
    assert result.manifest["event_eligible_point_count"] == int(result.points["event_eligible"].sum())
    assert result.lot_variants.loc[
        result.lot_variants["variant_type"].eq("BASE_1ORD"),
        ["plateau_point_count", "base_point_trades", "plateau_total_trades"],
    ].notna().all().all()
    assert (inputs.output_dir / "audit.xlsx").exists()
    workbook = load_workbook(inputs.output_dir / "audit.xlsx", read_only=True)
    assert workbook.sheetnames == [
        "00_Input_Audit",
        "01_Pair_History",
        "02_Filtering",
        "03_Refine_Required",
        "04_Plateau_Points",
        "05_Plateau_Library",
        "06_CloseMA_Profile",
        "07_Isolated_Peaks",
        "08_1ORD",
        "09_CloseMA_Families",
        "10_MRS3_Structures",
        "11_Lot_Variants",
        "12_Ready_JSON",
        "13_Deep_Gap_Research",
        "14_Recalibration",
        "15_Config",
        "16_Point_Events",
        "17_Plateau_Events",
    ]
    history = pd.read_excel(inputs.output_dir / "audit.xlsx", sheet_name="01_Pair_History")
    assert pd.to_datetime(history["report_start"], utc=True).eq(
        pd.Timestamp("2026-07-15", tz="UTC")
    ).all()
    assert pd.to_datetime(history["report_end"], utc=True).eq(
        pd.Timestamp("2026-08-06", tz="UTC")
    ).all()
    assert history["effective_days"].eq(22.0).all()


def test_pair_history_rejects_mixed_report_windows() -> None:
    points = pd.DataFrame(
        [
            {
                "symbol": "AAAUSDT",
                "side": "LONG",
                "listing_date": pd.Timestamp("2026-07-01", tz="UTC"),
                "report_start": pd.Timestamp("2026-07-15", tz="UTC"),
                "report_end": pd.Timestamp("2026-08-06", tz="UTC"),
                "effective_start": pd.Timestamp("2026-07-15", tz="UTC"),
                "effective_days": 22.0,
                "history_pass": True,
            },
            {
                "symbol": "AAAUSDT",
                "side": "LONG",
                "listing_date": pd.Timestamp("2026-07-01", tz="UTC"),
                "report_start": pd.Timestamp("2026-07-16", tz="UTC"),
                "report_end": pd.Timestamp("2026-08-07", tz="UTC"),
                "effective_start": pd.Timestamp("2026-07-16", tz="UTC"),
                "effective_days": 22.0,
                "history_pass": True,
            },
        ]
    )

    with pytest.raises(AssertionError, match="one report period"):
        _pair_history(points)


def test_pipeline_real_event_metadata_without_mappings_fails_closed(tmp_path: Path) -> None:
    paths = write_selection_inputs(tmp_path / "inputs")
    source = pd.read_csv(paths["csv"])
    source["event_mode"] = "real_independent_events"
    source["point_event_count"] = 3
    source["event_ids_hash"] = "sha256:sample"
    source.to_csv(paths["csv"], index=False)
    config = AlgorithmConfig.from_json(paths["config"])
    inputs = SelectionInputs(paths["csv"], paths["dates"], paths["template"], Side.LONG, tmp_path / "output")

    result = run_selection(inputs, config)

    assert result.manifest["event_mode"] == "real_independent_events"
    assert not result.plateaus["ready"].any()
    assert result.plateaus["plateau_event_count"].isna().all()
    assert result.plateaus["plateau_event_ids_hash"].isna().all()


def test_pipeline_package_manifest_and_plateaus_use_distinct_real_events(
    tmp_path: Path,
) -> None:
    paths = write_selection_inputs(tmp_path / "inputs")
    package, events_by_point = write_real_package(tmp_path / "package", paths)
    config = AlgorithmConfig.from_json(paths["config"])
    inputs = SelectionInputs(
        csv_path=None,
        dates_path=paths["dates"],
        template_path=paths["template"],
        side=Side.LONG,
        output_dir=tmp_path / "output",
        source_package_dir=package,
    )

    result = run_selection(inputs, config)

    assert result.manifest["event_mode"] == "real_independent_events"
    assert result.manifest["source_package_manifest_sha256"] == hashlib.sha256(
        (package / "package_manifest.json").read_bytes()
    ).hexdigest()
    for plateau in result.plateaus.itertuples(index=False):
        event_ids = sorted(
            {
                event_id
                for point_id in plateau.all_point_ids
                for event_id in events_by_point[point_id]
            }
        )
        assert plateau.plateau_event_count == len(event_ids)
        assert plateau.plateau_event_ids_hash == hashlib.sha256(
            "|".join(event_ids).encode("utf-8")
        ).hexdigest()


def test_pipeline_real_plateau_union_includes_event_ineligible_geometry_point(
    tmp_path: Path,
) -> None:
    paths = write_selection_inputs(tmp_path / "inputs")
    package, events_by_point = write_real_package(tmp_path / "package", paths)
    ineligible_point = next(
        point_id for point_id in events_by_point if "|190|" in point_id
    )
    events_by_point[ineligible_point] = ("ineligible-unique-a", "ineligible-unique-b")
    points = pd.read_csv(package / "points.csv")
    points.loc[points["point_id"].eq(ineligible_point), "point_event_count"] = 2
    points.loc[points["point_id"].eq(ineligible_point), "event_ids_hash"] = (
        hashlib.sha256("|".join(events_by_point[ineligible_point]).encode("utf-8")).hexdigest()
    )
    points.to_csv(package / "points.csv", index=False, lineterminator="\n")
    pd.DataFrame(
        [
            {"point_id": point_id, "event_id": event_id}
            for point_id, event_ids in events_by_point.items()
            for event_id in event_ids
        ]
    ).sort_values(["point_id", "event_id"], kind="mergesort").to_csv(
        package / "point_events.csv", index=False, lineterminator="\n"
    )
    config = AlgorithmConfig.from_json(paths["config"])
    inputs = SelectionInputs(
        csv_path=None,
        dates_path=paths["dates"],
        template_path=paths["template"],
        side=Side.LONG,
        output_dir=tmp_path / "output",
        source_package_dir=package,
    )

    result = run_selection(inputs, config)

    plateau = next(
        row
        for row in result.plateaus.itertuples(index=False)
        if ineligible_point in row.all_point_ids
    )
    assert not bool(
        result.points.loc[
            result.points["point_id"].eq(ineligible_point), "event_eligible"
        ].iloc[0]
    )
    expected_events = sorted(
        {
            event_id
            for point_id in plateau.all_point_ids
            for event_id in events_by_point[point_id]
        }
    )
    assert plateau.plateau_event_count == len(expected_events)
    assert plateau.plateau_event_ids_hash == hashlib.sha256(
        "|".join(expected_events).encode("utf-8")
    ).hexdigest()


def test_pipeline_monthly_event_scalar_missing_fails_plateau_admission() -> None:
    points = pd.DataFrame([
        {"point_id": "A", "event_mode": "real_independent_events", "_event_ids": ("e1",), "events_last_30d": None},
        {"point_id": "B", "event_mode": "real_independent_events", "_event_ids": ("e2",), "events_last_30d": 1},
    ])
    plateaus = pd.DataFrame([{
        "plateau_id": "P1",
        "all_point_ids": ("A", "B"),
        "status": "MRS3_USABLE",
        "ready": True,
    }])

    result = _apply_package_event_unions(points, plateaus)

    assert result.loc[0, "plateau_event_count"] is None
    assert result.loc[0, "status"] == "INSUFFICIENT_INDEPENDENT_EVENTS"
    assert not bool(result.loc[0, "ready"])


def test_pipeline_monthly_event_count_sums_points_without_deduplicating_ids() -> None:
    points = pd.DataFrame([
        {
            "point_id": "A",
            "event_mode": "real_independent_events",
            "_event_ids": ("shared",),
            "events_last_30d": 3,
        },
        {
            "point_id": "B",
            "event_mode": "real_independent_events",
            "_event_ids": ("shared",),
            "events_last_30d": 4,
        },
    ])
    plateaus = pd.DataFrame([{
        "plateau_id": "P1",
        "all_point_ids": ("A", "B"),
        "status": "MRS3_USABLE",
        "ready": True,
    }])

    result = _apply_package_event_unions(points, plateaus)

    assert result.loc[0, "plateau_event_count"] == 7
    assert result.loc[0, "plateau_event_ids_hash"] == hashlib.sha256(
        b"shared"
    ).hexdigest()


def test_pipeline_monthly_event_scalar_invalidates_only_its_plateau() -> None:
    points = pd.DataFrame([
        {
            "point_id": "A",
            "event_mode": "real_independent_events",
            "_event_ids": ("e1",),
            "events_last_30d": None,
        },
        {
            "point_id": "B",
            "event_mode": "real_independent_events",
            "_event_ids": ("e2",),
            "events_last_30d": 1,
        },
        {
            "point_id": "C",
            "event_mode": "real_independent_events",
            "_event_ids": ("e3",),
            "events_last_30d": 4,
        },
    ])
    plateaus = pd.DataFrame([
        {
            "plateau_id": "P1",
            "all_point_ids": ("A", "B"),
            "status": "MRS3_USABLE",
            "ready": True,
        },
        {
            "plateau_id": "P2",
            "all_point_ids": ("C",),
            "status": "MRS3_USABLE",
            "ready": True,
        },
    ])

    result = _apply_package_event_unions(points, plateaus)

    assert pd.isna(result.loc[0, "plateau_event_count"])
    assert result.loc[0, "status"] == "INSUFFICIENT_INDEPENDENT_EVENTS"
    assert not bool(result.loc[0, "ready"])
    assert result.loc[1, "plateau_event_count"] == 4
    assert result.loc[1, "status"] == "MRS3_USABLE"
    assert bool(result.loc[1, "ready"])


def test_pipeline_rejects_mixed_event_modes_before_plateau_union() -> None:
    points = pd.DataFrame([
        {
            "point_id": "A",
            "event_mode": "legacy_trades_proxy",
            "_event_ids": (),
        },
        {
            "point_id": "B",
            "event_mode": "real_independent_events",
            "_event_ids": ("e1",),
        },
    ])
    plateaus = pd.DataFrame([{
        "plateau_id": "P1",
        "all_point_ids": ("A", "B"),
        "status": "MRS3_USABLE",
        "ready": True,
    }])

    with pytest.raises(ValueError, match="mixed event modes"):
        _apply_package_event_unions(points, plateaus)


@pytest.mark.parametrize("mode", [None, "unsupported"])
def test_pipeline_rejects_missing_or_unknown_event_mode(mode: object) -> None:
    points = pd.DataFrame([{"point_id": "A", "event_mode": mode}])
    plateaus = pd.DataFrame([{
        "plateau_id": "P1", "all_point_ids": ("A",), "status": "MRS3_USABLE", "ready": True,
    }])

    with pytest.raises(ValueError, match="event_mode is required|unknown event mode"):
        _apply_package_event_unions(points, plateaus)


def test_same_inputs_produce_identical_digest_rows_and_json(tmp_path: Path) -> None:
    paths = write_selection_inputs(tmp_path / "inputs")
    config = AlgorithmConfig.from_json(paths["config"])
    first_inputs = SelectionInputs(paths["csv"], paths["dates"], paths["template"], Side.LONG, tmp_path / "a")
    second_inputs = SelectionInputs(paths["csv"], paths["dates"], paths["template"], Side.LONG, tmp_path / "b")

    first = run_selection(first_inputs, config)
    second = run_selection(second_inputs, config)

    assert first.manifest["deterministic_digest"] == second.manifest["deterministic_digest"]
    assert _digest_files(tmp_path / "a", "strategies/*.json") == _digest_files(
        tmp_path / "b", "strategies/*.json"
    )
    assert (tmp_path / "a" / "audit.xlsx").read_bytes() == (
        tmp_path / "b" / "audit.xlsx"
    ).read_bytes()


def test_coarse_domain_emits_refine_requests_without_fabricated_points(tmp_path: Path) -> None:
    paths = write_selection_inputs(tmp_path / "inputs")
    config = replace(
        AlgorithmConfig.from_json(paths["config"]),
        shift_domain_min_bp=30,
        shift_domain_max_bp=270,
        canonical_shifts_bp=(
            30,
            40,
            50,
            60,
            70,
            90,
            110,
            140,
            170,
            190,
            230,
            270,
        ),
        gap_rules=((30, 80, 80), (80, 200, 100), (200, 271, 130)),
    )
    inputs = SelectionInputs(paths["csv"], paths["dates"], paths["template"], Side.LONG, tmp_path / "output")

    result = run_selection(inputs, config)

    assert not result.refine_requests.empty
    exported_point_ids = {
        order["point_id"]
        for structure in result.structures.get("orders", [])
        for order in structure
    }
    assert exported_point_ids.issubset(set(result.points["point_id"]))


def test_pipeline_finishes_with_audit_when_no_plateau_is_ready(tmp_path: Path) -> None:
    paths = write_selection_inputs(tmp_path / "inputs")
    config = replace(
        AlgorithmConfig.from_json(paths["config"]),
        economic_min_win_rate_pct=Decimal("100"),
    )
    inputs = SelectionInputs(
        paths["csv"],
        paths["dates"],
        paths["template"],
        Side.LONG,
        tmp_path / "output",
    )

    result = run_selection(inputs, config)

    assert result.plateaus.empty
    assert result.structures.empty
    assert result.generated_strategies == []
    assert result.manifest["ready_plateau_count"] == 0
    assert result.manifest["ready_json_count"] == 0
    assert (inputs.output_dir / "audit.xlsx").exists()


def test_rerun_replaces_strategy_directory_without_stale_json(tmp_path: Path) -> None:
    paths = write_selection_inputs(tmp_path / "inputs")
    config = AlgorithmConfig.from_json(paths["config"])
    inputs = SelectionInputs(
        paths["csv"],
        paths["dates"],
        paths["template"],
        Side.LONG,
        tmp_path / "output",
    )
    first = run_selection(inputs, config)
    target = inputs.output_dir / "strategies"
    target_identity = (target.stat().st_dev, target.stat().st_ino)
    stale = inputs.output_dir / "strategies" / "STALE.json"
    stale.write_text('{"name":"STALE"}', encoding="utf-8")
    stale_dir = inputs.output_dir / "strategies" / "stale-directory"
    stale_dir.mkdir()
    (stale_dir / "nested.json").write_text('{"stale":true}', encoding="utf-8")

    second = run_selection(inputs, config)

    expected = sorted(second.lot_variants["json_filename"])
    actual = sorted(path.name for path in (inputs.output_dir / "strategies").glob("*.json"))
    assert len(first.generated_strategies) == len(second.generated_strategies)
    assert actual == expected
    assert not stale.exists()
    assert not stale_dir.exists()
    assert (target.stat().st_dev, target.stat().st_ino) == target_identity


def test_strategy_publication_rolls_back_after_install_failure(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    output = tmp_path / "output"
    target = output / "strategies"
    target.mkdir(parents=True)
    (target / "OLD.json").write_text('{"name":"OLD"}', encoding="utf-8")
    stale_dir = target / "old-directory"
    stale_dir.mkdir()
    (stale_dir / "nested.json").write_text("old", encoding="utf-8")
    identity = (target.stat().st_dev, target.stat().st_ino)
    variants = pd.DataFrame([{"json_filename": "NEW-1.json"}, {"json_filename": "NEW-2.json"}])
    generated = [{"name": "NEW-1"}, {"name": "NEW-2"}]
    original_replace = Path.replace
    failed = False

    def fail_second_install(self: Path, destination: Path) -> Path:
        nonlocal failed
        if self.name == "NEW-2.json" and destination.parent == target and not failed:
            failed = True
            raise OSError("injected install failure")
        return original_replace(self, destination)

    monkeypatch.setattr(Path, "replace", fail_second_install)
    with pytest.raises(OSError, match="injected install failure"):
        _publish_strategies(output, variants, generated)

    assert failed
    assert sorted(path.name for path in target.iterdir()) == ["OLD.json", "old-directory"]
    assert (target / "OLD.json").read_bytes() == b'{"name":"OLD"}'
    assert (stale_dir / "nested.json").read_bytes() == b"old"
    assert not (target / "NEW-1.json").exists()
    assert not (target / "NEW-2.json").exists()
    assert (target.stat().st_dev, target.stat().st_ino) == identity
    assert not list(output.glob(".strategies.mrs3-stage-*"))
    assert not (output / ".strategies.mrs3-backup").exists()


def test_strategy_publication_retains_backup_when_rollback_fails(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    output = tmp_path / "output"
    target = output / "strategies"
    target.mkdir(parents=True)
    old = target / "OLD.json"
    old.write_bytes(b"old")
    variants = pd.DataFrame([{"json_filename": "NEW.json"}])
    original_replace = Path.replace
    install_failed = False
    restore_failed = False

    def fail_restore(self: Path, destination: Path) -> Path:
        nonlocal install_failed, restore_failed
        if self.name == "NEW.json" and destination.parent == target and not install_failed:
            install_failed = True
            raise OSError("injected install failure")
        if self.name == "OLD.json" and destination.parent == target and install_failed and not restore_failed:
            restore_failed = True
            raise OSError("injected restore failure")
        return original_replace(self, destination)

    monkeypatch.setattr(Path, "replace", fail_restore)
    with pytest.raises(OSError, match="injected install failure"):
        _publish_strategies(output, variants, [{"name": "NEW"}])

    backup = output / ".strategies.mrs3-backup"
    assert install_failed and restore_failed
    assert (backup / "OLD.json").read_bytes() == b"old"
    assert not list(output.glob(".strategies.mrs3-stage-*"))
    with pytest.raises(RuntimeError, match="strategy backup requires recovery"):
        _publish_strategies(output, variants, [{"name": "NEW"}])
