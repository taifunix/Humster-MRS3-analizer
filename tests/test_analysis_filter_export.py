from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
import sys
from types import MappingProxyType, ModuleType

from openpyxl import load_workbook
import pytest


class _RunConnection:
    def execute(self, _query: str, parameters: list[str]):
        assert parameters == ["RUN-1"]
        return self

    def fetchone(self):
        return ("0.7-filter-v1",)


@dataclass(frozen=True)
class FilterResult:
    combined_rows: tuple[dict[str, object], ...]
    per_criterion: MappingProxyType


@dataclass(frozen=True)
class EngineFilterResult:
    rows: tuple[dict[str, object], ...]
    combined: tuple[dict[str, object], ...]
    standalone: MappingProxyType


def _order(pnl: float, dd: float, support: float, events: int) -> dict[str, object]:
    return {
        "source_pnl_pct": pnl,
        "source_efficiency": dd,
        "close_support": support,
        "point_event_count": events,
    }


def _row(
    candidate_id: str,
    behavior_key: str,
    status: str,
    *,
    deferred_by: str | None = None,
    deferred_by_candidate_id: str | None = None,
) -> dict[str, object]:
    return {
        "candidate_id": candidate_id,
        "comparison_key": behavior_key,
        "filter_status": status,
        "deferred_by": deferred_by,
        "deferred_by_candidate_id": deferred_by_candidate_id,
        "criteria": ("source_pnl", "close_support"),
        "orders_a": (_order(12.5, 2.5, 0.9, 11),),
        "orders_b": (_order(10.0, 3.5, 0.8, 9),),
    }


def _install_filter(monkeypatch, result: FilterResult) -> None:
    module = ModuleType("mrs3.analysis_shortlist")
    module.filter_analysis_candidates = lambda *_args: result
    monkeypatch.setitem(sys.modules, "mrs3.analysis_shortlist", module)


def test_audit_rows_require_comparison_key_and_do_not_use_trades_as_events() -> None:
    from mrs3.analysis_filter_export import _normal_row

    with pytest.raises(ValueError, match="comparison_key"):
        _normal_row({"candidate_id": "C1", "behavior_key": "legacy"}, ())
    normalized = _normal_row({
        "candidate_id": "C1", "comparison_key": "K1",
        "orders": ({"trades": 99},),
    }, ())
    assert normalized.get("order1_point_event_count_a") is None


def test_export_filter_audit_writes_ordered_sheets_and_sorted_numeric_rows(
    tmp_path: Path, monkeypatch
) -> None:
    from mrs3.analysis_filter_export import export_filter_audit

    result = EngineFilterResult(
        rows=(
            _row("C3", "A", "READY_AFTER_FILTERS"),
        ),
        combined=(
            _row("C2", "B", "DEFERRED_REDUNDANT", deferred_by="CloseSupport", deferred_by_candidate_id="C1"),
        ),
        standalone=MappingProxyType(
            {
                "source_pnl": (),
                "close_support": (
                    _row("C2", "B", "DEFERRED_REDUNDANT", deferred_by="CloseSupport", deferred_by_candidate_id="C1"),
                ),
            }
        ),
    )
    _install_filter(monkeypatch, result)

    output = export_filter_audit(
        _RunConnection(),
        "RUN-1",
        ("close_support", "source_pnl"),
        tmp_path / "audit-directory",
    )

    assert output == tmp_path / "audit-directory" / "phase2_filter_audit.xlsx"
    workbook = load_workbook(output, data_only=True)
    assert workbook.sheetnames == [
        "Summary",
        "READY_AFTER_FILTERS",
        "Source PnL",
        "CloseSupport",
        "DEFERRED_COMBINED",
    ]
    summary = {
        row[0]: row[1]
        for row in workbook["Summary"].iter_rows(min_row=2, values_only=True)
    }
    assert summary["input_count"] == 2
    assert summary["algorithm_version"] == "0.7-filter-v1"
    assert summary["ready_count"] == 1
    assert summary["deferred_count"] == 1

    ready = workbook["READY_AFTER_FILTERS"]
    assert list(ready.values)[1][0] == "C3"
    assert ready["H2"].value == 12.5
    assert isinstance(ready["H2"].value, float)

    source_pnl = workbook["Source PnL"]
    assert source_pnl.max_row == 1
    assert tuple(source_pnl.values)[0][:7] == (
        "candidate_id",
        "comparison_key",
        "filter_status",
        "deferred_by",
        "deferred_by_candidate_id",
        "criteria",
        "defer_reason",
    )

    combined = workbook["DEFERRED_COMBINED"]
    assert list(combined.values)[1][0:6] == (
        "C2",
        "B",
        "DEFERRED_REDUNDANT",
        "CloseSupport",
        "C1",
        "source_pnl, close_support",
    )
    assert combined["H2"].value == 12.5
    assert combined["K2"].value == 11
    assert combined["O2"].value == 9


def test_export_filter_audit_is_deterministic_for_file_output_and_all_criteria(
    tmp_path: Path, monkeypatch
) -> None:
    from mrs3.analysis_filter_export import export_filter_audit

    result = FilterResult(
        combined_rows=(_row("C1", "B", "DEFERRED_REDUNDANT", deferred_by="Source PnL", deferred_by_candidate_id="C0"),),
        per_criterion=MappingProxyType(
            {
                "source_pnl": (),
                "efficiency": (),
                "close_support": (),
                "point_event_count": (),
            }
        ),
    )
    _install_filter(monkeypatch, result)

    first = export_filter_audit(_RunConnection(), "RUN-1", ("point_event_count", "efficiency", "source_pnl", "close_support"), tmp_path / "one.xlsx")
    second = export_filter_audit(_RunConnection(), "RUN-1", ("source_pnl", "efficiency", "close_support", "point_event_count"), tmp_path / "two.xlsx")

    assert first.suffix == second.suffix == ".xlsx"
    assert first.read_bytes() == second.read_bytes()
    assert load_workbook(first, read_only=True).sheetnames == [
        "Summary",
        "READY_AFTER_FILTERS",
        "Source PnL",
        "PnL-DD",
        "CloseSupport",
        "PointEventCount",
        "DEFERRED_COMBINED",
    ]


def test_export_filter_audit_flattens_real_engine_order_vectors(tmp_path: Path) -> None:
    from mrs3.analysis_filter_export import export_filter_audit
    from tests.test_analysis_shortlist import _candidate, _connection, _order

    connection = _connection(
        _candidate("C_A", "STR_A", [_order("P1", 10, 5, 0.9)]),
        _candidate("C_B", "STR_B", [_order("P2", 9, 4, 0.8)]),
        events={"P1": ("e1",), "P2": ("e1",)},
    )
    try:
        output = export_filter_audit(
            connection, "R1", ("source_pnl", "efficiency", "close_support"),
            tmp_path / "real.xlsx",
        )
        sheet = load_workbook(output, data_only=True)["DEFERRED_COMBINED"]
        headers = [cell.value for cell in sheet[1]]
        values = dict(zip(headers, [cell.value for cell in sheet[2]], strict=True))

        assert values["order1_source_pnl_a"] == "10"
        assert values["order1_source_pnl_b"] == "9"
        assert values["order1_pnl_dd_a"] == "5"
        assert values["order1_close_support_b"] == "0.8"
        assert values["order1_point_event_count_a"] == 1
        summary = dict(
            sheet_row[:2]
            for sheet_row in load_workbook(output, data_only=True)["Summary"].iter_rows(
                min_row=2, values_only=True
            )
        )
        assert summary["algorithm_version"] == "v1"
    finally:
        connection.close()
