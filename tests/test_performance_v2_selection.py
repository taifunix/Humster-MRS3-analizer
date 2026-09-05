from __future__ import annotations

import json
from datetime import datetime, timedelta, timezone
from decimal import Decimal
from pathlib import Path

import duckdb
import pandas as pd
import pytest
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import mrs3.performance_v2_selection as selection_module

from mrs3.performance_v2_selection import (
    PerformanceV2SelectionError,
    SelectionConfig,
    _consistency_summary,
    _consistency_windows,
    _ab_metrics_from_windows,
    _holding_p95_minutes,
    _return_30d,
    _selection_windows,
    _trade_rate_30d,
    load_selection_candidates,
    load_selection_config,
    parse_selection_request,
    prepare_selection_window_cache,
    run_selection,
    selection_cache_missing_strategy_ids,
    selection_cache_status,
    write_selection_workbook,
)
from mrs3.performance_v2_store import initialize_performance_v2
from mrs3.performance_v2_windows import METRICS_VERSION, WindowMetrics


UTC = timezone.utc


@pytest.mark.parametrize(
    ("duration", "window_count"),
    [
        (timedelta(days=20, hours=23, minutes=59, seconds=59), 0),
        (timedelta(days=21), 3),
        (timedelta(days=27, hours=23, minutes=59, seconds=59), 3),
        (timedelta(days=28), 4),
        (timedelta(days=45), 4),
    ],
)
def test_consistency_windows_have_exact_fractional_calendar_boundaries(duration: timedelta, window_count: int) -> None:
    start = datetime(2026, 1, 1, tzinfo=UTC)
    end = start + duration

    windows = _consistency_windows(start, end)

    assert len(windows) == window_count
    if windows:
        assert windows[0][0] == start
        assert windows[-1][1] == end
        assert all(left[1] == right[0] for left, right in zip(windows, windows[1:]))


def test_three_consistency_windows_are_requested_without_q4_or_positional_tail() -> None:
    start = datetime(2026, 1, 1, tzinfo=UTC)
    end = start + timedelta(days=21)

    windows = _selection_windows(start, end, SelectionConfig())

    assert all(window in windows for window in _consistency_windows(start, end))
    assert not any(window[0] == start + timedelta(days=15.75) for window in windows)
    assert windows[-1][1] == end


def _consistency_metric(start: datetime, end: datetime, growth: str = "1.1") -> WindowMetrics:
    return WindowMetrics(
        1, start, end, METRICS_VERSION, start, end, "AVAILABLE", None,
        Decimal(growth), None, None, None, Decimal("2"), None,
        None, None, 1, Decimal("100"),
    )


def test_time_consistency_status_counts_positive_windows_and_handles_unavailable() -> None:
    start = datetime(2026, 1, 1, tzinfo=UTC)
    four = _consistency_windows(start, start + timedelta(days=28))
    three = _consistency_windows(start, start + timedelta(days=21))

    assert _consistency_summary([_consistency_metric(*window) for window in four[:3]] + [_consistency_metric(*four[3], "1")], four) == (3, 4, "PASS")
    assert _consistency_summary([_consistency_metric(*window) for window in four[:2]] + [_consistency_metric(*four[2], "1"), _consistency_metric(*four[3], "1")], four) == (2, 4, "FAIL")
    assert _consistency_summary([_consistency_metric(*window) for window in three[:2]] + [_consistency_metric(*three[2], "1")], three) == (2, 3, "PASS")
    assert _consistency_summary([_consistency_metric(*window) for window in three[:1]] + [_consistency_metric(*three[1], "1"), _consistency_metric(*three[2], "1")], three) == (1, 3, "FAIL")
    no_trades = WindowMetrics.unavailable(1, *four[3], "NO_TRADES")
    collapsed = WindowMetrics.unavailable(1, *four[3], "COLLAPSED")
    assert _consistency_summary([], ()) == (None, None, "UNAVAILABLE")
    assert _consistency_summary([_consistency_metric(*window) for window in four[:3]] + [no_trades], four) == (3, 3, "PASS")
    assert _consistency_summary([_consistency_metric(*window) for window in four[:2]] + [no_trades, no_trades], four) == (2, 2, "FAIL")
    assert _consistency_summary([no_trades] * 4, four) == (None, 0, "UNAVAILABLE")
    assert _consistency_summary([_consistency_metric(*window) for window in four[:3]] + [collapsed], four) == (None, None, "UNAVAILABLE")
    assert _consistency_summary([_consistency_metric(*window) for window in four[:3]] + [None], four) == (None, None, "UNAVAILABLE")


def test_consistency_summary_passes_each_subwindow_bounds_to_normalization(monkeypatch) -> None:
    start = datetime(2026, 1, 1, tzinfo=UTC)
    windows = _consistency_windows(start, start + timedelta(days=21))
    metrics = [_consistency_metric(*window) for window in windows]
    calls = []
    monkeypatch.setattr(selection_module, "_return_30d", lambda _metric, window_start, window_end: calls.append((window_start, window_end)) or Decimal("1"))

    assert _consistency_summary(metrics, windows) == (3, 3, "PASS")
    assert calls == list(windows)


def test_low_trades_filter_uses_calendar_rate_and_excludes_missing_rates() -> None:
    request = parse_selection_request({"symbol": "BTCUSDT", "side": "LONG", "stages": [
        {"id": "filter_low_trades", "enabled": True, "scope": "pair_side"},
    ]})
    result = run_selection(pd.DataFrame([
        _selection_row("a", strategy_id=1, total_trades=1, trades_30d=10),
        _selection_row("b", strategy_id=2, total_trades=1000, trades_30d=10),
        _selection_row("c", strategy_id=3, total_trades=2, trades_30d=10),
        _selection_row("missing", strategy_id=4, total_trades=0, trades_30d=None),
    ]), request).set_index("strategy_name")

    assert result["eliminated_by_filter_low_trades"].sum() == 0
    assert result.loc["missing", "finalist"]


def test_unavailable_time_consistency_survives_and_exports_na(tmp_path: Path) -> None:
    request = parse_selection_request({"symbol": "BTCUSDT", "side": "LONG", "stages": [
        {"id": "filter_time_consistency", "enabled": True, "scope": "pair_side_timeframe"},
    ]})
    result = run_selection(pd.DataFrame([
        _selection_row("unavailable", strategy_id=1, positive_quarter_count=3, positive_quarter_available_count=4, positive_quarter_status="UNAVAILABLE"),
        _selection_row("fail", strategy_id=2, positive_quarter_count=1, positive_quarter_available_count=4, positive_quarter_status="FAIL"),
    ]), request)
    result = result.set_index("strategy_name")

    assert result.loc["unavailable", "finalist"]
    assert result.loc["fail", "eliminated_by_filter_time_consistency"]
    book = load_workbook(write_selection_workbook(result.reset_index(), tmp_path / "consistency.xlsx", request), data_only=True)
    headers = [cell.value for cell in book["All candidates"][1]]
    assert "positive_quarter_status" not in headers
    data_rows = {
        book["All candidates"].cell(row, headers.index("Стратегия") + 1).value: row
        for row in range(2, book["All candidates"].max_row + 1)
    }
    unavailable_row = data_rows["unavailable"]
    assert book["All candidates"].cell(unavailable_row, headers.index("Positive windows") + 1).value == "N/A"
    assert book["All candidates"].cell(unavailable_row, headers.index("eliminated_by_filter_time_consistency") + 1).value == "N/A"


def test_unavailable_status_exports_na_without_count_fields(tmp_path: Path) -> None:
    request = parse_selection_request({"symbol": "BTCUSDT", "side": "LONG", "stages": []})
    result = run_selection(pd.DataFrame([_selection_row("unavailable", positive_quarter_status="UNAVAILABLE")]), request)
    result = result.drop(columns=["positive_quarter_count", "positive_quarter_available_count"], errors="ignore")

    book = load_workbook(write_selection_workbook(result, tmp_path / "status-only.xlsx", request), data_only=True)
    headers = [cell.value for cell in book["All candidates"][1]]
    assert book["All candidates"].cell(2, headers.index("Positive windows") + 1).value == "N/A"


def test_v21_cache_rows_are_stale_for_v22_selection_readiness(tmp_path: Path) -> None:
    connection = _candidate_db(tmp_path)
    database = tmp_path / "strategy_performance.duckdb"
    connection.close()
    request = parse_selection_request({"symbol": "BTCUSDT", "side": "LONG", "stages": []})
    prepare_selection_window_cache(database, request, SelectionConfig(), workers=1)
    with duckdb.connect(str(database)) as check:
        check.execute("update window_metrics set metrics_version = 'performance-window-v2.1'")
        assert METRICS_VERSION == "performance-window-v2.2"
        assert selection_cache_status(check, request, SelectionConfig()) == {"total": 1, "missing": 1, "ready": False}


def test_selection_30d_rates_use_calendar_window_with_sparse_events() -> None:
    start = datetime(2026, 1, 1, tzinfo=UTC)
    metrics = WindowMetrics(
        1, start, start + timedelta(days=14), "test",
        start, start + timedelta(days=2), "AVAILABLE", None,
        Decimal("1.1"), Decimal("10"), None, None, Decimal("2"), Decimal("5"),
        Decimal("0.1"), Decimal("2"), 5, Decimal("60"),
    )

    assert _return_30d(metrics).quantize(Decimal(".0001")) == Decimal("22.6588")
    assert _trade_rate_30d(metrics).quantize(Decimal(".0001")) == Decimal("10.7143")


def test_selection_ab_metrics_clamp_duration_to_report_bounds() -> None:
    start = datetime(2026, 1, 1, tzinfo=UTC)
    metrics = WindowMetrics(
        1, start, start + timedelta(days=14), "test",
        start, start + timedelta(days=2), "AVAILABLE", None,
        Decimal("1.1"), Decimal("10"), None, None, Decimal("2"), Decimal("5"),
        Decimal("0.1"), Decimal("2"), 5, Decimal("60"),
    )

    values = _ab_metrics_from_windows(
        metrics, metrics, report_start_utc=start + timedelta(days=3), report_end_utc=start + timedelta(days=10)
    )

    assert values["ab_calendar_days_a"] == Decimal("7")
    assert values["ab_calendar_days_b"] == Decimal("7")
    assert values["ab_trade_rate_a_30d"] == Decimal(5) * 30 / 7


def _config(path: Path, **selection: object) -> Path:
    path.write_text(
        json.dumps({
            "unified_performance_v2": {
                "database_root": "data/performance-v2",
                "finalist_selection": selection,
            },
        }),
        encoding="utf-8",
    )
    return path


def test_parse_selection_request_accepts_all_known_stages_in_order() -> None:
    stages = [
        "filter_holding_outlier", "filter_low_trades", "filter_min_shift", "ab_deterioration", "pareto_window_b", "pareto_window_b_dd_shift", "pareto_dd5_balanced",
        "pareto_plateau_points_per_order", "pareto_plateau_points_total", "pareto_efficiency_shift",
        "pareto_dd5_holding", "pareto_dd5_close_ma", "pareto_dd5_first_shift",
        "pareto_conditional_close_ma", "pareto_primary", "pareto_dd5_capital", "filter_lot_variant_redundancy",
    ]

    request = parse_selection_request({
        "symbol": "BTCUSDT",
        "side": "LONG",
        "stages": [
            {"id": stage_id, "enabled": index < 5,
             "scope": "pair_side" if index < 4 else "pair_side_timeframe",
             **({"min_shift_pct": "0.3"} if stage_id == "filter_min_shift" else {})}
            for index, stage_id in enumerate(stages)
        ],
    })

    assert request.symbol == "BTCUSDT"
    assert request.side == "LONG"
    assert [stage.id for stage in request.stages] == stages


def test_parse_selection_request_accepts_new_stages_and_requires_last_fixed_ranker() -> None:
    request = parse_selection_request({"symbol": "BTCUSDT", "side": "LONG", "stages": [
        {"id": "pareto_robust", "enabled": True, "scope": "pair_side_timeframe"},
        {"id": "pareto_shift_near_tie", "enabled": True, "scope": "pair_side_timeframe", "pnl_tolerance_pct": "10"},
        {"id": "rank_robust_top_n", "enabled": True, "scope": "pair_side", "top_n": 50},
    ]})

    assert request.stages[-1].top_n == 50
    assert request.stages[-2].pnl_tolerance_pct == Decimal("10")
    with pytest.raises(PerformanceV2SelectionError, match="INVALID_CONFIG_top_n"):
        parse_selection_request({"symbol": "BTCUSDT", "side": "LONG", "stages": [
            {"id": "rank_robust_top_n", "enabled": True, "scope": "pair_side", "top_n": 0},
        ]})
    with pytest.raises(PerformanceV2SelectionError, match="RANK_STAGE_MUST_BE_LAST"):
        parse_selection_request({"symbol": "BTCUSDT", "side": "LONG", "stages": [
            {"id": "rank_robust_top_n", "enabled": True, "scope": "pair_side", "top_n": 50},
            {"id": "pareto_robust", "enabled": True, "scope": "pair_side_timeframe"},
        ]})
    with pytest.raises(PerformanceV2SelectionError, match="DUPLICATE_STAGE"):
        parse_selection_request({"symbol": "BTCUSDT", "side": "LONG", "stages": [
            {"id": "rank_robust_top_n", "enabled": True, "scope": "pair_side", "top_n": 50},
            {"id": "rank_robust_top_n", "enabled": True, "scope": "pair_side", "top_n": 10},
        ]})


@pytest.mark.parametrize(
    ("payload", "code"),
    [
        ({"symbol": "BTCUSDT", "side": "LONG", "stages": [{"id": "unknown", "enabled": True, "scope": "pair_side"}]}, "UNKNOWN_STAGE"),
        ({"symbol": "BTCUSDT", "side": "LONG", "stages": [{"id": "ab_deterioration", "enabled": True, "scope": "pair_side"}, {"id": "ab_deterioration", "enabled": False, "scope": "pair_side"}]}, "DUPLICATE_STAGE"),
        ({"symbol": "BTCUSDT", "side": "LONG", "stages": [{"id": "ab_deterioration", "enabled": True, "scope": "global"}]}, "INVALID_SCOPE"),
        ({"symbol": "BTCUSDT", "side": "LONG", "stages": [{"id": "filter_lot_variant_redundancy", "enabled": True, "scope": "pair_side"}]}, "LOT_VARIANT_STAGE_SCOPE"),
    ],
)
def test_parse_selection_request_rejects_unknown_duplicate_and_invalid_scope(payload: dict[str, object], code: str) -> None:
    with pytest.raises(PerformanceV2SelectionError, match=code):
        parse_selection_request(payload)


def test_min_shift_filter_excludes_any_existing_order_below_threshold() -> None:
    request = parse_selection_request({"symbol": "BTCUSDT", "side": "LONG", "stages": [
        {"id": "filter_min_shift", "enabled": True, "scope": "pair_side", "min_shift_pct": "0.3"},
    ]})
    result = run_selection(pd.DataFrame([
        _selection_row("kept", strategy_id=1, order_count=2, order_1_shift_bp=30, order_2_shift_bp=40),
        _selection_row("excluded", strategy_id=2, order_count=2, order_1_shift_bp=40, order_2_shift_bp=20),
        _selection_row("missing", strategy_id=3, order_count=2, order_1_shift_bp=40),
    ]), request).set_index("strategy_name")

    assert not result.loc["kept", "eliminated_by_filter_min_shift"]
    assert result.loc["excluded", "eliminated_by_filter_min_shift"]
    assert not result.loc["missing", "eliminated_by_filter_min_shift"]


def test_window_b_pareto_eliminates_only_candidate_dominated_on_all_b_metrics() -> None:
    request = parse_selection_request({"symbol": "BTCUSDT", "side": "LONG", "stages": [
        {"id": "pareto_window_b", "enabled": True, "scope": "pair_side_timeframe"},
    ]})
    result = run_selection(pd.DataFrame([
        _selection_row("winner", ab_return_b_30d_pct=Decimal("20"), ab_trade_rate_b_30d=Decimal("30"), ab_drawdown_b_pct=Decimal("4"), ab_holding_p95_minutes=Decimal("60")),
        _selection_row("loser", ab_return_b_30d_pct=Decimal("10"), ab_trade_rate_b_30d=Decimal("20"), ab_drawdown_b_pct=Decimal("5"), ab_holding_p95_minutes=Decimal("90")),
        _selection_row("missing", ab_return_b_30d_pct=Decimal("5"), ab_trade_rate_b_30d=Decimal("10"), ab_drawdown_b_pct=Decimal("6")),
    ]), request).set_index("strategy_name")

    assert not result.loc["winner", "eliminated_by_pareto_window_b"]
    assert result.loc["loser", "eliminated_by_pareto_window_b"]
    assert not result.loc["missing", "eliminated_by_pareto_window_b"]


def test_window_b_dd_shift_pareto_eliminates_only_fully_dominated_candidate() -> None:
    request = parse_selection_request({"symbol": "BTCUSDT", "side": "LONG", "stages": [
        {"id": "pareto_window_b_dd_shift", "enabled": True, "scope": "pair_side_timeframe"},
    ]})
    result = run_selection(pd.DataFrame([
        _selection_row("winner", ab_return_b_30d_pct=Decimal("20"), max_drawdown_pct=Decimal("4"), first_shift_bp=200),
        _selection_row("loser", ab_return_b_30d_pct=Decimal("10"), max_drawdown_pct=Decimal("5"), first_shift_bp=100),
        _selection_row("tradeoff", ab_return_b_30d_pct=Decimal("25"), max_drawdown_pct=Decimal("5"), first_shift_bp=100),
    ]), request).set_index("strategy_name")

    assert not result.loc["winner", "eliminated_by_pareto_window_b_dd_shift"]
    assert result.loc["loser", "eliminated_by_pareto_window_b_dd_shift"]
    assert not result.loc["tradeoff", "eliminated_by_pareto_window_b_dd_shift"]


def test_selection_config_reads_agreed_defaults(tmp_path: Path) -> None:
    config = load_selection_config(_config(tmp_path / "config.performance.json"))

    assert config.ab_final_days == 14
    assert config.ab_return_floor_pct == 5
    assert config.ab_return_divisor == 10
    assert config.ab_win_rate_floor_pct == 58
    assert config.ab_trade_rate_divisor == 7
    assert config.plateau_points_pareto_pnl_multiplier == 2
    assert config.best_trade_max_profit_share_pct == 35
    assert config.best_trade_min_profitable_trades == 4
    assert config.shift_near_tie_min_advantage_bp == 10
    assert config.lot_variant_redundancy_enabled is True


def test_selection_config_reads_explicit_overrides(tmp_path: Path) -> None:
    config = load_selection_config(_config(
        tmp_path / "config.performance.json",
        ab_final_days=21,
        ab_return_floor_pct=4.5,
        ab_return_divisor=8,
        ab_win_rate_floor_pct=60,
        ab_trade_rate_divisor=6,
        plateau_points_pareto_pnl_multiplier=1.5,
        best_trade_max_profit_share_pct=40,
        best_trade_min_profitable_trades=5,
        shift_near_tie_min_advantage_bp=15,
        lot_variant_redundancy_enabled=False,
    ))

    assert config.ab_final_days == 21
    assert config.ab_return_floor_pct == Decimal("4.5")
    assert config.plateau_points_pareto_pnl_multiplier == Decimal("1.5")
    assert config.best_trade_max_profit_share_pct == 40
    assert config.best_trade_min_profitable_trades == 5
    assert config.shift_near_tie_min_advantage_bp == 15
    assert config.lot_variant_redundancy_enabled is False


@pytest.mark.parametrize(
    ("field", "value"),
    [
        ("ab_final_days", 0),
        ("ab_final_days", True),
        ("ab_return_floor_pct", 0),
        ("ab_return_divisor", "wrong"),
        ("ab_win_rate_floor_pct", float("nan")),
        ("ab_trade_rate_divisor", float("inf")),
        ("plateau_points_pareto_pnl_multiplier", False),
        ("best_trade_max_profit_share_pct", 100),
        ("best_trade_min_profitable_trades", 0),
        ("shift_near_tie_min_advantage_bp", 0),
        ("lot_variant_redundancy_enabled", "yes"),
    ],
)
def test_selection_config_rejects_invalid_values(tmp_path: Path, field: str, value: object) -> None:
    with pytest.raises(PerformanceV2SelectionError, match=f"INVALID_CONFIG_{field}"):
        load_selection_config(_config(tmp_path / "config.performance.json", **{field: value}))


def test_selection_config_rejects_malformed_v2_namespace(tmp_path: Path) -> None:
    path = tmp_path / "config.performance.json"
    path.write_text(json.dumps({"unified_performance_v2": []}), encoding="utf-8")

    with pytest.raises(PerformanceV2SelectionError, match="INVALID_CONFIG"):
        load_selection_config(path)


def test_real_selection_config_has_agreed_defaults() -> None:
    config = load_selection_config(Path(__file__).resolve().parents[1] / "config.performance.json")

    assert config == SelectionConfig()


def _candidate_db(tmp_path: Path) -> duckdb.DuckDBPyConnection:
    connection = duckdb.connect(str(tmp_path / "strategy_performance.duckdb"))
    initialize_performance_v2(connection)
    start = datetime(2026, 1, 1, tzinfo=UTC)
    strategy_id = connection.execute(
        """insert into strategies (strategy_name, symbol, side, timeframe, close_ma_len,
           order_count, analysis_run_id, candidate_identity, lifecycle_status,
           created_at_utc, updated_at_utc) values ('alpha', 'BTCUSDT', 'LONG', '1h',
           3, 1, 'run', 'candidate', 'ACTIVE', ?, ?) returning strategy_id""",
        [start, start],
    ).fetchone()[0]
    result_id = connection.execute(
        """insert into strategy_results (strategy_id, report_start_utc, report_end_utc, exchange,
           commission_rate, initial_balance, final_balance, total_pnl, total_pnl_pct,
           max_drawdown, max_drawdown_pct, total_fees, total_trades, imported_at_utc)
           values (?, ?, ?, 'Bybit', .0004, 100, 110, 10, 10, 5, 5, 2, 2, ?) returning result_id""",
        [strategy_id, start, datetime(2026, 1, 31, tzinfo=UTC), start],
    ).fetchone()[0]
    connection.execute("update strategies set current_result_id = ? where strategy_id = ?", [result_id, strategy_id])
    connection.execute("insert into analysis_plateaus values ('run', 'P1', 12, 34)")
    connection.execute(
        """insert into strategy_orders values (?, 1, 7, .995, 125, 1, 'run', 'P1', 8)""",
        [strategy_id],
    )
    connection.executemany(
        "insert into strategy_actions values (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
        [
            (result_id, 0, start, "BTCUSDT", 1, "closed", 1, 0, "", 0, 0, 100, None),
            (result_id, 1, datetime(2026, 1, 2, tzinfo=UTC), "BTCUSDT", 1, "opened", 1, 1, "long", 0, 0, 100, None),
            (result_id, 2, datetime(2026, 1, 3, tzinfo=UTC), "BTCUSDT", 1, "closed", 1, 0, "", 10, 2, 110, None),
        ],
    )
    connection.executemany(
        "insert into strategy_equity values (?, ?, ?, ?, ?)",
        [
            (result_id, 0, start, 100, 100),
            (result_id, 1, datetime(2026, 1, 2, tzinfo=UTC), 100, 100),
            (result_id, 2, datetime(2026, 1, 3, tzinfo=UTC), 110, 110),
            (result_id, 3, datetime(2026, 1, 31, tzinfo=UTC), 110, 110),
        ],
    )
    return connection


def test_loader_derives_proxy_holding_and_order_plateau_counts(tmp_path: Path) -> None:
    connection = _candidate_db(tmp_path)
    request = parse_selection_request({"symbol": "BTCUSDT", "side": "LONG", "stages": []})
    try:
        result_id = connection.execute("select result_id from strategy_results").fetchone()[0]
        connection.execute(
            "insert into strategy_actions values (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
            [result_id, 3, datetime(2026, 1, 3, 12, tzinfo=UTC), "BTCUSDT", 1, "fee", 1, 0, "", -2, 0, 108, None],
        )
        row = load_selection_candidates(connection, request).iloc[0]
    finally:
        connection.close()

    assert row["order_1_plateau_point_count"] == 12
    assert row["order_1_plateau_key"] == ("run", "P1")
    assert row["total_plateau_point_count"] == 12
    assert row["total_trades"] == 1
    assert row["dd5_proxy"] is not None
    assert row["dd5_proxy"] > 0
    assert row["holding_p95_minutes"] == 1440
    assert row["holding_median_minutes"] == 1440
    assert row["ab_pnl_change_30d_pct"] is None
    assert row["trades_30d"] == Decimal("1")
    assert row["total_pnl_pct"] == Decimal("10")
    assert row["positive_quarter_status"] == "UNAVAILABLE"
    assert pd.isna(row["positive_quarter_count"])
    assert pd.isna(row["positive_quarter_available_count"])
    assert row["best_trade_profit_share_pct"] == 100
    assert row["pnl_without_best_trade"] == 0
    assert row["pnl_without_best_trade_pct"] == 0
    assert row["completed_profitable_trade_count"] == 1
    assert row["best_trade_reliable"]


def test_signed_short_actions_feed_holding_and_best_trade_metrics(tmp_path: Path) -> None:
    connection = _candidate_db(tmp_path)
    result_id = connection.execute("select result_id from strategy_results").fetchone()[0]
    connection.execute("update strategies set side = 'SHORT'")
    connection.execute(
        "update strategy_actions set size = -1, post_size = -1, post_side = 'short' "
        "where result_id = ? and action = 'opened'",
        [result_id],
    )
    connection.execute(
        "update strategy_actions set post_size = 0, post_side = '' where result_id = ? and action = 'closed'",
        [result_id],
    )
    connection.executemany(
        "insert into strategy_actions values (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
        [
            (result_id, 3, datetime(2026, 1, 25, tzinfo=UTC), "BTCUSDT", 1, "opened", -1, -1, "short", 0, 0, 110, None),
            (result_id, 4, datetime(2026, 1, 26, tzinfo=UTC), "BTCUSDT", 1, "closed", 1, 0, "", 5, 0, 115, None),
        ],
    )
    try:
        row = load_selection_candidates(
            connection, parse_selection_request({"symbol": "BTCUSDT", "side": "SHORT", "stages": []})
        ).iloc[0]
    finally:
        connection.close()

    assert row["holding_p95_minutes"] == Decimal("1440")
    assert row["ab_holding_p95_minutes"] == Decimal("1440")
    assert row["best_trade_reliable"]
    assert row["completed_profitable_trade_count"] == 2
    assert row["pnl_without_best_trade"] == 5


def test_best_trade_facts_with_no_profitable_trip_are_not_evaluable(tmp_path: Path) -> None:
    connection = _candidate_db(tmp_path)
    request = parse_selection_request({"symbol": "BTCUSDT", "side": "LONG", "stages": []})
    try:
        connection.execute("update strategy_actions set pnl = -10 where action = 'closed'")
        row = load_selection_candidates(connection, request).iloc[0]
    finally:
        connection.close()

    assert row["best_trade_profit_share_pct"] is None
    assert row["pnl_without_best_trade"] is None
    assert row["completed_profitable_trade_count"] is None


def test_loader_exports_pnl_without_best_as_pct_of_initial_balance(tmp_path: Path) -> None:
    connection = _candidate_db(tmp_path)
    request = parse_selection_request({"symbol": "BTCUSDT", "side": "LONG", "stages": []})
    try:
        result_id = connection.execute("select result_id from strategy_results").fetchone()[0]
        connection.executemany(
            "insert into strategy_actions values (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
            [
                (result_id, 3, datetime(2026, 1, 4, tzinfo=UTC), "BTCUSDT", 1, "opened", 1, 1, "long", 0, 0, 110, None),
                (result_id, 4, datetime(2026, 1, 5, tzinfo=UTC), "BTCUSDT", 1, "closed", 1, 0, "", 5, 1, 115, None),
            ],
        )
        row = load_selection_candidates(connection, request).iloc[0]
    finally:
        connection.close()

    assert row["pnl_without_best_trade"] == 5
    assert row["pnl_without_best_trade_pct"] == 5


def test_holding_p95_uses_all_closed_positions(tmp_path: Path) -> None:
    connection = _candidate_db(tmp_path)
    request = parse_selection_request({"symbol": "BTCUSDT", "side": "LONG", "stages": []})
    result_id = connection.execute("select result_id from strategy_results").fetchone()[0]
    connection.executemany(
        "insert into strategy_actions values (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
        [
            (result_id, 3, datetime(2026, 1, 5, tzinfo=UTC), "BTCUSDT", 1, "opened", 1, 1, "long", 0, 0, 110, None),
            (result_id, 4, datetime(2026, 1, 7, tzinfo=UTC), "BTCUSDT", 1, "closed", 1, 0, "", 0, 0, 110, None),
            (result_id, 5, datetime(2026, 1, 8, tzinfo=UTC), "BTCUSDT", 1, "opened", 1, 1, "long", 0, 0, 110, None),
            (result_id, 6, datetime(2026, 1, 12, tzinfo=UTC), "BTCUSDT", 1, "closed", 1, 0, "", 0, 0, 110, None),
        ],
    )
    try:
        p95 = _holding_p95_minutes(connection, request)
    finally:
        connection.close()

    assert p95[result_id] == Decimal("5472.0")


def test_parallel_window_warmup_persists_default_selection_windows(tmp_path: Path) -> None:
    connection = _candidate_db(tmp_path)
    database = tmp_path / "strategy_performance.duckdb"
    connection.close()
    request = parse_selection_request({"symbol": "BTCUSDT", "side": "LONG", "stages": []})

    prepare_selection_window_cache(database, request, SelectionConfig(), workers=1)
    prepare_selection_window_cache(database, request, SelectionConfig(), workers=1)

    with duckdb.connect(str(database), read_only=True) as check:
        assert check.execute("select count(*) from window_metrics").fetchone() == (7,)


def test_missing_cache_strategy_ids_only_returns_current_results_without_facts(tmp_path: Path) -> None:
    connection = _candidate_db(tmp_path)
    database = tmp_path / "strategy_performance.duckdb"
    request = parse_selection_request({"symbol": "BTCUSDT", "side": "LONG", "stages": []})
    config = SelectionConfig()
    connection.close()
    prepare_selection_window_cache(database, request, config, workers=1)

    with duckdb.connect(str(database)) as check:
        start = datetime(2026, 1, 1, tzinfo=UTC)
        strategy_id = check.execute(
            """insert into strategies (strategy_name, symbol, side, timeframe, close_ma_len,
               order_count, analysis_run_id, candidate_identity, lifecycle_status,
               created_at_utc, updated_at_utc) values ('beta', 'BTCUSDT', 'LONG', '1h',
               3, 1, 'run', 'candidate-beta', 'ACTIVE', ?, ?) returning strategy_id""",
            [start, start],
        ).fetchone()[0]
        result_id = check.execute(
            """insert into strategy_results (strategy_id, report_start_utc, report_end_utc, exchange,
               commission_rate, initial_balance, final_balance, total_pnl, total_pnl_pct,
               max_drawdown, max_drawdown_pct, total_fees, total_trades, imported_at_utc)
               values (?, ?, ?, 'Bybit', .0004, 100, 110, 10, 10, 5, 5, 2, 2, ?) returning result_id""",
            [strategy_id, start, datetime(2026, 1, 31, tzinfo=UTC), start],
        ).fetchone()[0]
        check.execute("update strategies set current_result_id = ? where strategy_id = ?", [result_id, strategy_id])

        assert selection_cache_missing_strategy_ids(check, request, config) == (strategy_id,)


def test_legacy_full_ab_only_cache_is_not_ready(tmp_path: Path) -> None:
    connection = _candidate_db(tmp_path)
    database = tmp_path / "strategy_performance.duckdb"
    connection.close()
    request = parse_selection_request({"symbol": "BTCUSDT", "side": "LONG", "stages": []})
    config = SelectionConfig()
    prepare_selection_window_cache(database, request, config, workers=1)
    with duckdb.connect(str(database)) as check:
        result_id, start, end = check.execute("select result_id, report_start_utc, report_end_utc from strategy_results").fetchone()
        split = end - timedelta(days=config.ab_final_days)
        check.execute(
            "delete from window_metrics where result_id = ? and (requested_start_utc, requested_end_utc) not in ((?, ?), (?, ?), (?, ?))",
            [result_id, start, end, start, split, split, end],
        )
        assert selection_cache_status(check, request, config) == {"total": 1, "missing": 1, "ready": False}


def test_loader_leaves_incomplete_order_and_empty_candidate_facts_blank(tmp_path: Path) -> None:
    connection = _candidate_db(tmp_path)
    request = parse_selection_request({"symbol": "BTCUSDT", "side": "LONG", "stages": []})
    try:
        connection.execute("update strategies set order_count = 2")
        connection.execute("update strategy_results set max_drawdown_pct = 10")
        row = load_selection_candidates(connection, request).iloc[0]
        now = datetime(2026, 2, 1, tzinfo=UTC)
        no_orders_id = connection.execute(
            """insert into strategies (strategy_name, symbol, side, timeframe, close_ma_len,
               order_count, analysis_run_id, candidate_identity, lifecycle_status,
               created_at_utc, updated_at_utc) values ('without-orders', 'BTCUSDT', 'LONG', '1h',
               3, 1, 'run', 'candidate-2', 'ACTIVE', ?, ?) returning strategy_id""",
            [now, now],
        ).fetchone()[0]
        no_orders_result = connection.execute(
            """insert into strategy_results (strategy_id, report_start_utc, report_end_utc, exchange,
               commission_rate, initial_balance, final_balance, total_pnl, total_pnl_pct,
               max_drawdown, max_drawdown_pct, total_fees, total_trades, imported_at_utc)
               values (?, ?, ?, 'Bybit', .0004, 100, 100, 0, 0, 0, 0, 0, 0, ?) returning result_id""",
            [no_orders_id, now, now, now],
        ).fetchone()[0]
        connection.execute(
            "update strategies set current_result_id = ? where strategy_id = ?", [no_orders_result, no_orders_id]
        )
        all_rows = load_selection_candidates(connection, request).set_index("strategy_name")
        connection.execute("update strategy_results set max_drawdown_pct = 0 where strategy_id = 1")
        zero_dd = load_selection_candidates(connection, request).set_index("strategy_name").loc["alpha"]
        empty = load_selection_candidates(
            connection, parse_selection_request({"symbol": "ETHUSDT", "side": "LONG", "stages": []})
        )
    finally:
        connection.close()

    assert row["risk_scale"] == Decimal("0.5")
    assert row["dd5_proxy"] is not None
    assert row["scaled_lot_sum"] is None
    assert row["capital_proxy"] is None
    assert row["capital_efficiency"] is None
    assert set(all_rows.index) == {"alpha", "without-orders"}
    assert pd.isna(all_rows.loc["without-orders", "order_1_plateau_point_count"])
    assert pd.isna(all_rows.loc["without-orders", "capital_proxy"])
    assert zero_dd["dd5_proxy"] is None
    assert zero_dd["scaled_lot_sum"] is None
    assert "strategy_id" in empty.columns
    assert empty.empty


def _selection_row(name: str, **values: object) -> dict[str, object]:
    return {
        "strategy_id": 1 if name == "winner" else 2,
        "strategy_name": name,
        "timeframe": "1h",
        "order_count": 1,
        "dd5_proxy": Decimal("10") if name == "winner" else Decimal("5"),
        "first_shift_bp": 200 if name == "winner" else 100,
        "order_1_shift_bp": 30 if name == "winner" else 270,
        "order_2_plateau_point_count": Decimal("7.6"),
        "order_3_plateau_point_count": Decimal("8.4"),
        "order_4_plateau_point_count": Decimal("9.5"),
        "order_1_open_ma_len": Decimal("3.6"),
        "order_2_open_ma_len": Decimal("4.4"),
        "order_3_open_ma_len": Decimal("5.5"),
        "order_4_open_ma_len": Decimal("6.1"),
        "order_1_lot_x": Decimal("0.25"),
        "order_2_lot_x": Decimal("0.50"),
        "order_3_lot_x": Decimal("0.75"),
        "order_4_lot_x": Decimal("1.00"),
        "capital_proxy": Decimal("1") if name == "winner" else Decimal("2"),
        "capital_efficiency": Decimal("10") if name == "winner" else Decimal("2.5"),
        "holding_p95_minutes": Decimal("10") if name == "winner" else Decimal("20"),
        "close_ma_len": 3 if name == "winner" else 5,
        "total_trades": 100,
        "order_1_plateau_point_count": 20 if name == "winner" else 10,
        "total_plateau_point_count": 20 if name == "winner" else 10,
        **values,
    }


def _lot_variant_row(
    name: str,
    strategy_id: int,
    lots: tuple[str, str] = ("1", "2"),
    interval: tuple[datetime, datetime] = (datetime(2026, 1, 1, tzinfo=UTC), datetime(2026, 2, 1, tzinfo=UTC)),
    **metrics: object,
) -> dict[str, object]:
    start, end = interval
    return {
        "strategy_id": strategy_id,
        "strategy_name": name,
        "symbol": "BTCUSDT",
        "side": "LONG",
        "timeframe": "1h",
        "close_ma_len": 20,
        "order_count": 2,
        "order_1_open_ma_len": 5,
        "order_1_shift_bp": 100,
        "order_1_lot_x": Decimal(lots[0]),
        "order_2_open_ma_len": 10,
        "order_2_shift_bp": 200,
        "order_2_lot_x": Decimal(lots[1]),
        "report_start_utc": start,
        "report_end_utc": end,
        "effective_start_utc": start,
        "effective_end_utc": end,
        "dd5_proxy": Decimal("10"),
        "capital_proxy": Decimal("5"),
        "robust_pnl_30d_pct": Decimal("8"),
        "worst_drawdown_pct": Decimal("6"),
        "profit_factor": Decimal("1.5"),
        **metrics,
    }


def test_lot_variant_filter_is_default_on_first_and_keeps_loser_auditable() -> None:
    request = parse_selection_request({"symbol": "BTCUSDT", "side": "LONG", "stages": [
        {"id": "filter_best_trade_dependency", "enabled": True, "scope": "pair_side_timeframe"},
        {"id": "filter_lot_variant_redundancy", "enabled": True, "scope": "pair_side_timeframe"},
    ]})
    result = run_selection(pd.DataFrame([
        _lot_variant_row("winner", 1, lots=("1", "2"), dd5_proxy=Decimal("11")),
        _lot_variant_row(
            "loser", 2, lots=("3", "4"), best_trade_reliable=True,
            completed_profitable_trade_count=4, pnl_without_best_trade=Decimal("0"),
            best_trade_profit_share_pct=Decimal("40"),
        ),
    ]), request).set_index("strategy_name")

    assert result.loc["winner", "finalist"]
    assert not result.loc["loser", "finalist"]
    assert result.loc["loser", "eliminated_by_filter_lot_variant_redundancy"]
    assert not result.loc["loser", "eliminated_by_filter_best_trade_dependency"]
    assert result.loc["loser", "auto_status"] == "FILTERED"
    assert result.loc["loser", "elimination_reason"] == "LOT_VARIANT_REDUNDANT"
    assert result.loc["winner", "lot_variant_representative_strategy_id"] == 1
    assert result.loc["loser", "lot_variant_representative_strategy_id"] == 1
    assert result.loc["winner", "lot_variant_group_key"] == result.loc["loser", "lot_variant_group_key"]


def test_lot_variant_filter_can_be_disabled_and_fails_closed() -> None:
    frame = pd.DataFrame([
        _lot_variant_row("missing-a", 1, lots=("1", "2"), profit_factor=None),
        _lot_variant_row("missing-b", 2, lots=("3", "4"), profit_factor=None),
    ])
    request = parse_selection_request({"symbol": "BTCUSDT", "side": "LONG", "stages": []})

    disabled = run_selection(frame, request, SelectionConfig(lot_variant_redundancy_enabled=False))
    assert disabled["finalist"].all()
    assert disabled["lot_variant_group_key"].isna().all()

    failed_closed = run_selection(frame, request)
    assert failed_closed["finalist"].all()
    assert failed_closed["lot_variant_group_key"].isna().all()

    malformed = pd.DataFrame([
        _lot_variant_row("bad-a", 1, lots=("1", "2"), effective_start_utc="not-a-date"),
        _lot_variant_row("bad-b", 2, lots=("3", "4")),
    ])
    malformed_result = run_selection(malformed, request)
    assert malformed_result["finalist"].all()
    assert malformed_result["lot_variant_group_key"].isna().all()


def test_lot_variant_filter_isolated_by_interval_and_canonicalizes_order_permutation() -> None:
    first = _lot_variant_row("first", 1, lots=("1", "2"))
    second = _lot_variant_row("second", 2, lots=("3", "4"))
    second["order_1_open_ma_len"], second["order_2_open_ma_len"] = second["order_2_open_ma_len"], second["order_1_open_ma_len"]
    second["order_1_shift_bp"], second["order_2_shift_bp"] = second["order_2_shift_bp"], second["order_1_shift_bp"]
    second["order_1_lot_x"], second["order_2_lot_x"] = second["order_2_lot_x"], second["order_1_lot_x"]
    same_interval = run_selection(pd.DataFrame([first, second]), parse_selection_request({"symbol": "BTCUSDT", "side": "LONG", "stages": []})).set_index("strategy_name")
    assert same_interval.loc["first", "finalist"]
    assert not same_interval.loc["second", "finalist"]

    later = _lot_variant_row(
        "later", 3, lots=("3", "4"),
        interval=(datetime(2026, 2, 1, tzinfo=UTC), datetime(2026, 3, 1, tzinfo=UTC)),
    )
    separate = run_selection(pd.DataFrame([first, later]), parse_selection_request({"symbol": "BTCUSDT", "side": "LONG", "stages": []}))
    assert separate["finalist"].all()


@pytest.mark.parametrize(
    ("metric", "a_value", "b_value", "expected"),
    [
        ("dd5_proxy", Decimal("11"), Decimal("10"), "a"),
        ("capital_proxy", Decimal("4"), Decimal("5"), "a"),
        ("robust_pnl_30d_pct", Decimal("9"), Decimal("8"), "a"),
        ("worst_drawdown_pct", Decimal("5"), Decimal("6"), "a"),
        ("profit_factor", Decimal("1.6"), Decimal("1.5"), "a"),
        ("strategy_id", None, None, "b"),
    ],
)
def test_lot_variant_filter_uses_declared_winner_order(metric: str, a_value: object, b_value: object, expected: str) -> None:
    a_id, b_id = (2, 1) if metric == "strategy_id" else (1, 2)
    a = _lot_variant_row("a", a_id, lots=("1", "2"), **({metric: a_value} if a_value is not None else {}))
    b = _lot_variant_row("b", b_id, lots=("3", "4"), **({metric: b_value} if b_value is not None else {}))
    result = run_selection(pd.DataFrame([a, b]), parse_selection_request({"symbol": "BTCUSDT", "side": "LONG", "stages": []}))
    finalists = result.loc[result["finalist"], "strategy_name"].tolist()
    assert finalists == [expected]


@pytest.mark.parametrize("stage_id", [
    "pareto_dd5_balanced", "pareto_plateau_points_per_order", "pareto_plateau_points_total",
    "pareto_efficiency_shift", "pareto_dd5_holding", "pareto_dd5_close_ma",
    "pareto_dd5_first_shift", "pareto_primary", "pareto_dd5_capital",
])
def test_pareto_stages_eliminate_dominated_candidate(stage_id: str) -> None:
    request = parse_selection_request({"symbol": "BTCUSDT", "side": "LONG", "stages": [
        {"id": stage_id, "enabled": True, "scope": "pair_side"},
    ]})

    result = run_selection(pd.DataFrame([_selection_row("winner"), _selection_row("loser")]), request)
    result = result.set_index("strategy_name")

    assert not result.loc["winner", f"eliminated_by_{stage_id}"]
    assert result.loc["loser", f"eliminated_by_{stage_id}"]
    assert not result.loc["loser", "finalist"]


def test_ab_insufficient_data_does_not_eliminate() -> None:
    request = parse_selection_request({"symbol": "BTCUSDT", "side": "LONG", "stages": [
        {"id": "ab_deterioration", "enabled": True, "scope": "pair_side"},
    ]})
    frame = pd.DataFrame([_selection_row("winner")])

    result = run_selection(frame, request)

    assert result.loc[0, "finalist"]
    assert result.loc[0, "elimination_reason"] == "AB_NOT_EVALUATED_INSUFFICIENT_DATA"


def test_new_robust_filters_only_eliminate_evaluable_or_dominated_candidates() -> None:
    request = parse_selection_request({"symbol": "BTCUSDT", "side": "LONG", "stages": [
        {"id": "filter_best_trade_dependency", "enabled": True, "scope": "pair_side_timeframe"},
        {"id": "filter_time_consistency", "enabled": True, "scope": "pair_side_timeframe"},
        {"id": "pareto_robust", "enabled": True, "scope": "pair_side_timeframe"},
    ]})
    result = run_selection(pd.DataFrame([
        _selection_row("winner", strategy_id=1, best_trade_reliable=True, completed_profitable_trade_count=4,
                       best_trade_profit_share_pct=Decimal("20"), pnl_without_best_trade=Decimal("5"),
                       positive_quarter_count=4, positive_quarter_available_count=4,
                       robust_pnl_30d_pct=Decimal("20"), worst_drawdown_pct=Decimal("4"),
                       worst_holding_p95_minutes=Decimal("40"), first_shift_bp=200),
        _selection_row("dependent", strategy_id=2, best_trade_reliable=True, completed_profitable_trade_count=4,
                       best_trade_profit_share_pct=Decimal("36"), pnl_without_best_trade=Decimal("5"),
                       positive_quarter_count=4, positive_quarter_available_count=4),
        _selection_row("boundary", strategy_id=5, best_trade_reliable=True, completed_profitable_trade_count=4,
                       best_trade_profit_share_pct=Decimal("35"), pnl_without_best_trade=Decimal("5"),
                       positive_quarter_count=1, positive_quarter_available_count=3),
        _selection_row("inconsistent", strategy_id=3, best_trade_reliable=False, completed_profitable_trade_count=1,
                       positive_quarter_count=2, positive_quarter_available_count=4),
        _selection_row("dominated", strategy_id=4, best_trade_reliable=False, completed_profitable_trade_count=1,
                       positive_quarter_count=4, positive_quarter_available_count=4,
                       robust_pnl_30d_pct=Decimal("10"), worst_drawdown_pct=Decimal("5"),
                       worst_holding_p95_minutes=Decimal("50"), first_shift_bp=100),
    ]), request).set_index("strategy_name")

    assert result.loc["dependent", "eliminated_by_filter_best_trade_dependency"]
    assert not result.loc["boundary", "eliminated_by_filter_best_trade_dependency"]
    assert not result.loc["boundary", "eliminated_by_filter_time_consistency"]
    assert result.loc["inconsistent", "eliminated_by_filter_time_consistency"]
    assert result.loc["dominated", "eliminated_by_pareto_robust"]
    assert result.loc["winner", "finalist"]


def test_shift_near_tie_and_final_rank_keep_top_rankable_and_unranked_rows() -> None:
    request = parse_selection_request({"symbol": "BTCUSDT", "side": "LONG", "stages": [
        {"id": "pareto_shift_near_tie", "enabled": True, "scope": "pair_side_timeframe", "pnl_tolerance_pct": "10"},
        {"id": "rank_robust_top_n", "enabled": True, "scope": "pair_side", "top_n": 1},
    ]})
    result = run_selection(pd.DataFrame([
        _selection_row("shift-winner", strategy_id=1, robust_pnl_30d_pct=Decimal("100"), worst_drawdown_pct=Decimal("4"),
                       worst_holding_p95_minutes=Decimal("40"), ab_stability_ratio=Decimal(".9"), minimum_plateau_point_count=20, first_shift_bp=200),
        _selection_row("near-tie", strategy_id=2, robust_pnl_30d_pct=Decimal("95"), worst_drawdown_pct=Decimal("5"),
                       worst_holding_p95_minutes=Decimal("50"), ab_stability_ratio=Decimal(".8"), minimum_plateau_point_count=10, first_shift_bp=100),
        _selection_row("unranked", strategy_id=3, robust_pnl_30d_pct=None, worst_drawdown_pct=None,
                       worst_holding_p95_minutes=None, ab_stability_ratio=None, minimum_plateau_point_count=None, first_shift_bp=None, close_ma_len=None),
    ]), request).set_index("strategy_name")

    assert result.loc["near-tie", "eliminated_by_pareto_shift_near_tie"]
    assert result.loc["shift-winner", "final_rank"] == 1
    assert not result.loc["unranked", "finalist"]
    assert result.loc["unranked", "auto_status"] == "RESERVE"
    assert result.loc["unranked", "elimination_reason"] == "RANK_NOT_EVALUATED_INSUFFICIENT_DATA"


def test_disabled_final_ranker_is_inert() -> None:
    request = parse_selection_request({"symbol": "BTCUSDT", "side": "LONG", "stages": [
        {"id": "rank_robust_top_n", "enabled": False, "scope": "pair_side", "top_n": 1},
    ]})
    result = run_selection(pd.DataFrame([_selection_row("only", strategy_id=1)]), request).set_index("strategy_name")

    assert result.loc["only", "finalist"]
    assert result.loc["only", "elimination_reason"] is None
    assert "final_rank" not in result


def test_shift_near_tie_is_bp_based_and_permutation_independent() -> None:
    request = parse_selection_request({"symbol": "BTCUSDT", "side": "LONG", "stages": [
        {"id": "pareto_shift_near_tie", "enabled": True, "scope": "pair_side_timeframe", "pnl_tolerance_pct": "10"},
    ]})
    rows = [
        _selection_row("shift-110bp", strategy_id=1, robust_pnl_30d_pct=Decimal("100"), first_shift_bp=110,
                       worst_drawdown_pct=Decimal("4"), worst_holding_p95_minutes=Decimal("40")),
        _selection_row("shift-100bp", strategy_id=2, robust_pnl_30d_pct=Decimal("95"), first_shift_bp=100,
                       worst_drawdown_pct=Decimal("4"), worst_holding_p95_minutes=Decimal("40")),
    ]
    first = run_selection(pd.DataFrame(rows), request).set_index("strategy_name")
    second = run_selection(pd.DataFrame(list(reversed(rows))), request).set_index("strategy_name")

    assert first.loc["shift-100bp", "eliminated_by_pareto_shift_near_tie"]
    assert first["eliminated_by_pareto_shift_near_tie"].to_dict() == second["eliminated_by_pareto_shift_near_tie"].to_dict()


def test_close_ma_near_tie_prefers_only_strictly_smaller_close_ma() -> None:
    request = parse_selection_request({"symbol": "BTCUSDT", "side": "LONG", "stages": [
        {"id": "pareto_close_ma_near_tie", "enabled": True, "scope": "pair_side_timeframe", "pnl_tolerance_pct": "10"},
    ]})
    result = run_selection(pd.DataFrame([
        _selection_row("smaller-close", strategy_id=1, robust_pnl_30d_pct=Decimal("95"), close_ma_len=3,
                       worst_drawdown_pct=Decimal("4"), worst_holding_p95_minutes=Decimal("40")),
        _selection_row("larger-close", strategy_id=2, robust_pnl_30d_pct=Decimal("100"), close_ma_len=5,
                       worst_drawdown_pct=Decimal("4"), worst_holding_p95_minutes=Decimal("40")),
        _selection_row("same-close", strategy_id=3, robust_pnl_30d_pct=Decimal("100"), close_ma_len=3,
                       worst_drawdown_pct=Decimal("4"), worst_holding_p95_minutes=Decimal("40")),
    ]), request).set_index("strategy_name")

    assert result.loc["larger-close", "eliminated_by_pareto_close_ma_near_tie"]
    assert not result.loc["smaller-close", "eliminated_by_pareto_close_ma_near_tie"]
    assert not result.loc["same-close", "eliminated_by_pareto_close_ma_near_tie"]


def test_final_rank_uses_only_prior_stage_survivors_and_renormalizes_weights() -> None:
    request = parse_selection_request({"symbol": "BTCUSDT", "side": "LONG", "stages": [
        {"id": "filter_min_shift", "enabled": True, "scope": "pair_side_timeframe", "min_shift_pct": "0.3"},
        {"id": "rank_robust_top_n", "enabled": True, "scope": "pair_side", "top_n": 1},
    ]})
    result = run_selection(pd.DataFrame([
        _selection_row("filtered-top", strategy_id=1, order_1_shift_bp=20, first_shift_bp=300,
                       robust_pnl_30d_pct=Decimal("100"), worst_drawdown_pct=Decimal("1"), worst_holding_p95_minutes=Decimal("1"),
                       ab_stability_ratio=Decimal("1"), minimum_plateau_point_count=100),
        _selection_row("survivor", strategy_id=2, order_1_shift_bp=30, first_shift_bp=100,
                       robust_pnl_30d_pct=Decimal("10"), worst_drawdown_pct=Decimal("5"), worst_holding_p95_minutes=Decimal("50"),
                       ab_stability_ratio=Decimal(".5"), minimum_plateau_point_count=10),
    ]), request).set_index("strategy_name")

    assert result.loc["filtered-top", "eliminated_by_filter_min_shift"]
    assert result.loc["filtered-top", "elimination_reason"] == "FILTER_MIN_SHIFT"
    assert result.loc["survivor", "finalist"]
    assert result.loc["survivor", "final_rank"] == 1
    assert sum(result.loc["survivor", f"rank_weight_{name}"] for name in (
        "robust_pnl", "worst_drawdown", "ab_stability", "worst_holding",
        "first_shift", "minimum_plateau_points", "close_ma",
    )) == pytest.approx(1.0)


def test_final_rank_prefers_smaller_close_ma_with_approved_weight() -> None:
    request = parse_selection_request({"symbol": "BTCUSDT", "side": "LONG", "stages": [
        {"id": "rank_robust_top_n", "enabled": True, "scope": "pair_side", "top_n": 2},
    ]})
    common = {
        "robust_pnl_30d_pct": Decimal("100"), "worst_drawdown_pct": Decimal("4"),
        "worst_holding_p95_minutes": Decimal("40"), "ab_stability_ratio": Decimal(".8"),
        "minimum_plateau_point_count": 20, "first_shift_bp": 100,
    }
    result = run_selection(pd.DataFrame([
        _selection_row("small-close", strategy_id=1, close_ma_len=3, **common),
        _selection_row("large-close", strategy_id=2, close_ma_len=7, **common),
    ]), request).set_index("strategy_name")

    assert result.loc["small-close", "final_rank"] == 1
    assert result.loc["large-close", "final_rank"] == 2
    assert result.loc["small-close", "rank_quality_close_ma"] == pytest.approx(1.0)
    assert result.loc["large-close", "rank_quality_close_ma"] == pytest.approx(0.0)
    assert result.loc["small-close", "rank_weight_close_ma"] == pytest.approx(0.09)
    assert [result.loc["small-close", f"rank_weight_{name}"] for name in (
        "robust_pnl", "worst_drawdown", "ab_stability", "first_shift", "minimum_plateau_points",
    )] == pytest.approx([0.30, 0.15, 0.15, 0.10, 0.09])


def test_final_rank_uses_approved_weights_including_worst_holding() -> None:
    request = parse_selection_request({"symbol": "BTCUSDT", "side": "LONG", "stages": [
        {"id": "rank_robust_top_n", "enabled": True, "scope": "pair_side", "top_n": 20},
    ]})
    result = run_selection(pd.DataFrame([
        _selection_row(
            "only", strategy_id=1, robust_pnl_30d_pct=Decimal("10"), worst_drawdown_pct=Decimal("5"),
            ab_stability_ratio=Decimal(".8"), worst_holding_p95_minutes=Decimal("60"),
            first_shift_bp=100, minimum_plateau_point_count=20, close_ma_len=3,
        ),
    ]), request).iloc[0]

    assert [result[f"rank_weight_{name}"] for name in (
        "robust_pnl", "worst_drawdown", "ab_stability", "worst_holding",
        "first_shift", "minimum_plateau_points", "close_ma",
    )] == pytest.approx([.30, .15, .15, .12, .10, .09, .09])
    assert result["final_score"] == pytest.approx(100)


def test_final_rank_collapses_exact_analogs_before_top_n() -> None:
    request = parse_selection_request({"symbol": "BABAUSDT", "side": "LONG", "stages": [
        {"id": "rank_robust_top_n", "enabled": True, "scope": "pair_side", "top_n": 1},
    ]})
    common = {
        "symbol": "BABAUSDT", "side": "LONG", "timeframe": "1h", "order_count": 2,
            "close_ma_len": 3, "worst_drawdown_pct": Decimal("5"),
            "ab_stability_ratio": Decimal(".8"), "worst_holding_p95_minutes": Decimal("60"),
            "first_shift_bp": 100, "minimum_plateau_point_count": 20,
            "order_1_plateau_key": (77, 900), "order_2_plateau_key": (77, 901),
    }
    result = run_selection(pd.DataFrame([
        _selection_row("best-analog", strategy_id=5, robust_pnl_30d_pct=Decimal("20"), **common),
        _selection_row("other-analog", strategy_id=6, robust_pnl_30d_pct=Decimal("10"), **common),
        _selection_row("other-close", strategy_id=7, close_ma_len=5, robust_pnl_30d_pct=Decimal("15"),
                       **{key: value for key, value in common.items() if key != "close_ma_len"}),
    ]), request).set_index("strategy_name")

    assert result.loc["best-analog", "auto_status"] == "FINALIST"
    assert result.loc["other-analog", "auto_status"] == "ANALOG"
    assert result.loc["other-analog", "auto_analog_of_strategy_id"] == 5
    assert result.loc["other-close", "auto_status"] == "RESERVE"
    assert result["finalist"].sum() == 1


def test_final_rank_collapses_adjacent_close_ma_only_with_identical_order_plateaus() -> None:
    request = parse_selection_request({"symbol": "BABAUSDT", "side": "LONG", "stages": [
        {"id": "rank_robust_top_n", "enabled": True, "scope": "pair_side", "top_n": 2},
    ]})
    common = {
        "symbol": "BABAUSDT", "side": "LONG", "timeframe": "1h", "order_count": 1,
        "worst_drawdown_pct": Decimal("5"), "ab_stability_ratio": Decimal(".8"),
        "worst_holding_p95_minutes": Decimal("60"), "first_shift_bp": 100,
        "minimum_plateau_point_count": 20, "order_1_plateau_key": (77, 900),
    }
    result = run_selection(pd.DataFrame([
        _selection_row("close-5", strategy_id=5, close_ma_len=5, robust_pnl_30d_pct=Decimal("20"), **common),
        _selection_row("close-6", strategy_id=6, close_ma_len=6, robust_pnl_30d_pct=Decimal("10"), **common),
        _selection_row("close-7", strategy_id=7, close_ma_len=7, robust_pnl_30d_pct=Decimal("15"), **common),
    ]), request).set_index("strategy_name")

    assert result.loc["close-5", "auto_status"] == "FINALIST"
    assert result.loc["close-6", "auto_status"] == "ANALOG"
    assert result.loc["close-6", "auto_analog_of_strategy_id"] == 5
    assert result.loc["close-7", "auto_status"] == "FINALIST"
    assert result.loc["close-7", "auto_analog_of_strategy_id"] is pd.NA


def test_prior_rejected_representative_does_not_consume_top_n_slot() -> None:
    request = parse_selection_request({"symbol": "BTCUSDT", "side": "LONG", "stages": [
        {"id": "rank_robust_top_n", "enabled": True, "scope": "pair_side", "top_n": 1},
    ]})
    common = {
        "symbol": "BTCUSDT", "side": "LONG", "timeframe": "1h", "order_count": 1,
        "worst_drawdown_pct": Decimal("5"), "ab_stability_ratio": Decimal(".8"),
        "worst_holding_p95_minutes": Decimal("60"), "first_shift_bp": 100,
        "minimum_plateau_point_count": 20,
    }
    result = run_selection(pd.DataFrame([
        _selection_row("rejected", strategy_id=1, close_ma_len=2, robust_pnl_30d_pct=Decimal("30"),
                       prior_rejected=True, **common),
        _selection_row("selected", strategy_id=2, close_ma_len=3, robust_pnl_30d_pct=Decimal("20"),
                       prior_rejected=False, **common),
        _selection_row("reserve", strategy_id=3, close_ma_len=4, robust_pnl_30d_pct=Decimal("10"),
                       prior_rejected=False, **common),
    ]), request).set_index("strategy_name")

    assert result.loc["rejected", "auto_status"] == "RESERVE"
    assert result.loc["rejected", "elimination_reason"] == "PRIOR_USER_REJECTED"
    assert result.loc["selected", "auto_status"] == "FINALIST"
    assert result.loc["reserve", "auto_status"] == "RESERVE"


def test_ranker_does_not_collapse_rows_with_missing_structural_key_or_dd5() -> None:
    request = parse_selection_request({"symbol": "BTCUSDT", "side": "LONG", "stages": [
        {"id": "rank_robust_top_n", "enabled": True, "scope": "pair_side", "top_n": 2},
    ]})
    rows = pd.DataFrame([
        {"strategy_id": 1, "strategy_name": "one", "timeframe": "1h", "close_ma_len": 3,
         "robust_pnl_30d_pct": 20, "worst_drawdown_pct": 5},
        {"strategy_id": 2, "strategy_name": "two", "timeframe": "1h", "close_ma_len": 3,
         "robust_pnl_30d_pct": 10, "worst_drawdown_pct": 6},
    ])

    result = run_selection(rows, request).set_index("strategy_id")

    assert result["auto_status"].to_dict() == {1: "FINALIST", 2: "FINALIST"}
    assert result["auto_analog_of_strategy_id"].isna().all()


def test_final_rank_breaks_boundary_ties_by_strategy_id_independent_of_input_order() -> None:
    request = parse_selection_request({"symbol": "BTCUSDT", "side": "LONG", "stages": [
        {"id": "rank_robust_top_n", "enabled": True, "scope": "pair_side", "top_n": 1},
    ]})
    facts = {
            "robust_pnl_30d_pct": Decimal("10"), "worst_drawdown_pct": Decimal("5"),
            "ab_stability_ratio": Decimal(".5"), "first_shift_bp": 100,
            "minimum_plateau_point_count": 10, "close_ma_len": 3,
            "order_1_plateau_key": (77, 900),
    }
    rows = [_selection_row("two", strategy_id=2, **facts), _selection_row("one", strategy_id=1, **facts)]

    first = run_selection(pd.DataFrame(rows), request).set_index("strategy_id")
    second = run_selection(pd.DataFrame(list(reversed(rows))), request).set_index("strategy_id")

    for result in (first, second):
        assert result.loc[1, "finalist"]
        assert result.loc[1, "final_rank"] == 1
        assert result.loc[2, "elimination_reason"] == "ANALOG"
        assert result.loc[2, "auto_analog_of_strategy_id"] == 1


def test_workbook_keeps_rank_eliminated_rows_with_rank_diagnostics(tmp_path: Path) -> None:
    request = parse_selection_request({"symbol": "BTCUSDT", "side": "LONG", "stages": [
        {"id": "rank_robust_top_n", "enabled": True, "scope": "pair_side", "top_n": 1},
    ]})
    result = run_selection(pd.DataFrame([
        _selection_row("top", strategy_id=1, close_ma_len=2, robust_pnl_30d_pct=Decimal("20"), worst_drawdown_pct=Decimal("4"),
                       worst_holding_p95_minutes=Decimal("40"), ab_stability_ratio=Decimal(".8"), minimum_plateau_point_count=20),
        _selection_row("cut", strategy_id=2, close_ma_len=3, robust_pnl_30d_pct=Decimal("10"), worst_drawdown_pct=Decimal("5"),
                       worst_holding_p95_minutes=Decimal("50"), ab_stability_ratio=Decimal(".7"), minimum_plateau_point_count=10),
    ]), request)
    book = load_workbook(write_selection_workbook(result, tmp_path / "ranked.xlsx", request), data_only=True)
    sheet = book["All candidates"]
    headers = [cell.value for cell in sheet[1]]

    assert sheet.max_row == 3
    assert "Final rank" in headers and "Rank coverage, %" in headers
    assert headers[-1] == "eliminated_by_rank_robust_top_n"
    assert {sheet.cell(row, headers.index("Final rank") + 1).value for row in (2, 3)} == {1, 2}


@pytest.mark.parametrize(("stage_id", "field", "values"), [
    ("filter_holding_outlier", "holding_p95_minutes", [10, 10, 10, 100]),
    ("filter_low_trades", "trades_30d", [100, 100, 100, 1]),
])
def test_iqr_filters_eliminate_only_outlier(stage_id: str, field: str, values: list[int]) -> None:
    rows = [_selection_row(f"row-{index}", strategy_id=index, **{field: value}) for index, value in enumerate(values)]
    request = parse_selection_request({"symbol": "BTCUSDT", "side": "LONG", "stages": [
        {"id": stage_id, "enabled": True, "scope": "pair_side"},
    ]})

    result = run_selection(pd.DataFrame(rows), request).set_index("strategy_name")

    assert result.loc["row-3", f"eliminated_by_{stage_id}"]
    assert result["finalist"].sum() == 3


def test_conditional_close_ma_needs_more_than_three_survivors() -> None:
    rows = [_selection_row(f"row-{index}", strategy_id=index, close_ma_len=3 + index,
                           capital_efficiency=Decimal(10 - index)) for index in range(4)]
    request = parse_selection_request({"symbol": "BTCUSDT", "side": "LONG", "stages": [
        {"id": "pareto_conditional_close_ma", "enabled": True, "scope": "pair_side"},
    ]})

    result = run_selection(pd.DataFrame(rows), request)

    assert result["eliminated_by_pareto_conditional_close_ma"].sum() == 3


def test_workbook_keeps_all_candidates_and_ab_30d_columns(tmp_path: Path) -> None:
    request = parse_selection_request({"symbol": "BTCUSDT", "side": "LONG", "stages": [
        {"id": "pareto_dd5_capital", "enabled": True, "scope": "pair_side"},
    ]})
    result = run_selection(pd.DataFrame([
        _selection_row(
            "winner", ab_return_a_30d_pct=Decimal("10.75"), ab_return_b_30d_pct=Decimal("4.25"), ab_calendar_days_a=Decimal("31"), ab_calendar_days_b=Decimal("14"), ab_pnl_change_30d_pct=Decimal("1.6"),
            total_pnl_pct=Decimal("12.6"), pnl_30d_pct=Decimal("8.5"), dd5_proxy=Decimal("5.5"), profit_factor=Decimal("1.6"),
            capital_efficiency=Decimal("10.6"), win_rate_pct=Decimal("67.8"), pnl_without_best_trade_pct=Decimal("5.6"),
            holding_p95_minutes=Decimal("10.7"), holding_median_minutes=Decimal("5.6"),
                positive_quarter_count=3, positive_quarter_available_count=3, trades_30d=Decimal("3.75"),
        ),
        _selection_row("loser", ab_return_b_30d_pct=Decimal("2.50")),
    ]), request)
    result["final_rank"] = [1, 2]
    result.loc[1, "elimination_reason"] = "PARETO_PLATEAU_POINTS_PER_ORDER"
    result.loc[0, "order_1_plateau_point_count"] = 21.0
    result.loc[0, "order_2_plateau_point_count"] = 33.0
    result.loc[0, "order_3_plateau_point_count"] = None

    path = write_selection_workbook(result, tmp_path / "finalists.xlsx", request)
    book = load_workbook(path, data_only=True)
    headers = [cell.value for cell in book["All candidates"][1]]

    assert "PnL" not in headers
    assert "total_pnl_pct" not in headers
    assert "PnL/30" in headers and "Trades/30" in headers
    assert "eliminated_by_filter_lot_variant_redundancy" in headers
    assert "positive_quarter_status" not in headers
    strategy_column = headers.index("Стратегия") + 1
    winner_row = next(row for row in range(2, book["All candidates"].max_row + 1) if book["All candidates"].cell(row, strategy_column).value == "winner")
    assert book["All candidates"].cell(winner_row, headers.index("PnL/30") + 1).value == 9
    assert headers[:25] == [
        "ID", "Стратегия", "Пара", "Side", "ТФ", "ORD", "Close", "PnL/30", "PnL DD5/30",
        "∆ PnL A/B", "PnL A/30д, %", "Дней A", "PnL B/30д, %", "Дней B", "Positive windows", "CE", "PF", "DD", "W/R", "Trades", "Trades/30", "Lot DD5", "Hold p95", "Hold M", "PointsALL",
    ]
    assert "Shift 1" not in headers
    strategy_column = headers.index("Стратегия") + 1
    assert book["All candidates"].column_dimensions[get_column_letter(strategy_column)].hidden
    data_rows = {
        book["All candidates"].cell(row, strategy_column).value: row
        for row in range(2, book["All candidates"].max_row + 1)
    }
    winner_row = data_rows["winner"]
    loser_row = data_rows["loser"]
    a_column = headers.index("PnL A/30д, %") + 1
    b_column = headers.index("PnL B/30д, %") + 1
    assert book["All candidates"].cell(winner_row, a_column).value == 11
    assert {book["All candidates"].cell(row, b_column).value for row in (winner_row, loser_row)} == {3, 4}
    assert book["All candidates"].cell(winner_row, b_column).data_type == "n"
    for header in ("Дней A", "Дней B"):
        column = headers.index(header) + 1
        cell = book["All candidates"].cell(winner_row, column)
        assert cell.data_type == "n"
        assert cell.alignment.horizontal == "center"
        assert cell.font.color.type == "rgb"
        assert cell.font.color.rgb == "FF0000FF"
        assert book["All candidates"].column_dimensions[get_column_letter(column)].width == 6
    assert book["All candidates"].cell(winner_row, headers.index("Positive windows") + 1).value == "3/3"
    trades_30_column = headers.index("Trades/30") + 1
    assert book["All candidates"].cell(winner_row, trades_30_column).value == 3.75
    assert book["All candidates"].cell(winner_row, trades_30_column).data_type == "n"
    pnl_without_best_column = headers.index("PnL without best, %") + 1
    assert {book["All candidates"].cell(row, pnl_without_best_column).value for row in (2, 3)} == {6, None}
    assert book["All candidates"].cell(3, pnl_without_best_column).data_type == "n"
    assert book["All candidates"].cell(3, headers.index("Причина") + 1).value == "PARETO_PL_PTS_PER_ORDER"
    assert book["All candidates"].cell(3, headers.index("Причина") + 1).alignment.horizontal == "left"
    for header in ("PnL/30", "PnL DD5/30", "PF", "PnL A/30д, %", "PnL B/30д, %", "PnL without best, %"):
        assert book["All candidates"].cell(2, headers.index(header) + 1).number_format == "0"
    for header in (
        "Positive trades", "Robust PnL/30", "Worst DD", "Worst Hold p95", "A/B stability", "Rank q PnL", "Rank q DD",
        "Rank q A/B", "Rank q Shift", "Rank q Points", "Rank coverage, %", "Rank w PnL",
        "Rank w DD", "Rank w A/B", "Rank w Shift", "Rank w Points", "Rank w Close MA",
        "Rank q Close MA", "Final score (Pair+Side)", "Best trade, %", "PnL without best, %",
    ):
        assert book["All candidates"].column_dimensions[get_column_letter(headers.index(header) + 1)].hidden
    assert not book["All candidates"].column_dimensions[get_column_letter(headers.index("Final rank") + 1)].hidden
    assert headers.index("Final rank") + 1 == headers.index("Final")
    for header in ("Close", "DD", "Hold p95", "1 Shift", "Final rank"):
        assert book["All candidates"].cell(2, headers.index(header) + 1).font.bold
    assert headers.index("ORD") + 1 == headers.index("Close")
    for header, edge in (
        ("Positive windows", "right"),
        ("Hold p95", "left"),
        ("Hold M", "right"),
        ("1 Shift", "left"),
        ("4 Shift", "right"),
        ("Points", "left"),
        ("Points", "right"),
        ("MA", "left"),
        ("MA", "right"),
        ("Final rank", "left"),
        ("Final rank", "right"),
        ("Close", "left"),
        ("Close", "right"),
    ):
        assert getattr(book["All candidates"].cell(1, headers.index(header) + 1).border, edge).style == "double"
    assert headers.index("PointsALL") + 1 == headers.index("PointsMin")
    assert headers.index("CE") + 1 == headers.index("PF")
    shifts_start = headers.index("1 Shift")
    assert headers[shifts_start:shifts_start + 7] == [
        "1 Shift", "2 Shift", "3 Shift", "4 Shift", "Lots", "Points", "MA",
    ]
    first_order_shift = headers.index("1 Shift") + 1
    assert {book["All candidates"].cell(row, first_order_shift).value for row in (2, 3)} == {0.3, 2.7}
    assert book["All candidates"].cell(2, first_order_shift).number_format == "0.0"
    assert book["All candidates"].cell(2, headers.index("Final rank") + 1).number_format == "0"
    assert book["All candidates"].cell(winner_row, headers.index("∆ PnL A/B") + 1).value == 2
    assert "1 Points" not in headers
    assert {book["All candidates"].cell(row, headers.index("Points") + 1).value for row in (2, 3)} == {
        "20 / 8 / 8 / 10", "21 / 33 / - / 10",
    }
    assert {book["All candidates"].cell(row, headers.index("MA") + 1).value for row in (2, 3)} == {"4 / 4 / 6 / 6"}
    assert {book["All candidates"].cell(row, headers.index("Lots") + 1).value for row in (2, 3)} == {"25 / 50 / 75 / 100"}
    assert book["All candidates"].column_dimensions[get_column_letter(headers.index("1 Shift") + 1)].width == 5
    assert book["All candidates"].column_dimensions[get_column_letter(headers.index("PF") + 1)].width == 5
    assert book["All candidates"].column_dimensions[get_column_letter(headers.index("Lots") + 1)].width == 20
    assert book["All candidates"].column_dimensions[get_column_letter(headers.index("MA") + 1)].width == 15
    assert book["All candidates"].cell(2, 4).alignment.horizontal is None
    assert book["All candidates"].cell(2, 5).alignment.horizontal == "center"
    assert book["Finalists"].cell(2, 5).alignment.horizontal == "center"

    assert book.sheetnames == ["All candidates", "Finalists"]
    assert headers[-1] == "eliminated_by_pareto_dd5_capital"
    assert {book["All candidates"].cell(row, len(headers)).value for row in (2, 3)} == {"BLOCK", "PASS"}
    assert "result_id" not in headers
    assert "total_pnl" not in headers
    assert book["All candidates"].max_row == 3
    assert book["Finalists"].max_row == 2


def test_workbook_prefixes_applied_filter_reason_and_fills_rows(tmp_path: Path) -> None:
    request = parse_selection_request({"symbol": "BTCUSDT", "side": "LONG", "stages": [
        {"id": "pareto_plateau_points_per_order", "enabled": True, "scope": "pair_side"},
        {"id": "pareto_dd5_capital", "enabled": True, "scope": "pair_side"},
    ]})
    result = pd.DataFrame([_selection_row("winner"), _selection_row("loser")])
    result["finalist"] = [False, True]
    result["elimination_reason"] = ["PARETO_DD5_CAPITAL", None]

    book = load_workbook(write_selection_workbook(result, tmp_path / "finalists.xlsx", request), data_only=True)
    headers = [cell.value for cell in book["All candidates"][1]]

    assert book["All candidates"].cell(2, headers.index("Причина") + 1).value == "2. PARETO_DD5_CAPITAL"
    assert book["All candidates"].cell(2, 1).fill.fgColor.rgb == "00FAEFEF"
    assert book["All candidates"].cell(3, 1).fill.fgColor.rgb == "00D9EAD3"
    assert book["Finalists"].cell(2, 1).fill.fgColor.rgb == "00D9EAD3"


def test_workbook_keeps_only_enabled_filter_columns_in_request_order(tmp_path: Path) -> None:
    request = parse_selection_request({"symbol": "BTCUSDT", "side": "LONG", "stages": [
        {"id": "filter_low_trades", "enabled": False, "scope": "pair_side"},
        {"id": "pareto_dd5_capital", "enabled": True, "scope": "pair_side"},
        {"id": "ab_deterioration", "enabled": False, "scope": "pair_side"},
        {"id": "pareto_dd5_balanced", "enabled": True, "scope": "pair_side"},
    ]})
    result = run_selection(pd.DataFrame([_selection_row("winner"), _selection_row("loser")]), request)

    path = write_selection_workbook(result, tmp_path / "finalists.xlsx", request)
    headers = [cell.value for cell in load_workbook(path, data_only=True)["All candidates"][1]]

    assert headers[-2:] == ["eliminated_by_pareto_dd5_capital", "eliminated_by_pareto_dd5_balanced"]
    assert "eliminated_by_filter_low_trades" not in headers
    assert "eliminated_by_ab_deterioration" not in headers


def test_workbook_consolidated_ma_never_shows_decimal_places(tmp_path: Path) -> None:
    request = parse_selection_request({"symbol": "BTCUSDT", "side": "LONG", "stages": []})
    result = run_selection(pd.DataFrame([_selection_row(
        "sparse", order_1_open_ma_len=7.0, order_2_open_ma_len=3.0,
        order_3_open_ma_len=None, order_4_open_ma_len=None,
    )]), request)

    book = load_workbook(write_selection_workbook(result, tmp_path / "sparse.xlsx", request), data_only=True)
    headers = [cell.value for cell in book["All candidates"][1]]

    assert book["All candidates"].cell(2, headers.index("MA") + 1).value == "7 / 3"


def test_scope_timeframe_prevents_cross_timeframe_pareto_comparison() -> None:
    frame = pd.DataFrame([
        _selection_row("winner", timeframe="1h"),
        _selection_row("loser", timeframe="3h"),
    ])
    pair_side = parse_selection_request({"symbol": "BTCUSDT", "side": "LONG", "stages": [
        {"id": "pareto_dd5_capital", "enabled": True, "scope": "pair_side"},
    ]})
    timeframe = parse_selection_request({"symbol": "BTCUSDT", "side": "LONG", "stages": [
        {"id": "pareto_dd5_capital", "enabled": True, "scope": "pair_side_timeframe"},
    ]})

    all_scope = run_selection(frame, pair_side).set_index("strategy_name")
    split_scope = run_selection(frame, timeframe).set_index("strategy_name")

    assert not all_scope.loc["loser", "finalist"]
    assert split_scope["finalist"].all()


def test_stage_order_changes_survivors_and_keeps_first_elimination_trace() -> None:
    failing_dominator = _selection_row(
        "a", dd5_proxy=Decimal("10"), capital_proxy=Decimal("1"), ab_return_a_30d_pct=Decimal("10"), ab_return_b_30d_pct=Decimal("4"),
        ab_win_rate_b_pct=Decimal("60"), ab_trade_rate_a_30d=Decimal("10"), ab_trade_rate_b_30d=Decimal("10"),
    )
    passing_dominated = _selection_row(
        "b", strategy_id=3, dd5_proxy=Decimal("5"), capital_proxy=Decimal("2"),
        ab_return_a_30d_pct=Decimal("10"), ab_return_b_30d_pct=Decimal("10"),
        ab_win_rate_b_pct=Decimal("60"), ab_trade_rate_a_30d=Decimal("10"), ab_trade_rate_b_30d=Decimal("10"),
    )
    ab_first = parse_selection_request({"symbol": "BTCUSDT", "side": "LONG", "stages": [
        {"id": "ab_deterioration", "enabled": True, "scope": "pair_side"},
        {"id": "pareto_dd5_capital", "enabled": True, "scope": "pair_side"},
    ]})
    pareto_first = parse_selection_request({"symbol": "BTCUSDT", "side": "LONG", "stages": [
        {"id": stage.id, "enabled": stage.enabled, "scope": stage.scope}
        for stage in reversed(ab_first.stages)
    ]})

    first = run_selection(pd.DataFrame([failing_dominator, passing_dominated]), ab_first).set_index("strategy_name")
    second = run_selection(pd.DataFrame([failing_dominator, passing_dominated]), pareto_first).set_index("strategy_name")

    assert first.index[first["finalist"]].tolist() == ["b"]
    assert second.index[second["finalist"]].tolist() == []
    assert first.loc["a", "eliminated_by_ab_deterioration"]
    assert not first.loc["a", "eliminated_by_pareto_dd5_capital"]
    assert second.loc["b", "eliminated_by_pareto_dd5_capital"]
    assert not second.loc["b", "eliminated_by_ab_deterioration"]
    assert second.loc["a", "eliminated_by_ab_deterioration"]
    assert not second.loc["a", "eliminated_by_pareto_dd5_capital"]


def test_stage_counts_follow_the_applied_stage_order() -> None:
    failing_dominator = _selection_row(
        "a", dd5_proxy=Decimal("10"), capital_proxy=Decimal("1"), ab_return_a_30d_pct=Decimal("10"), ab_return_b_30d_pct=Decimal("4"),
        ab_win_rate_b_pct=Decimal("60"), ab_trade_rate_a_30d=Decimal("10"), ab_trade_rate_b_30d=Decimal("10"),
    )
    passing_dominated = _selection_row(
        "b", strategy_id=3, dd5_proxy=Decimal("5"), capital_proxy=Decimal("2"),
        ab_return_a_30d_pct=Decimal("10"), ab_return_b_30d_pct=Decimal("10"),
        ab_win_rate_b_pct=Decimal("60"), ab_trade_rate_a_30d=Decimal("10"), ab_trade_rate_b_30d=Decimal("10"),
    )
    request = parse_selection_request({"symbol": "BTCUSDT", "side": "LONG", "stages": [
        {"id": "ab_deterioration", "enabled": True, "scope": "pair_side"},
        {"id": "pareto_dd5_capital", "enabled": True, "scope": "pair_side"},
        {"id": "pareto_dd5_balanced", "enabled": False, "scope": "pair_side"},
    ]})

    result = run_selection(pd.DataFrame([failing_dominator, passing_dominated]), request)

    assert result.attrs["stage_counts"] == {
        "ab_deterioration": {"enabled": True, "eliminated": 1, "remaining": 1},
        "pareto_dd5_capital": {"enabled": True, "eliminated": 0, "remaining": 1},
        "pareto_dd5_balanced": {"enabled": False, "eliminated": 0, "remaining": 1},
    }


def test_missing_pareto_objective_neither_dominates_nor_is_eliminated() -> None:
    request = parse_selection_request({"symbol": "BTCUSDT", "side": "LONG", "stages": [
        {"id": "pareto_dd5_capital", "enabled": True, "scope": "pair_side"},
    ]})
    result = run_selection(pd.DataFrame([
        _selection_row("missing", capital_proxy=None), _selection_row("complete"),
    ]), request)

    assert result["finalist"].all()
    assert not result["eliminated_by_pareto_dd5_capital"].any()
