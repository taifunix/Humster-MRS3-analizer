"""Immutable finalist snapshots and strict XLSX review import."""

from __future__ import annotations

from dataclasses import asdict
from datetime import datetime, timezone
from decimal import Decimal
from hashlib import sha256
from io import BytesIO
import json
from typing import Mapping
from uuid import uuid4
import zipfile

import duckdb
from openpyxl import load_workbook
import pandas as pd

from .performance_v2_selection import SelectionConfig, SelectionRequest


SELECTION_CONTRACT_VERSION = "performance-v2-selection-review-v1"
WORKBOOK_SCHEMA_VERSION = "1"
META_SHEET = "_MRS_SELECTION_META"
MAX_UPLOAD_BYTES = 20 * 1024 * 1024
MAX_ZIP_ENTRIES = 256
MAX_UNCOMPRESSED_BYTES = 100 * 1024 * 1024
STATUSES = frozenset({"FINALIST", "RESERVE", "ANALOG", "FILTERED", "REJECTED"})


class SelectionReviewError(ValueError):
    def __init__(self, code: str, message: str = "", *, details: object = None) -> None:
        self.code = code
        self.details = details
        super().__init__(message or code)


def _rollback_quietly(connection: duckdb.DuckDBPyConnection) -> None:
    try:
        connection.execute("rollback")
    except Exception:
        pass


def _json_value(value: object) -> object:
    if isinstance(value, Decimal):
        return str(value)
    if isinstance(value, tuple):
        return [_json_value(item) for item in value]
    if isinstance(value, list):
        return [_json_value(item) for item in value]
    if isinstance(value, dict):
        return {str(key): _json_value(item) for key, item in value.items()}
    return value


def canonical_json(value: object) -> str:
    return json.dumps(_json_value(value), ensure_ascii=False, sort_keys=True, separators=(",", ":"))


def canonical_contract(request: SelectionRequest, config: SelectionConfig) -> tuple[str, str, str, str]:
    request_json = canonical_json(asdict(request))
    config_json = canonical_json(asdict(config))
    return request_json, sha256(request_json.encode()).hexdigest(), config_json, sha256(config_json.encode()).hexdigest()


def database_instance_id(connection: duckdb.DuckDBPyConnection) -> str:
    row = connection.execute("select value from schema_info where key = 'database_instance_id'").fetchone()
    if not row:
        raise SelectionReviewError("SELECTION_REVIEW_DATABASE_MISMATCH")
    return str(row[0])


def new_run_metadata(connection: duckdb.DuckDBPyConnection) -> dict[str, str]:
    return {
        "workbook_schema_version": WORKBOOK_SCHEMA_VERSION,
        "selection_run_id": str(uuid4()),
        "database_instance_id": database_instance_id(connection),
        "selection_contract_version": SELECTION_CONTRACT_VERSION,
        "exported_at_utc": datetime.now(timezone.utc).isoformat(),
    }


def apply_prior_rejected(connection: duckdb.DuckDBPyConnection, candidates: pd.DataFrame) -> pd.DataFrame:
    output = candidates.copy()
    rejected = {int(row[0]) for row in connection.execute("select strategy_id from strategy_tags where tag = 'REJECTED'").fetchall()}
    retest = {int(row[0]) for row in connection.execute("select strategy_id from strategy_tags where tag = 'RETEST'").fetchall()}
    output["prior_rejected"] = output["strategy_id"].map(lambda value: int(value) in rejected)
    output["prior_retest"] = output["strategy_id"].map(lambda value: int(value) in retest)
    return output


def _cell(value: object) -> object:
    try:
        if pd.isna(value):
            return None
    except (TypeError, ValueError):
        pass
    if isinstance(value, Decimal):
        return float(value)
    return value


def _insert_rows(
    connection: duckdb.DuckDBPyConnection, table: str, columns: tuple[str, ...], rows: list[list[object]]
) -> None:
    if not rows:
        return
    relation = f"_selection_input_{uuid4().hex}"
    connection.register(relation, pd.DataFrame(rows, columns=columns))
    try:
        names = ", ".join(columns)
        connection.execute(f"insert into {table} ({names}) select {names} from {relation}")
    finally:
        connection.unregister(relation)


def persist_selection_snapshot(
    connection: duckdb.DuckDBPyConnection,
    request: SelectionRequest,
    config: SelectionConfig,
    result: pd.DataFrame,
    metadata: Mapping[str, str],
    workbook_bytes: bytes,
) -> str:
    required = {"strategy_id", "result_id", "auto_status"}
    if not required.issubset(result.columns):
        raise SelectionReviewError("SELECTION_REVIEW_INVALID_SELECTION")
    run_id = metadata["selection_run_id"]
    request_json, request_hash, config_json, config_hash = canonical_contract(request, config)
    expected = {int(row.strategy_id): int(row.result_id) for row in result.itertuples()}
    rank_stage = next((stage for stage in request.stages if stage.id == "rank_robust_top_n"), None)
    top_n = rank_stage.top_n if rank_stage and rank_stage.top_n else 20
    representative_count = int(result["auto_status"].isin(["FINALIST", "RESERVE"]).sum())
    workbook_hash = sha256(workbook_bytes).hexdigest()
    connection.execute("begin transaction")
    try:
        current = dict(connection.execute(
            "select strategy_id, current_result_id from strategies where strategy_id in (select unnest(?::bigint[]))",
            [list(expected)],
        ).fetchall()) if expected else {}
        stale = sorted(strategy_id for strategy_id, result_id in expected.items() if current.get(strategy_id) != result_id)
        if stale:
            raise SelectionReviewError("SELECTION_REVIEW_STALE_RESULTS", details=stale)
        connection.execute(
            """insert into selection_runs (
                selection_run_id, database_instance_id, symbol, side, selection_contract_version,
                request_json, request_sha256, config_json, config_sha256, candidate_count,
                representative_count, auto_finalist_count, top_n, workbook_sha256, created_at_utc
            ) values (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            [run_id, metadata["database_instance_id"], request.symbol, request.side,
             metadata["selection_contract_version"], request_json, request_hash, config_json, config_hash,
             len(result), representative_count, int((result["auto_status"] == "FINALIST").sum()), top_n,
             workbook_hash, datetime.now(timezone.utc)],
        )
        stage_columns = [f"eliminated_by_{stage.id}" for stage in request.stages if stage.enabled]
        rows = []
        for row in result.to_dict(orient="records"):
            trace = canonical_json({column.removeprefix("eliminated_by_"): bool(row.get(column)) for column in stage_columns})
            rows.append([
                run_id, int(row["strategy_id"]), int(row["result_id"]), str(row["auto_status"]),
                _cell(row.get("final_score")), _cell(row.get("final_rank")), _cell(row.get("elimination_reason")),
                _cell(row.get("analog_group_key")), _cell(row.get("auto_analog_of_strategy_id")),
                bool(row.get("prior_rejected", False)), trace,
            ])
        _insert_rows(connection, "selection_results", (
            "selection_run_id", "strategy_id", "result_id_at_selection", "auto_status", "auto_score",
            "auto_rank", "auto_reason", "analog_group_key", "auto_analog_of_strategy_id", "prior_rejected",
            "stage_trace_json",
        ), rows)
        connection.execute("commit")
    except duckdb.ConstraintException as error:
        _rollback_quietly(connection)
        if "workbook_sha256" in str(error):
            raise SelectionReviewError("SELECTION_REVIEW_ALREADY_IMPORTED") from error
        raise
    except Exception:
        _rollback_quietly(connection)
        raise
    return run_id


def _bounded_xlsx(data: bytes) -> None:
    if not data or len(data) > MAX_UPLOAD_BYTES:
        raise SelectionReviewError("SELECTION_REVIEW_INVALID_FILE")
    try:
        with zipfile.ZipFile(BytesIO(data)) as archive:
            infos = archive.infolist()
            if len(infos) > MAX_ZIP_ENTRIES or sum(info.file_size for info in infos) > MAX_UNCOMPRESSED_BYTES:
                raise SelectionReviewError("SELECTION_REVIEW_INVALID_FILE")
    except (zipfile.BadZipFile, OSError):
        raise SelectionReviewError("SELECTION_REVIEW_INVALID_FILE") from None


def _whole_number(value: object, code: str, *, optional: bool = True) -> int | None:
    if value is None or value == "":
        if optional:
            return None
        raise SelectionReviewError(code)
    if isinstance(value, bool):
        raise SelectionReviewError(code)
    try:
        number = int(value)
    except (TypeError, ValueError, OverflowError):
        raise SelectionReviewError(code) from None
    if number <= 0 or float(value) != number:
        raise SelectionReviewError(code)
    return number


def _normalize_retest(value: object) -> bool:
    if value is None:
        return False
    if isinstance(value, str):
        normalized = value.strip()
        if not normalized:
            return False
        if normalized == "RETEST":
            return True
    raise SelectionReviewError("SELECTION_REVIEW_INVALID_RETEST")


def _parse_workbook(data: bytes) -> tuple[dict[str, str], list[dict[str, object]]]:
    _bounded_xlsx(data)
    try:
        workbook = load_workbook(BytesIO(data), data_only=False, read_only=True)
    except Exception:
        raise SelectionReviewError("SELECTION_REVIEW_INVALID_FILE") from None
    try:
        if any(cell.data_type == "f" for worksheet in workbook.worksheets for row in worksheet.iter_rows() for cell in row):
            raise SelectionReviewError("SELECTION_REVIEW_INVALID_FILE")
        if META_SHEET not in workbook.sheetnames or "All candidates" not in workbook.sheetnames:
            raise SelectionReviewError("SELECTION_REVIEW_SCHEMA_MISMATCH")
        meta_sheet = workbook[META_SHEET]
        metadata = {str(key): str(value) for key, value in meta_sheet.iter_rows(min_row=1, max_col=2, values_only=True) if key and value is not None}
        if metadata.get("workbook_schema_version") != WORKBOOK_SCHEMA_VERSION or metadata.get("selection_contract_version") != SELECTION_CONTRACT_VERSION:
            raise SelectionReviewError("SELECTION_REVIEW_SCHEMA_MISMATCH")
        sheet = workbook["All candidates"]
        header_cells = next(sheet.iter_rows(min_row=1, max_row=1), ())
        raw_headers = [cell.value for cell in header_cells]
        header_pairs = [
            (position, header)
            for position, header in enumerate(raw_headers)
            if isinstance(header, str) and header.strip()
        ]
        headers = [header for _, header in header_pairs]
        positions = {header: position for position, header in header_pairs}
        required = {"ID", "Result ID", "Стратегия", "Auto Status", "User Status", "RETEST", "Auto Rank", "User Rank", "Auto Analog Of ID", "Analog Of ID", "Comment"}
        if (len(headers) != len(set(headers)) or not required.issubset(headers)
                or positions["RETEST"] != positions["User Status"] + 1):
            raise SelectionReviewError("SELECTION_REVIEW_SCHEMA_MISMATCH")
        index = {header: position for position, header in header_pairs}
        rows: list[dict[str, object]] = []
        for cells in sheet.iter_rows(min_row=2):
            values = [cell.value for cell in cells]
            if all(value is None for value in values):
                continue
            rows.append({header: values[position] if position < len(values) else None for header, position in index.items()})
        return metadata, rows
    except SelectionReviewError:
        raise
    except Exception:
        raise SelectionReviewError("SELECTION_REVIEW_INVALID_FILE") from None
    finally:
        workbook.close()


def import_selection_review(connection: duckdb.DuckDBPyConnection, data: bytes) -> dict[str, object]:
    metadata, rows = _parse_workbook(data)
    if metadata.get("database_instance_id") != database_instance_id(connection):
        raise SelectionReviewError("SELECTION_REVIEW_DATABASE_MISMATCH")
    run_id = metadata.get("selection_run_id", "")
    run = connection.execute("select symbol, side from selection_runs where selection_run_id = ?", [run_id]).fetchone()
    if not run:
        raise SelectionReviewError("SELECTION_REVIEW_SCHEMA_MISMATCH")
    latest = connection.execute(
        "select selection_run_id from selection_runs where symbol = ? and side = ? order by created_at_utc desc, selection_run_id desc limit 1",
        list(run),
    ).fetchone()
    if latest != (run_id,):
        raise SelectionReviewError("SELECTION_REVIEW_NOT_LATEST_RUN")
    workbook_hash = sha256(data).hexdigest()
    if connection.execute("select 1 from selection_review_imports where workbook_sha256 = ?", [workbook_hash]).fetchone():
        raise SelectionReviewError("SELECTION_REVIEW_ALREADY_IMPORTED")
    snapshot_rows = connection.execute(
        """select strategy_id, result_id_at_selection, auto_status, auto_rank, auto_analog_of_strategy_id
             from selection_results where selection_run_id = ?""", [run_id]
    ).fetchall()
    snapshot = {int(row[0]): row[1:] for row in snapshot_rows}
    names = dict(connection.execute(
        "select strategy_id, strategy_name from strategies where strategy_id in (select unnest(?::bigint[]))", [list(snapshot)]
    ).fetchall()) if snapshot else {}
    submitted: dict[int, dict[str, object]] = {}
    for row in rows:
        strategy_id = _whole_number(row["ID"], "SELECTION_REVIEW_ROWSET_MISMATCH", optional=False)
        if strategy_id in submitted:
            raise SelectionReviewError("SELECTION_REVIEW_ROWSET_MISMATCH")
        submitted[strategy_id] = row
    retest_by_id: dict[int, bool] = {}
    ranks: set[int] = set()
    for strategy_id, row in submitted.items():
        status = str(row["User Status"] or "").strip().upper()
        if status not in STATUSES:
            raise SelectionReviewError("SELECTION_REVIEW_INVALID_STATUS")
        rank = _whole_number(row["User Rank"], "SELECTION_REVIEW_INVALID_RANK")
        if rank is not None:
            if status not in {"FINALIST", "RESERVE"} or rank in ranks:
                raise SelectionReviewError("SELECTION_REVIEW_INVALID_RANK")
            ranks.add(rank)
        comment = "" if row["Comment"] is None else str(row["Comment"])
        if len(comment) > 1000:
            raise SelectionReviewError("SELECTION_REVIEW_INVALID_FILE")
        retest_by_id[strategy_id] = _normalize_retest(row["RETEST"])
    if set(submitted) != set(snapshot):
        raise SelectionReviewError("SELECTION_REVIEW_ROWSET_MISMATCH")
    decisions: list[list[object]] = []
    for strategy_id, row in submitted.items():
        result_id, auto_status, auto_rank, auto_analog = snapshot[strategy_id]
        if (str(row["Стратегия"]) != names.get(strategy_id)
                or _whole_number(row["Result ID"], "SELECTION_REVIEW_AUTOMATIC_FIELDS_CHANGED", optional=False) != result_id
                or str(row["Auto Status"]) != auto_status
                or _whole_number(row["Auto Rank"], "SELECTION_REVIEW_AUTOMATIC_FIELDS_CHANGED") != auto_rank
                or _whole_number(row["Auto Analog Of ID"], "SELECTION_REVIEW_AUTOMATIC_FIELDS_CHANGED") != auto_analog):
            raise SelectionReviewError("SELECTION_REVIEW_AUTOMATIC_FIELDS_CHANGED")
        status = str(row["User Status"]).strip().upper()
        rank = _whole_number(row["User Rank"], "SELECTION_REVIEW_INVALID_RANK")
        analog = _whole_number(row["Analog Of ID"], "SELECTION_REVIEW_INVALID_ANALOG")
        if status == "ANALOG":
            if analog is None or analog == strategy_id or analog not in submitted:
                raise SelectionReviewError("SELECTION_REVIEW_INVALID_ANALOG")
        elif analog is not None:
            raise SelectionReviewError("SELECTION_REVIEW_INVALID_ANALOG")
        decisions.append([strategy_id, status, rank, analog, "" if row["Comment"] is None else str(row["Comment"])])
    statuses = {int(row[0]): str(row[1]) for row in decisions}
    if any(row[3] is not None and statuses[int(row[3])] not in {"FINALIST", "RESERVE"} for row in decisions):
        raise SelectionReviewError("SELECTION_REVIEW_INVALID_ANALOG")
    current = dict(connection.execute(
        "select strategy_id, current_result_id from strategies where strategy_id in (select unnest(?::bigint[]))", [list(snapshot)]
    ).fetchall())
    stale = sorted(strategy_id for strategy_id, values in snapshot.items() if current.get(strategy_id) != values[0])
    if stale:
        raise SelectionReviewError("SELECTION_REVIEW_STALE_RESULTS", details=stale)
    review_id = str(uuid4())
    now = datetime.now(timezone.utc)
    connection.execute("begin transaction")
    try:
        latest_again = connection.execute(
            "select selection_run_id from selection_runs where symbol = ? and side = ? order by created_at_utc desc, selection_run_id desc limit 1", list(run)
        ).fetchone()
        if latest_again != (run_id,):
            raise SelectionReviewError("SELECTION_REVIEW_NOT_LATEST_RUN")
        current_again = dict(connection.execute(
            "select strategy_id, current_result_id from strategies where strategy_id in (select unnest(?::bigint[]))", [list(snapshot)]
        ).fetchall())
        stale = sorted(strategy_id for strategy_id, values in snapshot.items() if current_again.get(strategy_id) != values[0])
        if stale:
            raise SelectionReviewError("SELECTION_REVIEW_STALE_RESULTS", details=stale)
        connection.execute(
            """insert into selection_review_imports (
                review_import_id, selection_run_id, workbook_sha256, imported_at_utc, row_count
            ) values (?, ?, ?, ?, ?)""",
            [review_id, run_id, workbook_hash, now, len(decisions)],
        )
        _insert_rows(connection, "selection_review_rows", (
            "review_import_id", "strategy_id", "user_status", "user_rank", "user_analog_of_strategy_id", "comment",
        ), [[review_id, *decision] for decision in decisions])
        ids = list(snapshot)
        connection.execute("delete from strategy_tags where tag = 'REJECTED' and strategy_id in (select unnest(?::bigint[]))", [ids])
        rejected = [[strategy_id, "REJECTED", "SELECTION_REVIEW", review_id, now] for strategy_id, status, *_ in decisions if status == "REJECTED"]
        _insert_rows(connection, "strategy_tags", (
            "strategy_id", "tag", "source", "source_ref", "updated_at_utc",
        ), rejected)
        blank_retest_ids = [strategy_id for strategy_id in ids if not retest_by_id[strategy_id]]
        if blank_retest_ids:
            connection.execute(
                "delete from strategy_tags where tag = 'RETEST' and source = 'SELECTION_REVIEW' and strategy_id in (select unnest(?::bigint[]))",
                [blank_retest_ids],
            )
        asserted_retest = [
            [strategy_id, "RETEST", "SELECTION_REVIEW", review_id, now]
            for strategy_id in ids if retest_by_id[strategy_id]
        ]
        if asserted_retest:
            connection.executemany(
                """insert into strategy_tags (strategy_id, tag, source, source_ref, updated_at_utc)
                   values (?, ?, ?, ?, ?)
                   on conflict (strategy_id, tag) do update set
                       source = excluded.source,
                       source_ref = excluded.source_ref,
                       updated_at_utc = excluded.updated_at_utc""",
                asserted_retest,
            )
        connection.execute("commit")
    except duckdb.ConstraintException as error:
        _rollback_quietly(connection)
        if "workbook_sha256" in str(error):
            raise SelectionReviewError("SELECTION_REVIEW_ALREADY_IMPORTED") from error
        raise
    except Exception:
        _rollback_quietly(connection)
        raise
    return {"review_import_id": review_id, "selection_run_id": run_id, "row_count": len(decisions), "finalist_count": sum(row[1] == "FINALIST" for row in decisions)}


def latest_effective_finalists(connection: duckdb.DuckDBPyConnection, symbol: str) -> tuple[bool, set[int]]:
    runs = connection.execute(
        """select selection_run_id from (
               select selection_run_id, row_number() over (partition by side order by created_at_utc desc, selection_run_id desc) as rn
               from selection_runs where symbol = ?
           ) where rn = 1""", [symbol]
    ).fetchall()
    finalists: set[int] = set()
    for (run_id,) in runs:
        review = connection.execute(
            "select review_import_id from selection_review_imports where selection_run_id = ? order by imported_at_utc desc, review_import_id desc limit 1", [run_id]
        ).fetchone()
        if review:
            finalists.update(int(row[0]) for row in connection.execute(
                "select strategy_id from selection_review_rows where review_import_id = ? and user_status = 'FINALIST'", [review[0]]
            ).fetchall())
        else:
            finalists.update(int(row[0]) for row in connection.execute(
                """select strategy_id from selection_results
                   where selection_run_id = ? and auto_status = 'FINALIST' and not prior_rejected""", [run_id]
            ).fetchall())
    return bool(runs), finalists
