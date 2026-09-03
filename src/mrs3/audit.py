from __future__ import annotations

from copy import copy
from datetime import datetime
from decimal import Decimal
import json
from pathlib import Path
import re
import shutil
import tempfile
import zipfile
from typing import Mapping, Sequence

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows


FIXED_EXCEL_TIME = datetime(2000, 1, 1, 0, 0, 0)


def serialize_value(value: object) -> object:
    if value is None:
        return None
    if isinstance(value, (pd.Timestamp, datetime)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return str(value)
    if isinstance(value, (tuple, list, dict, set)):
        serializable = sorted(value) if isinstance(value, set) else value
        return json.dumps(serializable, ensure_ascii=False, sort_keys=True, default=str)
    try:
        if pd.isna(value):
            return None
    except (TypeError, ValueError):
        pass
    return value


def serializable_frame(frame: pd.DataFrame) -> pd.DataFrame:
    output = frame.copy()
    for column in output.columns:
        output[column] = output[column].map(serialize_value)
    return output


def canonical_frame_json(frame: pd.DataFrame) -> str:
    serializable = serializable_frame(frame)
    return json.dumps(
        serializable.to_dict(orient="records"),
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
        default=str,
    )


def _normalize_xlsx_archive(path: Path) -> None:
    normalized = path.with_suffix(".normalized.xlsx")
    with zipfile.ZipFile(path, "r") as source, zipfile.ZipFile(
        normalized, "w", compression=zipfile.ZIP_DEFLATED, compresslevel=9
    ) as target:
        for name in sorted(source.namelist()):
            original = source.getinfo(name)
            info = zipfile.ZipInfo(name, (2000, 1, 1, 0, 0, 0))
            info.compress_type = zipfile.ZIP_DEFLATED
            info.external_attr = original.external_attr
            info.create_system = original.create_system
            contents = source.read(name)
            if name == "docProps/core.xml":
                text = contents.decode("utf-8")
                text, replacements = re.subn(
                    r"(<dcterms:modified\b[^>]*>)[^<]*(</dcterms:modified>)",
                    r"\g<1>2000-01-01T00:00:00Z\g<2>",
                    text,
                    count=1,
                )
                if replacements != 1:
                    raise ValueError("workbook core properties do not contain modified timestamp")
                contents = text.encode("utf-8")
            target.writestr(info, contents)
    normalized.replace(path)


def write_audit_workbook(
    tables: Mapping[str, pd.DataFrame], path: Path, *, data_widths_only: bool = False,
    minimum_width: int = 10, hidden_columns: frozenset[str] = frozenset(), decimal_comma: bool = False,
    numeric_decimals: bool = False, number_formats: Mapping[str, str] | None = None,
    bold_columns: frozenset[str] = frozenset(), column_edge_borders: Mapping[str, tuple[str, ...]] | None = None,
    center_from_column: int | None = None, left_aligned_columns: frozenset[str] = frozenset(),
    row_fill_colors: Mapping[str, Sequence[str | None]] | None = None, font_colors: Mapping[str, str] | None = None,
) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    workbook.remove(workbook.active)
    workbook.properties.creator = "MRS3 v0.6"
    workbook.properties.created = FIXED_EXCEL_TIME
    workbook.properties.modified = FIXED_EXCEL_TIME
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for sheet_name, raw_frame in tables.items():
        worksheet = workbook.create_sheet(sheet_name)
        frame_source = raw_frame
        if numeric_decimals:
            frame_source = raw_frame.copy()
            for column in frame_source.columns:
                frame_source[column] = frame_source[column].map(
                    lambda value: float(value) if isinstance(value, Decimal) else value
                )
        frame = serializable_frame(frame_source)
        if decimal_comma:
            for column in frame.columns:
                frame[column] = frame[column].map(
                    lambda value: re.sub(r"(?<=\d)\.(?=\d+$)", ",", value) if isinstance(value, str) else value
                )
        for row in dataframe_to_rows(frame, index=False, header=True):
            worksheet.append(row)
        if center_from_column is not None:
            for row in worksheet.iter_rows(min_col=center_from_column):
                for cell in row:
                    cell.alignment = Alignment(horizontal="center")
        for index, column in enumerate(frame.columns, start=1):
            if str(column) in left_aligned_columns:
                for cell in worksheet.iter_cols(min_col=index, max_col=index, min_row=2):
                    for value_cell in cell:
                        value_cell.alignment = Alignment(horizontal="left")
        if numeric_decimals:
            for row in worksheet.iter_rows(min_row=2):
                for cell in row:
                    if isinstance(cell.value, float):
                        cell.number_format = "0.00"
        if number_formats:
            for index, column in enumerate(frame.columns, start=1):
                if number_format := number_formats.get(str(column)):
                    for cell in worksheet.iter_cols(min_col=index, max_col=index, min_row=2):
                        for value_cell in cell:
                            if isinstance(value_cell.value, (int, float)):
                                value_cell.number_format = number_format
        if bold_columns or column_edge_borders or font_colors:
            edge_side = Side(style="double", color="8FA3B8")
            for index, column in enumerate(frame.columns, start=1):
                cells = list(worksheet.iter_cols(min_col=index, max_col=index, min_row=1, max_row=worksheet.max_row))[0]
                if str(column) in bold_columns:
                    for value_cell in cells[1:]:
                        value_cell.font = Font(bold=True)
                if color := (font_colors or {}).get(str(column)):
                    for value_cell in cells[1:]:
                        font = copy(value_cell.font)
                        font.color = color
                        value_cell.font = font
                for edge in (column_edge_borders or {}).get(str(column), ()):
                    border = Border(**{edge: edge_side})
                    for value_cell in cells:
                        value_cell.border += border
        for row, color in zip(worksheet.iter_rows(min_row=2), (row_fill_colors or {}).get(sheet_name, ())):
            if color:
                for cell in row:
                    cell.fill = PatternFill("solid", fgColor=color)
        if len(frame.columns):
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
            worksheet.freeze_panes = "A2"
            worksheet.auto_filter.ref = worksheet.dimensions
            for index, column in enumerate(frame.columns, start=1):
                values = [
                    "" if value is None else str(value)
                    for value in frame[column].head(2000)
                ]
                if not data_widths_only:
                    values.insert(0, str(column))
                width = min(70, max(minimum_width, max((len(value) for value in values), default=0) + 2))
                dimension = worksheet.column_dimensions[get_column_letter(index)]
                dimension.width = width
                dimension.hidden = str(column) in hidden_columns
    with tempfile.NamedTemporaryFile(
        prefix=f".{path.stem}.", suffix=".xlsx", dir=path.parent, delete=False
    ) as handle:
        temporary = Path(handle.name)
    try:
        workbook.save(temporary)
        _normalize_xlsx_archive(temporary)
        temporary.replace(path)
    finally:
        temporary.unlink(missing_ok=True)
    return path


def write_audit_csvs(tables: Mapping[str, pd.DataFrame], directory: Path) -> None:
    target_directory = directory.resolve()
    target_directory.parent.mkdir(parents=True, exist_ok=True)
    backup = target_directory.with_name(f".{target_directory.name}.mrs3-backup")
    if backup.exists():
        raise RuntimeError(f"audit CSV backup requires recovery: {backup}")
    staging = Path(
        tempfile.mkdtemp(
            prefix=f".{target_directory.name}.mrs3-stage-",
            dir=target_directory.parent,
        )
    )
    moved_existing = False
    installed = False
    try:
        for sheet_name, frame in tables.items():
            serializable_frame(frame).to_csv(
                staging / f"{sheet_name}.csv",
                index=False,
                float_format="%.17g",
                lineterminator="\n",
            )
        if target_directory.exists():
            if not target_directory.is_dir():
                raise RuntimeError(
                    f"audit CSV target is not a directory: {target_directory}"
                )
            target_directory.replace(backup)
            moved_existing = True
        staging.replace(target_directory)
        installed = True
        if moved_existing:
            try:
                shutil.rmtree(backup)
            except Exception:
                failed = target_directory.with_name(
                    f".{target_directory.name}.mrs3-failed"
                )
                if failed.exists():
                    raise RuntimeError(
                        f"audit CSV failed directory requires recovery: {failed}"
                    )
                target_directory.replace(failed)
                backup.replace(target_directory)
                shutil.rmtree(failed)
                installed = False
                raise
    except Exception:
        if moved_existing and backup.exists() and not target_directory.exists():
            backup.replace(target_directory)
        raise
    finally:
        if staging.exists():
            shutil.rmtree(staging)
        if installed and backup.exists():
            shutil.rmtree(backup)
