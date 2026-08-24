from __future__ import annotations

from dataclasses import dataclass, fields
from datetime import datetime, timezone
from decimal import Decimal, ROUND_HALF_UP
import json
from pathlib import Path
import shutil
import tempfile
import uuid

import duckdb
import pandas as pd
from openpyxl import load_workbook

from .config import AlgorithmConfig
from .models import Side
from .performance_metrics import PreciseMetricError, derive_precise_metrics
from .posttest import compare_posttest, write_posttest_outputs


@dataclass(frozen=True, slots=True)
class PerformanceDd5Artifacts:
    workbook: Path
    csv_directory: Path
    manifest: Path
    manifest_json: dict[str, object]
    dd5_run_id: str


_DECIMAL_CONFIG_FIELDS = {
    "history_min_days", "economic_min_pnl_pct", "economic_min_win_rate_pct",
    "economic_max_dd_pct", "economic_min_efficiency", "core_link_min",
    "plateau_envelope_min", "supported_link_min", "isolated_peak_relative",
    "equivalent_tolerance", "close_core_min", "close_supported_min",
    "numeric_tolerance", "initial_lot_sum", "target_dd_pct",
    "close_multiplier_long", "close_multiplier_short",
}

_PERSISTED_DECIMAL_PLACES = Decimal("0.000000000001")
_REQUIRED_WORKBOOK_SHEETS = frozenset({
    "00_Selection_Summary",
    "01_Finalists",
    "16_Raw_MRS3_Results",
    "17_DD5_Normalized",
    "18_Final_Comparison",
    "19_Position_Holding_Cycles",
    "20_Position_Holding_Exclusions",
})


def _config_json(config: AlgorithmConfig) -> str:
    value: dict[str, object] = {}
    for item in fields(config):
        current = getattr(config, item.name)
        if item.name == "side_columns":
            current = {key.value: columns for key, columns in current.items()}
        elif item.name == "shift_factors":
            current = [[boundary, str(factor)] for boundary, factor in current]
        elif item.name in _DECIMAL_CONFIG_FIELDS:
            current = str(current)
        elif item.name == "base_rates":
            current = {key: str(rate) for key, rate in current.items()}
        value[item.name] = current
    return json.dumps(value, sort_keys=True)


def _config_from_json(raw: str) -> AlgorithmConfig:
    value = json.loads(raw)
    defaults = AlgorithmConfig.defaults()
    kwargs = {item.name: value.get(item.name, getattr(defaults, item.name)) for item in fields(defaults)}
    kwargs["side_columns"] = {Side(key): dict(columns) for key, columns in kwargs["side_columns"].items()}
    kwargs["base_rates"] = {key: Decimal(str(rate)) for key, rate in kwargs["base_rates"].items()}
    kwargs["shift_factors"] = tuple((int(boundary), Decimal(str(factor))) for boundary, factor in kwargs["shift_factors"])
    for name in _DECIMAL_CONFIG_FIELDS:
        kwargs[name] = Decimal(str(kwargs[name]))
    return AlgorithmConfig(**kwargs)


def _settings_side(settings: object, stored_side: object) -> str:
    if not isinstance(settings, dict):
        raise ValueError("strategy settings must be a JSON object")
    if stored_side is not None:
        if not isinstance(stored_side, str):
            raise ValueError(
                f"strategy_versions.side has unsupported value: {stored_side!r}"
            )
        if stored_side.strip():
            candidate = stored_side.strip().upper()
            if candidate == "LONG":
                return "LONG"
            if candidate == "SHORT":
                return "SHORT"
            raise ValueError(
                f"strategy_versions.side has unsupported value: {stored_side!r}"
            )
    basic = settings.get("basic")
    if isinstance(basic, dict):
        if "use_long" in basic:
            return "LONG" if bool(basic["use_long"]) else "SHORT"
        side = basic.get("side")
        if side is not None:
            if not isinstance(side, str) or not side.strip():
                raise ValueError(f"strategy settings basic.side is invalid: {side!r}")
            candidate = side.strip().upper()
            if candidate == "LONG":
                return "LONG"
            if candidate == "SHORT":
                return "SHORT"
            raise ValueError(f"strategy settings basic.side is invalid: {side!r}")
    raise ValueError("strategy settings have no side")


def _settings_symbol(settings: object, stored_symbol: object) -> str | None:
    if stored_symbol is not None and str(stored_symbol).strip():
        return str(stored_symbol).strip()
    basic = settings.get("basic") if isinstance(settings, dict) else None
    if isinstance(basic, dict):
        symbol = basic.get("symbol")
        if isinstance(symbol, str) and symbol.strip():
            return symbol.strip()
    return None


def _settings_timeframe(settings: object, stored_timeframe: object) -> str | None:
    if stored_timeframe is not None and str(stored_timeframe).strip():
        return str(stored_timeframe).strip()
    basic = settings.get("basic") if isinstance(settings, dict) else None
    if isinstance(basic, dict):
        for key in ("time_frame", "timeframe"):
            timeframe = basic.get(key)
            if isinstance(timeframe, str) and timeframe.strip():
                return timeframe.strip()
    return None


def _settings_metadata(settings: object, side: str) -> dict[str, object]:
    if not isinstance(settings, dict):
        raise ValueError("strategy settings must be a JSON object")
    mrs3 = settings.get("mrs3")
    if not isinstance(mrs3, dict):
        raise ValueError("strategy settings have no mrs3 section")
    is_long = str(side).upper() == "LONG"
    orders = mrs3.get("ma_long" if is_long else "ma_short")
    if not isinstance(orders, list) or not orders:
        raise ValueError(f"strategy settings have no active MRS3 orders: {side}")
    shifts: list[str] = []
    has_shift_multipliers = True
    for order in orders:
        if not isinstance(order, dict):
            raise ValueError("strategy settings have an invalid MRS3 order")
        if "multiplier" not in order:
            has_shift_multipliers = False
            continue
        multiplier = Decimal(str(order["multiplier"]))
        change = Decimal("1") - multiplier if is_long else multiplier - Decimal("1")
        shifts.append(str(int((change * Decimal("10000")).to_integral_value(rounding=ROUND_HALF_UP))))
    close = mrs3.get("ma_close_long" if is_long else "ma_close_short")
    common_close_ma = int(close["len"]) if isinstance(close, dict) and "len" in close else None
    return {
        "lots": [Decimal(str(order["lot_x"])) for order in orders],
        "shift_bp_vector": " / ".join(shifts) if has_shift_multipliers else None,
        "side": "LONG" if is_long else "SHORT",
        "order_count": len(orders),
        "common_close_ma": common_close_ma,
        "first_shift_bp": int(shifts[0]) if has_shift_multipliers else None,
    }


_DECLARED_METRIC_NAMES = (
    ("Initial balance", "InitialBalance"),
    ("Final balance", "FinalBalance"),
    ("Total PnL", "TotalPnL"),
    ("Total PnL, %", "TotalPnLPercent"),
    ("Max Drawdown", "MaxDrawdown"),
    ("Max Drawdown, %", "MaxDrawdownPercent"),
    ("Win Rate, %", "WinRate"),
)


def _declared_metrics(
    encoded: object,
) -> dict[str, object] | None:
    if isinstance(encoded, str):
        try:
            decoded = json.loads(encoded)
        except json.JSONDecodeError:
            decoded = None
        if isinstance(decoded, dict) and all(
            any(name in decoded for name in aliases)
            for aliases in _DECLARED_METRIC_NAMES
        ):
            return decoded
    return None


def _read_rows(database: Path, import_id: str, *, dd5_run_id: str | None = None) -> list[dict[str, object]]:
    params: list[object] = [import_id]
    run_filter = ""
    if dd5_run_id is not None:
        run_filter = " and r.test_run_id in (select test_run_id from dd5_results where dd5_run_id = ?)"
        params.append(dd5_run_id)
    with duckdb.connect(str(database), read_only=True) as connection:
        run = connection.execute(
            "select status, quarantined_count from import_runs where import_id = ?",
            [import_id],
        ).fetchone()
        if run is None or run[0] != "COMMITTED" or run[1] != 0:
            raise ValueError("DD5 requires a committed import with zero quarantine")
        rows = connection.execute(
            f"""
            select r.test_run_id, s.strategy_name, s.symbol, s.side, s.timeframe,
                   s.settings_json, m.profit_factor, m.profit_factor_status,
                   m.total_trades, m.days_in_test, r.initial_balance,
                   m.win_trades, m.metrics_json, f.status
            from backtest_runs r
            join backtest_metrics m on m.test_run_id = r.test_run_id
            join strategy_versions s on s.strategy_version_id = r.strategy_version_id
            join import_files f on f.test_run_id = r.test_run_id
            where f.import_id = ?{run_filter}
            order by r.test_run_id
            """,
            params,
        ).fetchall()
        equity_rows = connection.execute(
            f"""
            select r.test_run_id, e.timestamp_utc, e.wallet, e.equity
            from backtest_runs r
            join import_files f on f.test_run_id = r.test_run_id
            join backtest_equity e on e.test_run_id = r.test_run_id
            where f.import_id = ?{run_filter}
            order by r.test_run_id, e.sample_index
            """,
            params,
        ).fetchall()
        action_rows = connection.execute(
            f"""
            select r.test_run_id, a.action_index, a.raw_action_json
            from backtest_runs r
            join import_files f on f.test_run_id = r.test_run_id
            left join backtest_actions a on a.test_run_id = r.test_run_id
            where f.import_id = ?{run_filter}
            order by r.test_run_id, a.action_index
            """,
            params,
        ).fetchall()
    rows = [row for row in rows if row[-1] in {"IMPORTED", "SKIPPED"}]
    if not rows:
        raise ValueError("committed import has no backtest results")
    actions_by_run: dict[str, list[dict[str, object]]] = {}
    equity_by_run: dict[str, list[tuple[object, Decimal, Decimal]]] = {}
    kept_run_ids = {str(row[0]) for row in rows}
    for test_run_id, _action_index, raw_action_json in action_rows:
        if raw_action_json is None or str(test_run_id) not in kept_run_ids:
            continue
        actions_by_run.setdefault(str(test_run_id), []).append(json.loads(raw_action_json))
    for test_run_id, timestamp, wallet, equity in equity_rows:
        if str(test_run_id) in kept_run_ids:
            equity_by_run.setdefault(str(test_run_id), []).append(
                (timestamp, Decimal(str(wallet)), Decimal(str(equity)))
            )
    result: list[dict[str, object]] = []
    for row in rows:
        (
            test_run_id,
            name,
            symbol,
            side,
            timeframe,
            settings_json,
            profit_factor,
            profit_factor_status,
            trades,
            days,
            initial_balance,
            win_trades,
            metrics_json,
            _status,
        ) = row
        settings = json.loads(settings_json)
        symbol = _settings_symbol(settings, symbol)
        side = _settings_side(settings, side)
        timeframe = _settings_timeframe(settings, timeframe)
        metadata = _settings_metadata(settings, side)
        evidence = equity_by_run.get(str(test_run_id))
        if not evidence:
            raise ValueError(f"missing equity evidence for test run: {test_run_id}")
        declared_metrics = _declared_metrics(metrics_json)
        if declared_metrics is None:
            raise ValueError(
                f"missing declared metrics diagnostics for {test_run_id}"
            )
        try:
            precise = derive_precise_metrics(
                initial_balance,
                [wallet for _, wallet, _ in evidence],
                [equity for _, _, equity in evidence],
                win_trades,
                trades,
                timestamps=[timestamp for timestamp, _, _ in evidence],
                declared_metrics=declared_metrics,
            )
        except PreciseMetricError as error:
            raise ValueError(
                f"invalid precise backtest metrics for {test_run_id}: {error}"
            ) from error
        result.append(
            {
                "test_run_id": test_run_id,
                "strategy_name": name,
                "symbol": symbol,
                "side": side,
                "timeframe": timeframe,
                "pnl_pct": precise.total_pnl_pct,
                "dd_pct": precise.max_drawdown_pct,
                "win_rate_pct": precise.win_rate_pct,
                "profit_factor": Decimal(str(profit_factor)) if profit_factor is not None else None,
                "profit_factor_status": str(profit_factor_status),
                "trades": int(trades),
                "effective_days": Decimal(str(days)),
                "trades_json": json.dumps(actions_by_run.get(str(test_run_id), []), ensure_ascii=False),
                "lots": metadata["lots"],
                "shift_bp_vector": metadata["shift_bp_vector"],
                "order_count": metadata["order_count"],
                "common_close_ma": metadata["common_close_ma"],
                "first_shift_bp": metadata["first_shift_bp"],
            }
        )
    return result


def _verify_dd5_readback(database: Path, dd5_run_id: str, import_id: str, expected_count: int) -> None:
    with duckdb.connect(str(database), read_only=True) as connection:
        run = connection.execute(
            "select import_id, status, input_test_count from dd5_runs where dd5_run_id = ?",
            [dd5_run_id],
        ).fetchone()
        result_count = connection.execute(
            "select count(*) from dd5_results where dd5_run_id = ?",
            [dd5_run_id],
        ).fetchone()[0]
    if run != (import_id, "CALCULATION_ONLY", expected_count) or result_count != expected_count:
        raise ValueError("DD5 readback verification failed")


def _read_persisted_results(database: Path, dd5_run_id: str, import_id: str, config: AlgorithmConfig):
    with duckdb.connect(str(database), read_only=True) as connection:
        run = connection.execute(
            "select import_id, status, input_test_count from dd5_runs where dd5_run_id = ?", [dd5_run_id]
        ).fetchone()
        rows = connection.execute(
            "select test_run_id, projected_pnl_dd5, projected_dd_pct, projected_pnl30_dd5, pareto_rank, pareto from dd5_results where dd5_run_id = ? order by test_run_id",
            [dd5_run_id],
        ).fetchall()
    if run is None or run[0] != import_id or run[1] != "CALCULATION_ONLY" or run[2] != len(rows) or not rows:
        raise ValueError("DD5 persisted run readback failed")
    raw_rows = _read_rows(database, import_id, dd5_run_id=dd5_run_id)
    if [row["test_run_id"] for row in raw_rows] != [row[0] for row in rows]:
        raise ValueError("DD5 persisted run readback failed")
    persisted: dict[str, tuple[object, ...]] = {}
    for test_run_id, projected_pnl, projected_dd, projected_pnl30, pareto_rank, pareto in rows:
        persisted[test_run_id] = (projected_pnl, projected_dd, projected_pnl30, pareto_rank, pareto)
    raw_frame = pd.DataFrame(raw_rows).drop(columns=["lots"])
    variants = pd.DataFrame([{"strategy_name": row["strategy_name"], "lots": row["lots"]} for row in raw_rows])
    tables = compare_posttest(raw_frame, variants, config)
    for row in tables.comparison.to_dict(orient="records"):
        stored = persisted[row["test_run_id"]]
        if stored[4] != bool(row["pareto"]) or stored[3] != int(row["near_tie_rank"]):
            raise ValueError("DD5 persisted result readback failed")
        for actual, expected in zip(stored[:3], (row["projected_pnl_dd5"], row["projected_dd_pct"], row["pnl30_dd5"]), strict=True):
            expected_at_storage_precision = Decimal(str(expected)).quantize(
                _PERSISTED_DECIMAL_PLACES, rounding=ROUND_HALF_UP
            )
            if Decimal(str(actual)) != expected_at_storage_precision:
                raise ValueError("DD5 persisted result readback failed")
    return tables


def _persist_dd5(
    database: Path, import_id: str, config: AlgorithmConfig
) -> tuple[str, object]:
    database = Path(database).resolve()
    rows = _read_rows(database, import_id)
    raw = pd.DataFrame(rows).drop(columns=["lots"])
    variants = pd.DataFrame(
        [{"strategy_name": row["strategy_name"], "lots": row["lots"]} for row in rows]
    )
    tables = compare_posttest(raw, variants, config)
    dd5_run_id = uuid.uuid4().hex
    with duckdb.connect(str(database)) as connection:
        connection.execute("begin transaction")
        try:
            connection.execute(
                "insert into dd5_runs values (?, ?, ?, ?, ?, ?, ?)",
                [
                    dd5_run_id,
                    import_id,
                    datetime.now(timezone.utc),
                    config.target_dd_pct,
                    _config_json(config),
                    len(tables.normalized),
                    "CALCULATION_ONLY",
                ],
            )
            for row in tables.comparison.to_dict(orient="records"):
                raw_json = {
                    key: ([str(value) for value in row[key]] if key == "lots" else str(row[key]) if isinstance(row[key], Decimal) else row[key])
                    for key in ("test_run_id", "strategy_name", "pnl_pct", "dd_pct", "win_rate_pct", "profit_factor", "profit_factor_status", "trades", "effective_days", "lots")
                }
                connection.execute(
                    "insert into dd5_results (dd5_run_id, test_run_id, projected_pnl_dd5, projected_dd_pct, projected_pnl30_dd5, scaled_lots_json, capital_requirement_proxy, holding_filter, pareto_rank, raw_json, pareto) values (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                    [
                        dd5_run_id,
                        row["test_run_id"],
                        row["projected_pnl_dd5"],
                        row["projected_dd_pct"],
                        row["pnl30_dd5"],
                        json.dumps([str(item) for item in row["scaled_lots"]]),
                        row["capital_requirement_proxy"],
                        None,
                        int(row["near_tie_rank"]),
                        json.dumps(raw_json, sort_keys=True),
                        bool(row["pareto"]),
                    ],
                )
            connection.execute("commit")
        except Exception:
            connection.execute("rollback")
            raise

    _verify_dd5_readback(database, dd5_run_id, import_id, len(tables.normalized))
    tables = _read_persisted_results(database, dd5_run_id, import_id, config)
    return dd5_run_id, tables


def _manifest_json(
    database: Path, import_id: str, dd5_run_id: str, tables: object, config: AlgorithmConfig
) -> dict[str, object]:
    return {
        "database": database.name,
        "import_id": import_id,
        "dd5_run_id": dd5_run_id,
        "raw_result_count": len(tables.raw),
        "profit_factor_unavailable_count": int((tables.raw["profit_factor_status"] != "AVAILABLE").sum()),
        "pareto_count": int(tables.comparison["pareto"].sum()),
        "target_dd_pct": str(config.target_dd_pct),
        "dd5_mode": "CALCULATION_ONLY",
        "scaled_strategy_count": 0,
        "scaled_strategies_require_retest": False,
    }


def run_performance_dd5(
    database: Path,
    import_id: str,
    output_dir: Path,
    config: AlgorithmConfig,
) -> PerformanceDd5Artifacts:
    database = Path(database).resolve()
    if Path(output_dir).suffix.casefold() == ".xlsx":
        return run_performance_dd5_atomic(database, import_id, output_dir, config)
    dd5_run_id, tables = _persist_dd5(database, import_id, config)
    return regenerate_performance_dd5(database, dd5_run_id, output_dir)


def _verify_committed_import(database: Path, import_id: str) -> None:
    try:
        with duckdb.connect(str(database), read_only=True) as connection:
            schema = connection.execute(
                "select value from schema_info where key = 'schema_version'"
            ).fetchone()
            row = connection.execute(
                "select status, quarantined_count from import_runs where import_id = ?",
                [import_id],
            ).fetchone()
    except Exception as error:
        raise ValueError("DD5 requires a readable Performance DB") from error
    if schema != ("1",) or row != ("COMMITTED", 0):
        raise ValueError("DD5 requires committed zero-quarantine import evidence")


def _verify_workbook(path: Path) -> None:
    workbook = None
    try:
        workbook = load_workbook(path, read_only=True)
        actual = frozenset(workbook.sheetnames)
    except Exception as error:
        raise ValueError("DD5 workbook readback failed") from error
    finally:
        if workbook is not None:
            workbook.close()
    if not _REQUIRED_WORKBOOK_SHEETS.issubset(actual):
        raise ValueError("DD5 workbook readback failed")


def run_performance_dd5_atomic(
    database: Path,
    import_id: str,
    workbook_path: Path,
    config: AlgorithmConfig,
) -> PerformanceDd5Artifacts:
    """Persist DD5, stage the workbook, then replace only its fixed target."""
    database = Path(database).resolve()
    target = Path(workbook_path).resolve()
    _verify_committed_import(database, import_id)
    dd5_run_id, tables = _persist_dd5(database, import_id, config)
    target.parent.mkdir(parents=True, exist_ok=True)
    staging = Path(tempfile.mkdtemp(prefix=f".{target.stem}.stage-", dir=target.parent))
    csv_target = target.with_name(f"{target.stem}.csv")
    manifest_path = target.with_name(f"{target.stem}.manifest.json")
    manifest_json = _manifest_json(database, import_id, dd5_run_id, tables, config)
    try:
        write_posttest_outputs(tables, staging)
        staged_workbook = staging / "posttest.xlsx"
        _verify_workbook(staged_workbook)
        staged_workbook.replace(target)
        _verify_workbook(target)
        staged_csv = staging / "posttest_csv"
        if csv_target.exists():
            shutil.rmtree(csv_target)
        staged_csv.replace(csv_target)
        manifest_path.write_text(
            json.dumps(manifest_json, indent=2, sort_keys=True) + "\n", encoding="utf-8"
        )
    finally:
        if staging.exists():
            shutil.rmtree(staging)
    return PerformanceDd5Artifacts(target, csv_target, manifest_path, manifest_json, dd5_run_id)


def regenerate_performance_dd5(
    database: Path, dd5_run_id: str, output_dir: Path
) -> PerformanceDd5Artifacts:
    database = Path(database).resolve()
    with duckdb.connect(str(database), read_only=True) as connection:
        run = connection.execute(
            "select import_id, config_json from dd5_runs where dd5_run_id = ?",
            [dd5_run_id],
        ).fetchone()
    if run is None:
        raise ValueError(f"unknown DD5 run: {dd5_run_id}")
    import_id, config_json = run
    config = _config_from_json(config_json)
    tables = _read_persisted_results(database, dd5_run_id, import_id, config)
    manifest_json: dict[str, object] = {
        "database": database.name,
        "import_id": import_id,
        "dd5_run_id": dd5_run_id,
        "raw_result_count": len(tables.raw),
        "profit_factor_unavailable_count": int((tables.raw["profit_factor_status"] != "AVAILABLE").sum()),
        "pareto_count": int(tables.comparison["pareto"].sum()),
        "target_dd_pct": str(config.target_dd_pct),
        "dd5_mode": "CALCULATION_ONLY",
        "scaled_strategy_count": 0,
        "scaled_strategies_require_retest": False,
    }
    output_dir = Path(output_dir).resolve()
    output_dir.mkdir(parents=True, exist_ok=True)
    write_posttest_outputs(tables, output_dir)
    manifest = output_dir / "posttest_manifest.json"
    manifest.write_text(json.dumps(manifest_json, indent=2, sort_keys=True) + "\n", encoding="utf-8")
    return PerformanceDd5Artifacts(output_dir / "posttest.xlsx", output_dir / "posttest_csv", manifest, manifest_json, dd5_run_id)
