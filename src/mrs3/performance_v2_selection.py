"""Typed, closed request/config contract for Performance v2 finalist selection."""

from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime, timedelta, timezone
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP, localcontext
from concurrent.futures import ThreadPoolExecutor
from colorsys import hls_to_rgb
import json
from pathlib import Path
from typing import Callable, Literal, Mapping, Sequence

import duckdb
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation

from .performance_v2_windows import (
    METRICS_VERSION, WindowMetrics, _METRIC_COLUMNS, _cached, _calculate,
    _load_source, _metric_from_row, _persist, calendar_window_days, get_or_calculate_window, get_or_calculate_window_pair,
)
from .audit import _normalize_xlsx_archive, write_audit_workbook

StageScope = Literal["pair_side", "pair_side_timeframe"]

_STAGE_IDS = frozenset((
    "filter_lot_variant_redundancy",
    "filter_holding_outlier",
    "filter_low_trades",
    "filter_min_shift",
    "ab_deterioration",
    "pareto_window_b",
    "pareto_window_b_dd_shift",
    "pareto_dd5_balanced",
    "pareto_plateau_points_per_order",
    "pareto_plateau_points_total",
    "pareto_efficiency_shift",
    "pareto_dd5_holding",
    "pareto_dd5_close_ma",
    "pareto_dd5_first_shift",
    "pareto_conditional_close_ma",
    "pareto_primary",
    "pareto_dd5_capital",
    "filter_best_trade_dependency",
    "filter_time_consistency",
    "pareto_robust",
    "pareto_shift_near_tie",
    "pareto_close_ma_near_tie",
    "rank_robust_top_n",
))
_SCOPES = frozenset(("pair_side", "pair_side_timeframe"))
_CANDIDATE_COLUMNS = (
    "strategy_id", "strategy_name", "symbol", "side", "timeframe", "close_ma_len", "order_count",
    "result_id", "total_pnl", "total_pnl_pct", "max_drawdown", "max_drawdown_pct", "total_fees",
    "total_trades", "trades_30d", "pnl_30d_pct", "profit_factor", "win_rate_pct", "risk_scale", "dd5_proxy", "holding_p95_minutes",
    "holding_median_minutes",
    "report_start_utc", "report_end_utc", "reported_start_utc", "reported_end_utc",
    "effective_start_utc", "effective_end_utc",
    "ab_pnl_change_30d_pct", "ab_return_b_pct", "first_shift_bp", "scaled_lot_sum", "capital_proxy",
    "capital_efficiency", "total_plateau_point_count",
    "ab_return_a_30d_pct", "ab_return_b_30d_pct", "ab_calendar_days_a", "ab_calendar_days_b", "ab_win_rate_b_pct",
    "ab_trade_rate_a_30d", "ab_trade_rate_b_30d", "ab_drawdown_b_pct", "ab_holding_p95_minutes",
    "best_trade_profit_share_pct", "pnl_without_best_trade", "pnl_without_best_trade_pct", "completed_profitable_trade_count", "best_trade_reliable",
    "positive_quarter_count", "positive_quarter_available_count", "positive_quarter_status", "robust_pnl_30d_pct", "worst_drawdown_pct", "worst_holding_p95_minutes",
    "ab_stability_ratio", "minimum_plateau_point_count",
    "lot_variant_group_key", "lot_variant_representative_strategy_id",
    "rank_quality_robust_pnl", "rank_quality_worst_drawdown", "rank_quality_ab_stability", "rank_quality_worst_holding",
    "rank_quality_first_shift", "rank_quality_minimum_plateau_points", "rank_quality_close_ma", "rank_weight_coverage_pct",
    "rank_weight_robust_pnl", "rank_weight_worst_drawdown", "rank_weight_ab_stability", "rank_weight_worst_holding",
    "rank_weight_first_shift", "rank_weight_minimum_plateau_points", "rank_weight_close_ma", "final_score", "final_rank",
    *(f"order_{order}_{field}" for order in range(1, 5) for field in (
        "open_ma_len", "open_multiplier", "shift_bp", "lot_x", "plateau_point_count", "plateau_key",
    )),
)


class PerformanceV2SelectionError(ValueError):
    """Stable error code for an invalid finalist-selection request/config."""


@dataclass(frozen=True, slots=True)
class SelectionStage:
    id: str
    enabled: bool
    scope: StageScope
    min_shift_pct: Decimal | None = None
    pnl_tolerance_pct: Decimal | None = None
    top_n: int | None = None


@dataclass(frozen=True, slots=True)
class SelectionRequest:
    symbol: str
    side: Literal["LONG", "SHORT"]
    stages: tuple[SelectionStage, ...]


@dataclass(frozen=True, slots=True)
class SelectionConfig:
    ab_final_days: int = 14
    ab_return_floor_pct: Decimal = Decimal("5")
    ab_return_divisor: Decimal = Decimal("10")
    ab_win_rate_floor_pct: Decimal = Decimal("58")
    ab_trade_rate_divisor: Decimal = Decimal("7")
    plateau_points_pareto_pnl_multiplier: Decimal = Decimal("2")
    best_trade_max_profit_share_pct: Decimal = Decimal("35")
    best_trade_min_profitable_trades: int = 4
    shift_near_tie_min_advantage_bp: int = 10
    lot_variant_redundancy_enabled: bool = True


def _error(code: str) -> PerformanceV2SelectionError:
    return PerformanceV2SelectionError(code)


def parse_selection_request(payload: Mapping[str, object]) -> SelectionRequest:
    if set(payload) != {"symbol", "side", "stages"}:
        raise _error("INVALID_REQUEST")
    symbol = payload["symbol"]
    side = payload["side"]
    stages = payload["stages"]
    if not isinstance(symbol, str) or not symbol.strip():
        raise _error("INVALID_SYMBOL")
    if side not in {"LONG", "SHORT"}:
        raise _error("INVALID_SIDE")
    if not isinstance(stages, list):
        raise _error("INVALID_STAGES")

    parsed: list[SelectionStage] = []
    seen: set[str] = set()
    for raw in stages:
        if not isinstance(raw, Mapping):
            raise _error("INVALID_STAGE")
        stage_id = raw.get("id")
        expected = {"id", "enabled", "scope"}
        if stage_id == "filter_min_shift":
            expected.add("min_shift_pct")
        if stage_id in {"pareto_shift_near_tie", "pareto_close_ma_near_tie"}:
            expected.add("pnl_tolerance_pct")
        if stage_id == "rank_robust_top_n":
            expected.add("top_n")
        if set(raw) != expected:
            raise _error("INVALID_STAGE")
        enabled, scope = raw["enabled"], raw["scope"]
        if not isinstance(stage_id, str) or stage_id not in _STAGE_IDS:
            raise _error("UNKNOWN_STAGE")
        if stage_id in seen:
            raise _error("DUPLICATE_STAGE")
        if not isinstance(enabled, bool):
            raise _error("INVALID_STAGE")
        if scope not in _SCOPES:
            raise _error("INVALID_SCOPE")
        min_shift_pct = _positive_decimal(raw["min_shift_pct"], "min_shift_pct") if stage_id == "filter_min_shift" and enabled else None
        pnl_tolerance_pct = _bounded_decimal(raw["pnl_tolerance_pct"], "pnl_tolerance_pct", Decimal(0), Decimal(100)) if stage_id in {"pareto_shift_near_tie", "pareto_close_ma_near_tie"} and enabled else None
        top_n = _positive_int(raw["top_n"], "top_n") if stage_id == "rank_robust_top_n" else None
        if stage_id == "rank_robust_top_n" and scope != "pair_side":
            raise _error("RANK_STAGE_SCOPE")
        if stage_id == "filter_lot_variant_redundancy" and scope != "pair_side_timeframe":
            raise _error("LOT_VARIANT_STAGE_SCOPE")
        seen.add(stage_id)
        parsed.append(SelectionStage(stage_id, enabled, scope, min_shift_pct, pnl_tolerance_pct, top_n))
    if any(stage.id == "rank_robust_top_n" for stage in parsed) and parsed[-1].id != "rank_robust_top_n":
        raise _error("RANK_STAGE_MUST_BE_LAST")
    return SelectionRequest(symbol.strip(), side, tuple(parsed))


def _positive_decimal(value: object, name: str) -> Decimal:
    if isinstance(value, bool):
        raise _error(f"INVALID_CONFIG_{name}")
    try:
        parsed = Decimal(str(value))
    except (InvalidOperation, ValueError):
        raise _error(f"INVALID_CONFIG_{name}") from None
    if not parsed.is_finite() or parsed <= 0:
        raise _error(f"INVALID_CONFIG_{name}")
    return parsed


def _bounded_decimal(value: object, name: str, lower: Decimal, upper: Decimal) -> Decimal:
    if isinstance(value, bool):
        raise _error(f"INVALID_CONFIG_{name}")
    try:
        parsed = Decimal(str(value))
    except (InvalidOperation, ValueError):
        raise _error(f"INVALID_CONFIG_{name}") from None
    if not parsed.is_finite() or parsed < lower or parsed >= upper:
        raise _error(f"INVALID_CONFIG_{name}")
    return parsed


def _positive_int(value: object, name: str) -> int:
    if isinstance(value, bool) or not isinstance(value, int) or value < 1:
        raise _error(f"INVALID_CONFIG_{name}")
    return value


def load_selection_config(path: Path) -> SelectionConfig:
    try:
        raw = json.loads(Path(path).read_text(encoding="utf-8"))
        section = raw["unified_performance_v2"]
    except (OSError, UnicodeDecodeError, json.JSONDecodeError, KeyError, TypeError):
        raise _error("INVALID_CONFIG") from None
    if not isinstance(section, Mapping):
        raise _error("INVALID_CONFIG")
    selected = section.get("finalist_selection", {})
    if not isinstance(selected, Mapping):
        raise _error("INVALID_CONFIG")
    final_days = selected.get("ab_final_days", 14)
    if isinstance(final_days, bool) or not isinstance(final_days, int) or final_days < 1:
        raise _error("INVALID_CONFIG_ab_final_days")
    best_trade_min = selected.get("best_trade_min_profitable_trades", 4)
    shift_advantage = selected.get("shift_near_tie_min_advantage_bp", 10)
    lot_variant_enabled = selected.get("lot_variant_redundancy_enabled", True)
    if not isinstance(lot_variant_enabled, bool):
        raise _error("INVALID_CONFIG_lot_variant_redundancy_enabled")
    return SelectionConfig(
        ab_final_days=final_days,
        ab_return_floor_pct=_positive_decimal(selected.get("ab_return_floor_pct", 5), "ab_return_floor_pct"),
        ab_return_divisor=_positive_decimal(selected.get("ab_return_divisor", 10), "ab_return_divisor"),
        ab_win_rate_floor_pct=_positive_decimal(selected.get("ab_win_rate_floor_pct", 58), "ab_win_rate_floor_pct"),
        ab_trade_rate_divisor=_positive_decimal(selected.get("ab_trade_rate_divisor", 7), "ab_trade_rate_divisor"),
        plateau_points_pareto_pnl_multiplier=_positive_decimal(
            selected.get("plateau_points_pareto_pnl_multiplier", 2),
            "plateau_points_pareto_pnl_multiplier",
        ),
        best_trade_max_profit_share_pct=_bounded_decimal(
            selected.get("best_trade_max_profit_share_pct", 35), "best_trade_max_profit_share_pct", Decimal(0), Decimal(100)
        ),
        best_trade_min_profitable_trades=_positive_int(best_trade_min, "best_trade_min_profitable_trades"),
        shift_near_tie_min_advantage_bp=_positive_int(shift_advantage, "shift_near_tie_min_advantage_bp"),
        lot_variant_redundancy_enabled=lot_variant_enabled,
    )


def _decimal_or_none(value: object) -> Decimal | None:
    if value is None:
        return None
    try:
        result = Decimal(str(value))
    except (InvalidOperation, TypeError, ValueError):
        return None
    return result if result.is_finite() else None


def _holding_quantiles_minutes(
    connection: duckdb.DuckDBPyConnection, request: SelectionRequest
) -> dict[int, tuple[Decimal, Decimal]]:
    rows = connection.execute(
        """with actions as (
                 select a.result_id, a.timestamp_utc, a.action_index, lower(a.action) as kind,
                        a.post_size, lower(a.post_side) as post_side
                   from strategy_actions a
                   join strategy_results r on r.result_id = a.result_id
                   join strategies s on s.strategy_id = r.strategy_id and s.current_result_id = r.result_id
                  where s.lifecycle_status = 'ACTIVE' and s.symbol = ? and s.side = ?
                    and lower(a.action) in ('opened', 'increased', 'decreased', 'closed')
             ), numbered as (
                 select *, sum(case when kind = 'opened' and post_size > 0 and post_side in ('long', 'short') then 1 else 0 end)
                    over (partition by result_id order by timestamp_utc, action_index rows unbounded preceding) as position_number
                   from actions
             ), intervals as (
                 select result_id, position_number,
                        min(timestamp_utc) filter (where kind = 'opened' and post_size > 0 and post_side in ('long', 'short')) as opened_at,
                        min(timestamp_utc) filter (where kind = 'closed' and post_size = 0) as closed_at
                   from numbered
                  group by result_id, position_number
             )
             select result_id,
                    cast(quantile_cont(date_diff('second', opened_at, closed_at) / cast(60 as decimal(20, 6)), .95) as decimal(38, 6)),
                    cast(quantile_cont(date_diff('second', opened_at, closed_at) / cast(60 as decimal(20, 6)), .5) as decimal(38, 6))
               from intervals
              where opened_at is not null and closed_at is not null and closed_at >= opened_at
              group by result_id""",
        [request.symbol, request.side],
    ).fetchall()
    return {
        int(result_id): (Decimal(str(p95)), Decimal(str(median)))
        for result_id, p95, median in rows if p95 is not None and median is not None
    }


def _holding_p95_minutes(
    connection: duckdb.DuckDBPyConnection, request: SelectionRequest
) -> dict[int, Decimal]:
    return {result_id: values[0] for result_id, values in _holding_quantiles_minutes(connection, request).items()}


def _window_b_holding_p95_minutes(
    connection: duckdb.DuckDBPyConnection, request: SelectionRequest, config: SelectionConfig
) -> dict[int, Decimal]:
    rows = connection.execute(
        """with actions as (
                 select a.result_id, a.timestamp_utc, a.action_index, lower(a.action) as kind, a.post_size, lower(a.post_side) as post_side,
                        r.report_end_utc - (? * interval '1 day') as b_start, r.report_end_utc
                   from strategy_actions a
                   join strategy_results r on r.result_id = a.result_id
                   join strategies s on s.strategy_id = r.strategy_id and s.current_result_id = r.result_id
                  where s.lifecycle_status = 'ACTIVE' and s.symbol = ? and s.side = ?
                    and lower(a.action) in ('opened', 'increased', 'decreased', 'closed')
             ), numbered as (
                 select *, sum(case when kind = 'opened' and post_size > 0 and post_side in ('long', 'short') then 1 else 0 end)
                    over (partition by result_id order by timestamp_utc, action_index rows unbounded preceding) as position_number
                   from actions
             ), intervals as (
                 select result_id, position_number, min(b_start) as b_start, min(report_end_utc) as report_end_utc,
                        min(timestamp_utc) filter (where kind = 'opened' and post_size > 0 and post_side in ('long', 'short')) as opened_at,
                        min(timestamp_utc) filter (where kind = 'closed' and post_size = 0) as closed_at
                   from numbered group by result_id, position_number
             )
             select result_id, cast(quantile_cont(date_diff('second', opened_at, closed_at) / cast(60 as decimal(20, 6)), .95) as decimal(38, 6))
               from intervals
              where opened_at is not null and closed_at between b_start and report_end_utc and closed_at >= opened_at
              group by result_id""",
        [config.ab_final_days, request.symbol, request.side],
    ).fetchall()
    return {int(result_id): Decimal(str(p95)) for result_id, p95 in rows if p95 is not None}


def _best_trade_facts(
    connection: duckdb.DuckDBPyConnection, request: SelectionRequest
) -> dict[int, tuple[Decimal | None, Decimal | None, int | None, bool]]:
    rows = connection.execute(
        """with actions as (
                 select a.result_id, a.timestamp_utc, a.action_index, lower(a.action) as kind,
                        a.post_size, lower(a.post_side) as post_side, a.pnl, lower(s.side) as expected_side,
                        case when a.post_size > 0 and lower(a.post_side) in ('long', 'short')
                                  and lower(a.post_side) <> lower(s.side) then 1 else 0 end as side_flip
                   from strategy_actions a
                   join strategy_results r on r.result_id = a.result_id
                   join strategies s on s.strategy_id = r.strategy_id and s.current_result_id = r.result_id
                  where s.lifecycle_status = 'ACTIVE' and s.symbol = ? and s.side = ?
                    and lower(a.action) in ('opened', 'increased', 'decreased', 'closed')
             ), numbered as (
                 select *, sum(case when kind = 'opened' and post_size > 0 and post_side = expected_side then 1 else 0 end)
                    over (partition by result_id order by timestamp_utc, action_index rows unbounded preceding) as position_number
                   from actions
             ), trips as (
                 select result_id, position_number,
                        max(side_flip) as side_flip,
                        max(case when kind = 'closed' and post_size = 0 then 1 else 0 end) as completed,
                        sum(pnl) filter (where kind in ('decreased', 'closed')) as trip_pnl
                   from numbered
                  group by result_id, position_number
             ), completed as (
                 select result_id, trip_pnl
                   from trips
                  where position_number > 0 and completed = 1 and side_flip = 0 and trip_pnl is not null
             ), reliable as (
                 select result_id, max(side_flip) = 0 as reliable from numbered group by result_id
             ), summary as (
                 select result_id,
                        max(trip_pnl) filter (where trip_pnl > 0) as best_trade_pnl,
                        sum(trip_pnl) as completed_pnl,
                        sum(case when trip_pnl > 0 then trip_pnl else 0 end) as gross_positive_pnl,
                        count(*) filter (where trip_pnl > 0) as positive_count
                   from completed group by result_id
             )
             select reliable.result_id, summary.best_trade_pnl, summary.completed_pnl,
                    summary.gross_positive_pnl, summary.positive_count, reliable.reliable
               from reliable left join summary using (result_id)""",
        [request.symbol, request.side],
    ).fetchall()
    facts: dict[int, tuple[Decimal | None, Decimal | None, int | None, bool]] = {}
    for result_id, best, total, gross, count, reliable in rows:
        best_value, total_value, gross_value = (_decimal_or_none(value) for value in (best, total, gross))
        if not reliable or best_value is None or total_value is None or gross_value is None or gross_value <= 0:
            facts[int(result_id)] = (None, None, None, bool(reliable))
            continue
        facts[int(result_id)] = (best_value / gross_value * 100, total_value - best_value, int(count), True)
    return facts


def _return_30d(
    metrics: WindowMetrics,
    report_start_utc: datetime | None = None,
    report_end_utc: datetime | None = None,
) -> Decimal | None:
    if not metrics.available or metrics.growth_factor is None:
        return None
    elapsed = calendar_window_days(metrics, report_start_utc, report_end_utc)
    if elapsed is None:
        return None
    if elapsed < 1 or metrics.growth_factor < 0:
        return None
    if metrics.growth_factor == 0:
        return Decimal(-100)
    try:
        with localcontext() as context:
            context.prec = 34
            return ((Decimal(30) * metrics.growth_factor.ln() / elapsed).exp() - 1) * 100
    except (ArithmeticError, ValueError):
        return None


def _trade_rate_30d(
    metrics: WindowMetrics,
    report_start_utc: datetime | None = None,
    report_end_utc: datetime | None = None,
) -> Decimal | None:
    if not metrics.available or metrics.trade_count is None:
        return None
    days = calendar_window_days(metrics, report_start_utc, report_end_utc)
    if days is None:
        return None
    return Decimal(metrics.trade_count) * 30 / days if days >= 1 else None


def _ab_metrics(
    connection: duckdb.DuckDBPyConnection,
    result_id: int,
    report_start: datetime,
    report_end: datetime,
    config: SelectionConfig,
) -> dict[str, Decimal | None]:
    split = report_end - timedelta(days=config.ab_final_days)
    if split <= report_start:
        return _empty_ab_metrics()
    metrics_a, metrics_b = get_or_calculate_window_pair(
        connection, result_id, (report_start, split), (split, report_end)
    )
    return _ab_metrics_from_windows(metrics_a, metrics_b, report_start, report_end)


def _ab_metrics_from_windows(
    metrics_a: WindowMetrics,
    metrics_b: WindowMetrics,
    report_start_utc: datetime | None = None,
    report_end_utc: datetime | None = None,
) -> dict[str, Decimal | None]:
    return_a = _return_30d(metrics_a, report_start_utc, report_end_utc)
    return_b = _return_30d(metrics_b, report_start_utc, report_end_utc)
    return {
        "ab_pnl_change_30d_pct": None if return_a is None or return_b is None or return_a <= 0 else (return_b / return_a - 1) * 100,
        "ab_return_b_pct": metrics_b.return_pct,
        "ab_return_a_30d_pct": return_a, "ab_calendar_days_a": calendar_window_days(metrics_a, report_start_utc, report_end_utc),
        "ab_return_b_30d_pct": return_b, "ab_calendar_days_b": calendar_window_days(metrics_b, report_start_utc, report_end_utc),
        "ab_win_rate_b_pct": metrics_b.win_rate_pct, "ab_trade_rate_a_30d": _trade_rate_30d(metrics_a, report_start_utc, report_end_utc),
        "ab_trade_rate_b_30d": _trade_rate_30d(metrics_b, report_start_utc, report_end_utc), "ab_drawdown_b_pct": metrics_b.max_drawdown_pct,
    }


def _consistency_windows(report_start: datetime, report_end: datetime) -> tuple[tuple[datetime, datetime], ...]:
    span = report_end - report_start
    if span >= timedelta(days=28):
        count = 4
    elif span >= timedelta(days=21):
        count = 3
    else:
        return ()
    return tuple(
        (report_start + span * index / count, report_end if index == count - 1 else report_start + span * (index + 1) / count)
        for index in range(count)
    )


def _selection_windows(report_start: datetime, report_end: datetime, config: SelectionConfig) -> tuple[tuple[datetime, datetime], ...]:
    split = report_end - timedelta(days=config.ab_final_days)
    windows = [(report_start, report_end)]
    if split > report_start:
        windows.extend(((report_start, split), (split, report_end)))
    windows.extend(_consistency_windows(report_start, report_end))
    return tuple(dict.fromkeys(windows))


def _empty_ab_metrics() -> dict[str, Decimal | None]:
    return {key: None for key in (
        "ab_pnl_change_30d_pct", "ab_return_b_pct", "ab_return_a_30d_pct", "ab_calendar_days_a", "ab_calendar_days_b", "ab_return_b_30d_pct",
        "ab_win_rate_b_pct", "ab_trade_rate_a_30d", "ab_trade_rate_b_30d", "ab_drawdown_b_pct",
    )}


def _selection_cached_metrics(connection: duckdb.DuckDBPyConnection, request: SelectionRequest) -> dict[tuple[int, datetime, datetime], WindowMetrics]:
    rows = connection.execute(
        "select " + ", ".join(f"wm.{column}" for column in _METRIC_COLUMNS) +
        " from window_metrics wm"
        " join strategy_results r on r.result_id = wm.result_id"
        " join strategies s on s.strategy_id = r.strategy_id and s.current_result_id = r.result_id"
        " where s.lifecycle_status = 'ACTIVE' and s.symbol = ? and s.side = ? and wm.metrics_version = ?",
        [request.symbol, request.side, METRICS_VERSION],
    ).fetchall()
    metrics = (_metric_from_row(row) for row in rows)
    return {(metric.result_id, metric.requested_start_utc, metric.requested_end_utc): metric for metric in metrics}


def _cached_selection_metrics(
    connection: duckdb.DuckDBPyConnection,
    result_id: int,
    report_start: datetime,
    report_end: datetime,
    config: SelectionConfig,
    cached_metrics: Mapping[tuple[int, datetime, datetime], WindowMetrics] | None = None,
) -> tuple[WindowMetrics | None, dict[str, Decimal | None]]:
    cached = (
        (lambda start, end: cached_metrics.get((result_id, start, end)))
        if cached_metrics is not None
        else (lambda start, end: _cached(connection, result_id, start, end, METRICS_VERSION))
    )
    full = cached(report_start, report_end)
    split = report_end - timedelta(days=config.ab_final_days)
    if split <= report_start:
        return (full, _empty_ab_metrics())
    a = cached(report_start, split)
    b = cached(split, report_end)
    if a is None or b is None:
        return (full, _empty_ab_metrics())
    return (full, _ab_metrics_from_windows(a, b, report_start, report_end))


def _consistency_summary(
    metrics: Sequence[WindowMetrics | None], windows: Sequence[tuple[datetime, datetime]],
) -> tuple[int | None, int | None, str]:
    windows = tuple(windows)
    window_count = len(windows)
    if not window_count or len(metrics) != window_count:
        return None, None, "UNAVAILABLE"
    positive = 0
    assessed = 0
    for metric, (window_start, window_end) in zip(metrics, windows):
        if metric is None:
            return None, None, "UNAVAILABLE"
        if metric.availability_status == "NO_TRADES" or metric.unavailable_reason == "NO_TRADES":
            continue
        if not metric.available:
            return None, None, "UNAVAILABLE"
        value = _return_30d(metric, window_start, window_end)
        if value is None:
            return None, None, "UNAVAILABLE"
        assessed += 1
        positive += value > 0
    if assessed == 0:
        return None, 0, "UNAVAILABLE"
    threshold = 3 if window_count == 4 else 2
    return positive, assessed, "PASS" if positive >= threshold else "FAIL"


def _cached_positive_quarters(
    result_id: int, report_start: datetime, report_end: datetime, config: SelectionConfig,
    cached_metrics: Mapping[tuple[int, datetime, datetime], WindowMetrics],
) -> tuple[int | None, int | None, str]:
    windows = _consistency_windows(report_start, report_end)
    metrics = [cached_metrics.get((result_id, start, end)) for start, end in windows]
    return _consistency_summary(metrics, windows)


def _selection_window_job(database: str, result_id: int, report_start: datetime, report_end: datetime, final_days: int) -> tuple[WindowMetrics, ...]:
    """Compute default selection windows through a read-only worker connection."""
    windows = _selection_windows(report_start, report_end, SelectionConfig(ab_final_days=final_days))
    with duckdb.connect(database, read_only=True) as connection:
        connection.execute("set threads to 1")
        cached = [_cached(connection, result_id, start, end, METRICS_VERSION) for start, end in windows]
        if all(cached):
            return tuple(cached)
        source = _load_source(connection, result_id)
        return tuple(
            metric if metric is not None else _calculate(result_id, start, end, METRICS_VERSION, *source)
            for metric, (start, end) in zip(cached, windows)
        )


def _selection_window_job_from_args(args: tuple[str, int, datetime, datetime, int]) -> tuple[WindowMetrics, ...]:
    return _selection_window_job(*args)


def _selection_cache_missing_strategy_ids(
    connection: duckdb.DuckDBPyConnection, request: SelectionRequest, config: SelectionConfig,
) -> tuple[int, ...]:
    rows = connection.execute(
        """select s.strategy_id, r.result_id, r.report_start_utc, r.report_end_utc from strategies s
             join strategy_results r on r.result_id = s.current_result_id and r.strategy_id = s.strategy_id
            where s.lifecycle_status = 'ACTIVE' and s.symbol = ? and s.side = ?
            order by s.strategy_id""",
        [request.symbol, request.side],
    ).fetchall()
    cached_metrics = _selection_cached_metrics(connection, request)
    missing: list[int] = []
    for strategy_id, result_id, report_start, report_end in rows:
        windows = _selection_windows(report_start, report_end, config)
        if any(cached_metrics.get((int(result_id), window_start, window_end)) is None for window_start, window_end in windows):
            missing.append(int(strategy_id))
    return tuple(missing)


def selection_cache_missing_strategy_ids(
    connection: duckdb.DuckDBPyConnection, request: SelectionRequest, config: SelectionConfig,
) -> tuple[int, ...]:
    """Return active strategies whose current result lacks a required cache window."""
    return _selection_cache_missing_strategy_ids(connection, request, config)


def prepare_selection_window_cache(
    database: Path, request: SelectionRequest, config: SelectionConfig, workers: int,
    strategy_ids: Sequence[int] | None = None,
) -> None:
    """Warm default windows in independent readers, then persist them through one writer."""
    selected_ids = None if strategy_ids is None else tuple(dict.fromkeys(int(strategy_id) for strategy_id in strategy_ids))
    if selected_ids == ():
        return
    with duckdb.connect(str(database), read_only=True) as connection:
        where = "where s.lifecycle_status = 'ACTIVE' and s.symbol = ? and s.side = ?"
        parameters: list[object] = [request.symbol, request.side]
        if selected_ids is not None:
            where += " and s.strategy_id in (" + ",".join("?" for _ in selected_ids) + ")"
            parameters.extend(selected_ids)
        rows = connection.execute(
            """select r.result_id, r.report_start_utc, r.report_end_utc from strategies s
                 join strategy_results r on r.result_id = s.current_result_id and r.strategy_id = s.strategy_id
                """ + where,
            parameters,
        ).fetchall()
    if not rows:
        return
    jobs = [(str(database), int(result_id), report_start, report_end, config.ab_final_days) for result_id, report_start, report_end in rows]
    with ThreadPoolExecutor(max_workers=min(workers, len(jobs))) as executor:
        metrics = [metric for result in executor.map(_selection_window_job_from_args, jobs) for metric in result]
    with duckdb.connect(str(database)) as connection:
        for metric in metrics:
            _persist(connection, metric)


def selection_cache_status(connection: duckdb.DuckDBPyConnection, request: SelectionRequest, config: SelectionConfig) -> dict[str, int | bool]:
    rows = connection.execute(
        """select r.result_id, r.report_start_utc, r.report_end_utc from strategies s
             join strategy_results r on r.result_id = s.current_result_id and r.strategy_id = s.strategy_id
            where s.lifecycle_status = 'ACTIVE' and s.symbol = ? and s.side = ?""",
        [request.symbol, request.side],
    ).fetchall()
    cached_metrics = _selection_cached_metrics(connection, request)
    missing = 0
    for result_id, start, end in rows:
        windows = _selection_windows(start, end, config)
        if any(cached_metrics.get((int(result_id), window_start, window_end)) is None for window_start, window_end in windows):
            missing += 1
    return {"total": len(rows), "missing": missing, "ready": bool(rows) and missing == 0}


def load_selection_candidates(
    connection: duckdb.DuckDBPyConnection,
    request: SelectionRequest,
    config: SelectionConfig = SelectionConfig(),
    *, cache_only: bool = False,
) -> pd.DataFrame:
    """Load all current ACTIVE candidates for one Pair + Side without filtering them."""
    holding_minutes = _holding_quantiles_minutes(connection, request)
    b_holding_minutes = _window_b_holding_p95_minutes(connection, request, config)
    best_trade_facts = _best_trade_facts(connection, request)
    cached_metrics = _selection_cached_metrics(connection, request) if cache_only else None
    rows = connection.execute(
        """select s.strategy_id, s.strategy_name, s.symbol, s.side, s.timeframe, s.close_ma_len,
                  s.order_count, r.result_id, r.report_start_utc, r.report_end_utc,
                  r.reported_start_utc, r.reported_end_utc, r.effective_start_utc, r.effective_end_utc,
                  r.initial_balance,
                  r.total_pnl, r.total_pnl_pct, r.max_drawdown, r.max_drawdown_pct,
                  r.total_fees, r.total_trades, o.order_id, o.analysis_run_id, o.plateau_id,
                  o.open_ma_len, o.open_multiplier, o.shift_bp, o.lot_x, p.plateau_point_count
             from strategies s
             join strategy_results r on r.result_id = s.current_result_id and r.strategy_id = s.strategy_id
             left join strategy_orders o on o.strategy_id = s.strategy_id
             left join analysis_plateaus p on p.analysis_run_id = o.analysis_run_id and p.plateau_id = o.plateau_id
            where s.lifecycle_status = 'ACTIVE' and s.symbol = ? and s.side = ?
            order by s.strategy_name, s.strategy_id, o.order_id""",
        [request.symbol, request.side],
    ).fetchall()
    candidates: dict[int, dict[str, object]] = {}
    for row in rows:
        (
            strategy_id, strategy_name, symbol, side, timeframe, close_ma_len, order_count,
            result_id, report_start, report_end, reported_start, reported_end, effective_start, effective_end,
            initial_balance, total_pnl, total_pnl_pct, max_drawdown,
            max_drawdown_pct, total_fees, total_trades, order_id, analysis_run_id, plateau_id,
            open_ma_len, open_multiplier, shift_bp, lot_x, plateau_count,
        ) = row
        candidate = candidates.get(int(strategy_id))
        if candidate is None:
            result_id = int(result_id)
            if cache_only:
                full_metrics, ab_metrics = _cached_selection_metrics(
                    connection, result_id, report_start, report_end, config, cached_metrics
                )
                positive_quarters = _cached_positive_quarters(result_id, report_start, report_end, config, cached_metrics or {})
            else:
                full_metrics = get_or_calculate_window_pair(connection, result_id, (report_start, report_end), (report_start, report_end))[0]
                ab_metrics = _ab_metrics(connection, result_id, report_start, report_end, config)
                quarter_metrics = [
                    get_or_calculate_window(connection, result_id, start, end)
                    for start, end in _consistency_windows(report_start, report_end)
                ]
                positive_quarters = _consistency_summary(
                    quarter_metrics, _consistency_windows(report_start, report_end)
                )
            daily_log = None if full_metrics is None else full_metrics.daily_log_return
            pnl_30d = None if full_metrics is None else _return_30d(full_metrics, report_start, report_end)
            drawdown = _decimal_or_none(max_drawdown_pct)
            risk_scale = Decimal(5) / drawdown if drawdown is not None and drawdown > 0 else None
            best_share, pnl_without_best, positive_trade_count, best_trade_reliable = best_trade_facts.get(result_id, (None, None, None, False))
            initial_balance_value = _decimal_or_none(initial_balance)
            pnl_without_best_pct = (
                pnl_without_best / initial_balance_value * 100
                if pnl_without_best is not None and initial_balance_value is not None and initial_balance_value > 0 else None
            )
            candidate = {
                "strategy_id": int(strategy_id), "strategy_name": str(strategy_name), "symbol": str(symbol),
                "side": str(side), "timeframe": str(timeframe), "close_ma_len": int(close_ma_len),
                "order_count": int(order_count), "result_id": result_id, "total_pnl": _decimal_or_none(total_pnl),
                "total_pnl_pct": _decimal_or_none(total_pnl_pct), "max_drawdown": _decimal_or_none(max_drawdown),
                "max_drawdown_pct": drawdown, "total_fees": _decimal_or_none(total_fees),
                "report_start_utc": report_start, "report_end_utc": report_end,
                "reported_start_utc": reported_start, "reported_end_utc": reported_end,
                "effective_start_utc": effective_start, "effective_end_utc": effective_end,
                "total_trades": None if full_metrics is None else full_metrics.trade_count,
                "trades_30d": None if full_metrics is None else _trade_rate_30d(full_metrics, report_start, report_end),
                "pnl_30d_pct": pnl_30d,
                "profit_factor": None if full_metrics is None else full_metrics.profit_factor,
                "win_rate_pct": None if full_metrics is None else full_metrics.win_rate_pct,
                "risk_scale": risk_scale, "dd5_proxy": pnl_30d * risk_scale if pnl_30d is not None and risk_scale is not None else None,
                "holding_p95_minutes": holding_minutes.get(result_id, (None, None))[0],
                "holding_median_minutes": holding_minutes.get(result_id, (None, None))[1],
                **ab_metrics, "ab_holding_p95_minutes": b_holding_minutes.get(result_id),
                "first_shift_bp": None, "scaled_lot_sum": None, "capital_proxy": None,
                "capital_efficiency": None, "total_plateau_point_count": None,
                "best_trade_profit_share_pct": best_share, "pnl_without_best_trade": pnl_without_best,
                "pnl_without_best_trade_pct": pnl_without_best_pct,
                "completed_profitable_trade_count": positive_trade_count, "best_trade_reliable": best_trade_reliable,
                "positive_quarter_count": positive_quarters[0],
                "positive_quarter_available_count": positive_quarters[1],
                "positive_quarter_status": positive_quarters[2],
                "robust_pnl_30d_pct": None,
                "worst_drawdown_pct": None, "worst_holding_p95_minutes": None,
                "ab_stability_ratio": None, "minimum_plateau_point_count": None,
                "lot_variant_group_key": None, "lot_variant_representative_strategy_id": pd.NA,
            }
            candidates[int(strategy_id)] = candidate
        if order_id is not None:
            number = int(order_id)
            lot = _decimal_or_none(lot_x)
            points = None if plateau_count is None else int(plateau_count)
            candidate[f"order_{number}_open_ma_len"] = int(open_ma_len)
            candidate[f"order_{number}_open_multiplier"] = _decimal_or_none(open_multiplier)
            candidate[f"order_{number}_shift_bp"] = int(shift_bp)
            candidate[f"order_{number}_lot_x"] = lot
            candidate[f"order_{number}_plateau_point_count"] = points
            candidate[f"order_{number}_plateau_key"] = (
                (str(analysis_run_id), str(plateau_id))
                if analysis_run_id is not None and plateau_id is not None else None
            )
            if number == 1:
                candidate["first_shift_bp"] = int(shift_bp)
            lots = candidate.setdefault("_lots", [])
            points_list = candidate.setdefault("_points", [])
            if lot is not None:
                lots.append(lot)
            if points is not None:
                points_list.append(points)
    for candidate in candidates.values():
        lots = candidate.pop("_lots", [])
        points = candidate.pop("_points", [])
        risk_scale = candidate["risk_scale"]
        if risk_scale is not None and len(lots) == candidate["order_count"]:
            scaled_lot_sum = sum(lots, Decimal(0)) * risk_scale
            candidate["scaled_lot_sum"] = scaled_lot_sum
            candidate["capital_proxy"] = scaled_lot_sum + Decimal("0.05")
            if candidate["dd5_proxy"] is not None:
                candidate["capital_efficiency"] = candidate["dd5_proxy"] / candidate["capital_proxy"]
        if len(points) == candidate["order_count"]:
            candidate["total_plateau_point_count"] = sum(points)
            candidate["minimum_plateau_point_count"] = min(points)
        a, b = candidate["ab_return_a_30d_pct"], candidate["ab_return_b_30d_pct"]
        if a is not None and b is not None:
            candidate["robust_pnl_30d_pct"] = min(a, b)
            if a > 0 and b > 0:
                candidate["ab_stability_ratio"] = min(a, b) / max(a, b)
        full_dd, b_dd = candidate["max_drawdown_pct"], candidate["ab_drawdown_b_pct"]
        if full_dd is not None and b_dd is not None and full_dd >= 0 and b_dd >= 0:
            candidate["worst_drawdown_pct"] = max(full_dd, b_dd)
        full_hold, b_hold = candidate["holding_p95_minutes"], candidate["ab_holding_p95_minutes"]
        if full_hold is not None and b_hold is not None:
            candidate["worst_holding_p95_minutes"] = max(full_hold, b_hold)
    return pd.DataFrame.from_records(list(candidates.values())).reindex(columns=_CANDIDATE_COLUMNS)


_PARETO_OBJECTIVES = {
    "pareto_window_b": (("ab_return_b_30d_pct", "ab_trade_rate_b_30d"), ("ab_drawdown_b_pct", "ab_holding_p95_minutes")),
    "pareto_window_b_dd_shift": (("ab_return_b_30d_pct", "first_shift_bp"), ("max_drawdown_pct",)),
    "pareto_dd5_balanced": (("dd5_proxy", "first_shift_bp"), ("capital_proxy", "holding_p95_minutes", "close_ma_len")),
    "pareto_efficiency_shift": (("capital_efficiency", "first_shift_bp"), ()),
    "pareto_dd5_holding": (("dd5_proxy",), ("holding_p95_minutes",)),
    "pareto_dd5_close_ma": (("dd5_proxy",), ("close_ma_len",)),
    "pareto_dd5_first_shift": (("dd5_proxy", "first_shift_bp"), ()),
    "pareto_conditional_close_ma": (("capital_efficiency",), ("close_ma_len",)),
    "pareto_primary": (("dd5_proxy",), ("capital_proxy",)),
    "pareto_dd5_capital": (("dd5_proxy",), ("capital_proxy",)),
    "pareto_robust": (("robust_pnl_30d_pct", "first_shift_bp"), ("worst_drawdown_pct", "worst_holding_p95_minutes")),
}


def _present(value: object) -> bool:
    return value is not None and not pd.isna(value)


def _analog_group_keys(survivors: pd.DataFrame) -> pd.Series:
    """Partition exact plateau structures into non-transitive adjacent Close-MA groups."""
    groups: dict[tuple[object, ...], list[tuple[int, int, object]]] = {}
    keys: dict[object, tuple[object, ...]] = {}
    required = ("symbol", "side", "timeframe", "order_count", "close_ma_len")
    for index, row in survivors.iterrows():
        if not all(_present(row.get(column)) for column in required):
            keys[index] = ("__strategy__", int(row["strategy_id"]))
            continue
        order_count = int(row["order_count"])
        plateaus = tuple(row.get(f"order_{order}_plateau_key") for order in range(1, order_count + 1))
        if not plateaus or not all(_present(plateau) for plateau in plateaus):
            keys[index] = ("__strategy__", int(row["strategy_id"]))
            continue
        base = (row["symbol"], row["side"], row["timeframe"], order_count, *plateaus)
        groups.setdefault(base, []).append((int(row["close_ma_len"]), int(row["strategy_id"]), index))
    for base, members in groups.items():
        start: int | None = None
        bucket = 0
        for close_ma, _, index in sorted(members):
            if start is None or close_ma > start + 1:
                start = close_ma
                bucket += 1
            keys[index] = (*base, "close_ma", start, bucket)
    return pd.Series(keys)


def _dominates(other: pd.Series, candidate: pd.Series, maximize: tuple[str, ...], minimize: tuple[str, ...]) -> bool:
    columns = (*maximize, *minimize)
    if any(column not in other or not _present(other[column]) or not _present(candidate[column]) for column in columns):
        return False
    no_worse = all(other[column] >= candidate[column] for column in maximize) and all(
        other[column] <= candidate[column] for column in minimize
    )
    return no_worse and (any(other[column] > candidate[column] for column in maximize) or any(
        other[column] < candidate[column] for column in minimize
    ))


def _pareto_eliminated(group: pd.DataFrame, maximize: tuple[str, ...], minimize: tuple[str, ...]) -> list[object]:
    """Return dominated indexes without Python row-pair iteration."""
    columns = (*maximize, *minimize)
    values = group.loc[:, columns].to_numpy(dtype=object)
    valid = ~pd.isna(values).any(axis=1)
    comparable = values[valid]
    comparable_indexes = group.index[valid]
    eliminated: list[object] = []
    for candidate_index, candidate in enumerate(comparable):
        no_worse = np.ones(len(comparable), dtype=bool)
        strictly_better = np.zeros(len(comparable), dtype=bool)
        for column_index in range(len(maximize)):
            no_worse &= comparable[:, column_index] >= candidate[column_index]
            strictly_better |= comparable[:, column_index] > candidate[column_index]
        for column_index in range(len(maximize), len(columns)):
            no_worse &= comparable[:, column_index] <= candidate[column_index]
            strictly_better |= comparable[:, column_index] < candidate[column_index]
        no_worse[candidate_index] = False
        if np.any(no_worse & strictly_better):
            eliminated.append(comparable_indexes[candidate_index])
    return eliminated


def _plateau_pareto_eliminated(group: pd.DataFrame, stage_id: str, config: SelectionConfig) -> list[object]:
    eliminated: list[object] = []
    for _, same_order_count in group.groupby("order_count", sort=False):
        points = (
            tuple(f"order_{order}_plateau_point_count" for order in range(1, int(same_order_count["order_count"].iloc[0]) + 1))
            if stage_id.endswith("per_order") else ("total_plateau_point_count",)
        )
        columns = ("dd5_proxy", *points)
        values = same_order_count.loc[:, columns].to_numpy(dtype=object)
        valid = ~pd.isna(values).any(axis=1)
        comparable = values[valid]
        comparable_indexes = same_order_count.index[valid]
        for candidate_index, candidate in enumerate(comparable):
            dominates = comparable[:, 0] >= candidate[0] * config.plateau_points_pareto_pnl_multiplier
            for column_index in range(1, len(columns)):
                dominates &= comparable[:, column_index] >= candidate[column_index]
            dominates[candidate_index] = False
            if np.any(dominates):
                eliminated.append(comparable_indexes[candidate_index])
    return eliminated


def _near_tie_eliminated(
    group: pd.DataFrame, stage: SelectionStage, preference_column: str, *, higher_is_better: bool, advantage: Decimal = Decimal(0),
) -> list[object]:
    columns = ("robust_pnl_30d_pct", preference_column, "worst_drawdown_pct", "worst_holding_p95_minutes")
    values = group.loc[:, columns].to_numpy(dtype=object)
    valid = ~pd.isna(values).any(axis=1)
    comparable = values[valid]
    indexes = group.index[valid]
    tolerance = Decimal(1) - (stage.pnl_tolerance_pct or Decimal(10)) / Decimal(100)
    eliminated: list[object] = []
    for candidate_index, candidate in enumerate(comparable):
        pnl, preference, drawdown, holding = candidate
        if pnl <= 0:
            continue
        preferred = comparable[:, 1] >= preference + float(advantage) if higher_is_better else comparable[:, 1] < preference
        dominates = (
            (comparable[:, 0] > 0)
            & preferred
            & (comparable[:, 0] >= pnl * tolerance)
            & (comparable[:, 2] <= drawdown)
            & (comparable[:, 3] <= holding)
        )
        if np.any(dominates):
            eliminated.append(indexes[candidate_index])
    return eliminated


def _shift_near_tie_eliminated(group: pd.DataFrame, stage: SelectionStage, config: SelectionConfig) -> list[object]:
    return _near_tie_eliminated(
        group, stage, "first_shift_bp", higher_is_better=True, advantage=Decimal(config.shift_near_tie_min_advantage_bp),
    )


def _close_ma_near_tie_eliminated(group: pd.DataFrame, stage: SelectionStage) -> list[object]:
    return _near_tie_eliminated(group, stage, "close_ma_len", higher_is_better=False)


_RANK_COMPONENTS = (
    ("robust_pnl", "robust_pnl_30d_pct", Decimal(".30"), True, False),
    ("worst_drawdown", "worst_drawdown_pct", Decimal(".15"), False, False),
    ("ab_stability", "ab_stability_ratio", Decimal(".15"), True, False),
    ("worst_holding", "worst_holding_p95_minutes", Decimal(".12"), False, False),
    ("first_shift", "first_shift_bp", Decimal(".10"), True, False),
    ("minimum_plateau_points", "minimum_plateau_point_count", Decimal(".09"), True, True),
    ("close_ma", "close_ma_len", Decimal(".09"), False, False),
)


def _quality_percentiles(values: pd.Series, *, higher_is_better: bool, by_timeframe: pd.Series | None = None) -> pd.Series:
    quality = pd.Series(np.nan, index=values.index, dtype=float)
    groups = [(None, values)] if by_timeframe is None else values.groupby(by_timeframe, sort=False)
    for _, group in groups:
        numeric = pd.to_numeric(group, errors="coerce")
        numeric = numeric[numeric.notna()]
        if numeric.empty:
            continue
        if len(numeric) == 1:
            quality.loc[numeric.index] = 1.0
            continue
        ranks = numeric.rank(method="average", ascending=higher_is_better)
        quality.loc[numeric.index] = (ranks - 1) / (len(numeric) - 1)
    return quality


def _rank_robust(group: pd.DataFrame) -> tuple[pd.DataFrame, list[object]]:
    ranked = group.copy()
    quality_columns: list[str] = []
    for name, source, _, higher, within_timeframe in _RANK_COMPONENTS:
        if source not in ranked:
            ranked[source] = np.nan
        column = f"rank_quality_{name}"
        quality_columns.append(column)
        ranked[column] = _quality_percentiles(
            ranked[source], higher_is_better=higher,
            by_timeframe=ranked["timeframe"] if within_timeframe else None,
        )
    if "dd5_proxy" not in ranked:
        ranked["dd5_proxy"] = np.nan
    weights = np.array([float(weight) for _, _, weight, _, _ in _RANK_COMPONENTS])
    qualities = ranked.loc[:, quality_columns].to_numpy(dtype=float)
    present = ~np.isnan(qualities)
    coverage = present @ weights
    weighted = np.nan_to_num(qualities, nan=0.0) @ weights
    ranked["rank_weight_coverage_pct"] = coverage * 100
    for column_index, (name, _, _, _, _) in enumerate(_RANK_COMPONENTS):
        rank_weight = np.full(len(coverage), np.nan)
        np.divide(weights[column_index], coverage, out=rank_weight, where=coverage > 0)
        ranked[f"rank_weight_{name}"] = np.where(present[:, column_index], rank_weight, 0.0)
    final_score = np.full(len(coverage), np.nan)
    np.divide(weighted * 100, coverage, out=final_score, where=coverage > 0)
    ranked["final_score"] = final_score
    rankable = ranked["final_score"].notna()
    ordered = ranked.loc[rankable].sort_values(
        ["final_score", "dd5_proxy", "robust_pnl_30d_pct", "worst_drawdown_pct", "ab_stability_ratio",
         "worst_holding_p95_minutes", "first_shift_bp", "minimum_plateau_point_count", "close_ma_len", "strategy_id"],
        ascending=[False, False, False, True, False, True, False, False, True, True],
        na_position="last", kind="stable",
    )
    ranked.loc[ordered.index, "final_rank"] = range(1, len(ordered) + 1)
    return ranked, ordered.index.tolist()


def _scope_groups(frame: pd.DataFrame, scope: StageScope):
    return frame.groupby([] if scope == "pair_side" else ["timeframe"], dropna=False, sort=False) if scope != "pair_side" else [(None, frame)]


def _ab_eliminates(row: pd.Series, config: SelectionConfig) -> bool | None:
    fields = ("ab_return_a_30d_pct", "ab_return_b_30d_pct", "ab_win_rate_b_pct", "ab_trade_rate_a_30d", "ab_trade_rate_b_30d")
    if any(field not in row or not _present(row[field]) for field in fields):
        return None
    a, b, win_b, trades_a, trades_b = (row[field] for field in fields)
    return (
        b <= config.ab_return_floor_pct
        or (a > 0 and b <= a / config.ab_return_divisor)
        or win_b < config.ab_win_rate_floor_pct
        or (trades_a > 0 and trades_b <= trades_a / config.ab_trade_rate_divisor)
    )


_LOT_VARIANT_STAGE_ID = "filter_lot_variant_redundancy"
_LOT_VARIANT_METRICS = ("dd5_proxy", "capital_proxy", "robust_pnl_30d_pct", "worst_drawdown_pct", "profit_factor")


def _utc_datetime(value: object) -> datetime | None:
    if isinstance(value, pd.Timestamp):
        value = value.to_pydatetime()
    if isinstance(value, str):
        try:
            value = datetime.fromisoformat(value.replace("Z", "+00:00"))
        except ValueError:
            return None
    if not isinstance(value, datetime):
        return None
    if value.tzinfo is None:
        return None
    return value.astimezone(timezone.utc)


def _comparison_interval(row: pd.Series) -> tuple[datetime, datetime] | None:
    for start_name, end_name in (
        ("comparison_interval_start_utc", "comparison_interval_end_utc"),
        ("effective_start_utc", "effective_end_utc"),
        ("reported_start_utc", "reported_end_utc"),
        ("report_start_utc", "report_end_utc"),
    ):
        if start_name not in row or end_name not in row:
            continue
        raw_start, raw_end = row[start_name], row[end_name]
        if not _present(raw_start) and not _present(raw_end):
            continue
        start, end = _utc_datetime(raw_start), _utc_datetime(raw_end)
        if start is not None and end is not None and end >= start:
            return start, end
        return None
    return None


def _integer(value: object) -> int | None:
    parsed = _decimal_or_none(value)
    if parsed is None or parsed != parsed.to_integral_value():
        return None
    return int(parsed)


def _lot_variant_structure(row: pd.Series) -> tuple[tuple[object, ...], tuple[Decimal, ...]] | None:
    close_ma = _integer(row.get("close_ma_len"))
    order_count = _integer(row.get("order_count"))
    if close_ma is None or order_count is None or order_count < 1:
        return None
    fields: list[tuple[int, int, Decimal]] = []
    for order in range(1, order_count + 1):
        open_ma = _integer(row.get(f"order_{order}_open_ma_len"))
        shift = _integer(row.get(f"order_{order}_shift_bp"))
        lot = _decimal_or_none(row.get(f"order_{order}_lot_x"))
        if open_ma is None or shift is None or lot is None:
            return None
        fields.append((open_ma, shift, lot))
    fields.sort(key=lambda value: (value[0], value[1], value[2]))
    symbol, side, timeframe = (row.get(name) for name in ("symbol", "side", "timeframe"))
    if any(value is None or pd.isna(value) for value in (symbol, side, timeframe)):
        return None
    pairs = tuple((open_ma, shift) for open_ma, shift, _ in fields)
    lots = tuple(lot for _, _, lot in fields)
    return (str(symbol), str(side), str(timeframe), close_ma, order_count, pairs), lots


def _lot_variant_group_key(structure: tuple[object, ...]) -> str:
    symbol, side, timeframe, close_ma, order_count, pairs = structure
    return json.dumps(
        [symbol, side, timeframe, close_ma, order_count, [[open_ma, shift] for open_ma, shift in pairs]],
        ensure_ascii=False,
        separators=(",", ":"),
    )


def _lot_variant_eliminated(result: pd.DataFrame, group: pd.DataFrame) -> list[object]:
    candidates: dict[tuple[object, ...], list[tuple[object, tuple[Decimal, ...]]]] = {}
    for index, row in group.iterrows():
        structure = _lot_variant_structure(row)
        interval = _comparison_interval(row)
        if structure is None or interval is None:
            continue
        group_id = (*structure[0], interval[0], interval[1])
        candidates.setdefault(group_id, []).append((index, structure[1]))

    eliminated: list[object] = []
    for group_id, members in candidates.items():
        if len(members) < 2 or len({lots for _, lots in members}) < 2:
            continue
        rows = result.loc[[index for index, _ in members]]
        if any(
            _decimal_or_none(rows.iloc[position].get(metric)) is None
            for position in range(len(rows))
            for metric in _LOT_VARIANT_METRICS
        ):
            continue
        strategy_ids: dict[object, int] = {}
        for index in rows.index:
            raw_id = rows.at[index, "strategy_id"]
            if isinstance(raw_id, bool) or not isinstance(raw_id, (int, np.integer)) or int(raw_id) < 1:
                break
            strategy_ids[index] = int(raw_id)
        else:
            def winner_key(index: object) -> tuple[object, ...]:
                row = rows.loc[index]
                values = [_decimal_or_none(row[metric]) for metric in _LOT_VARIANT_METRICS]
                assert all(value is not None for value in values)
                dd5, capital, robust, drawdown, profit_factor = values
                return (-dd5, capital, -robust, drawdown, -profit_factor, strategy_ids[index], row.get("_source_order", index))

            winner = min(rows.index, key=winner_key)
            structure = group_id[:-2]
            key = _lot_variant_group_key(structure)
            result.loc[rows.index, "lot_variant_group_key"] = key
            result.loc[rows.index, "lot_variant_representative_strategy_id"] = strategy_ids[winner]
            losers = [index for index in rows.index if index != winner]
            if losers:
                eliminated.extend(losers)
                result.loc[losers, "auto_status"] = "FILTERED"
                result.loc[losers, "elimination_reason"] = "LOT_VARIANT_REDUNDANT"
    return eliminated


def run_selection(
    candidates: pd.DataFrame, request: SelectionRequest, config: SelectionConfig = SelectionConfig()
) -> pd.DataFrame:
    """Apply the submitted stages in order; input candidates remain fully represented."""
    result = candidates.copy()
    result["_source_order"] = np.arange(len(result))
    result = result.sort_values(["strategy_name", "strategy_id"], kind="stable").reset_index(drop=True)
    if "symbol" not in result:
        result["symbol"] = request.symbol
    if "side" not in result:
        result["side"] = request.side
    if "prior_rejected" not in result:
        result["prior_rejected"] = False
    result["prior_rejected"] = result["prior_rejected"].fillna(False).astype(bool)
    result["finalist"] = True
    result["elimination_reason"] = None
    result["auto_status"] = None
    result["analog_group_key"] = None
    result["auto_analog_of_strategy_id"] = pd.NA
    result["lot_variant_group_key"] = None
    result["lot_variant_representative_strategy_id"] = pd.NA
    stage_counts: dict[str, dict[str, int | bool]] = {}
    explicit_stage_ids = {stage.id for stage in request.stages}
    stages = list(request.stages)
    if _LOT_VARIANT_STAGE_ID not in explicit_stage_ids and config.lot_variant_redundancy_enabled:
        stages.insert(0, SelectionStage(_LOT_VARIANT_STAGE_ID, True, "pair_side_timeframe"))
    elif _LOT_VARIANT_STAGE_ID in explicit_stage_ids:
        lot_stage = next(stage for stage in stages if stage.id == _LOT_VARIANT_STAGE_ID)
        stages = [lot_stage, *(stage for stage in stages if stage.id != _LOT_VARIANT_STAGE_ID)]
    implicit_lot_variant_stage = _LOT_VARIANT_STAGE_ID not in explicit_stage_ids
    for stage in stages:
        column = f"eliminated_by_{stage.id}"
        result[column] = False
        if not stage.enabled or (stage.id == _LOT_VARIANT_STAGE_ID and not config.lot_variant_redundancy_enabled):
            if stage.id != _LOT_VARIANT_STAGE_ID or not implicit_lot_variant_stage:
                stage_counts[stage.id] = {"enabled": False, "eliminated": 0, "remaining": int(result["finalist"].sum())}
            continue
        if stage.id == _LOT_VARIANT_STAGE_ID:
            survivors = result.loc[result["finalist"]]
            for _, group in _scope_groups(survivors, stage.scope):
                eliminated = _lot_variant_eliminated(result, group)
                if eliminated:
                    result.loc[eliminated, column] = True
                    result.loc[eliminated, "finalist"] = False
            if stage.id != _LOT_VARIANT_STAGE_ID or not implicit_lot_variant_stage:
                stage_counts[stage.id] = {"enabled": True, "eliminated": int(result[column].sum()), "remaining": int(result["finalist"].sum())}
            continue
        if stage.id == "rank_robust_top_n":
            survivors = result.loc[result["finalist"]]
            ranked, ordered = _rank_robust(survivors)
            rank_columns = [column for column in ranked.columns if column.startswith("rank_") or column in {"final_score", "final_rank"}]
            result.loc[ranked.index, rank_columns] = ranked.loc[:, rank_columns]
            order_position = {index: position for position, index in enumerate(ordered)}
            representatives: list[object] = []
            analog_keys = _analog_group_keys(result.loc[survivors.index])
            for key, group in result.loc[survivors.index].groupby(analog_keys, sort=False):
                group_key = json.dumps(key, ensure_ascii=False, separators=(",", ":"))
                result.loc[group.index, "analog_group_key"] = group_key
                not_rejected = group.index[~result.loc[group.index, "prior_rejected"]]
                pool = not_rejected if len(not_rejected) else group.index
                rankable_pool = [index for index in pool if index in order_position]
                representative = (
                    min(rankable_pool, key=order_position.get)
                    if rankable_pool
                    else result.loc[pool, "strategy_id"].astype(int).idxmin()
                )
                representatives.append(representative)
                analogs = group.index.difference([representative], sort=False)
                if len(analogs):
                    result.loc[analogs, "auto_status"] = "ANALOG"
                    result.loc[analogs, "auto_analog_of_strategy_id"] = int(result.at[representative, "strategy_id"])
                    result.loc[analogs, "finalist"] = False
                    result.loc[analogs, column] = True
                    result.loc[analogs, "elimination_reason"] = "ANALOG"

            ranked_representatives = sorted(
                (index for index in representatives if index in order_position), key=order_position.get
            )
            unranked_representatives = sorted(
                (index for index in representatives if index not in order_position),
                key=lambda index: int(result.at[index, "strategy_id"]),
            )
            result.loc[survivors.index, "final_rank"] = np.nan
            result.loc[ranked_representatives, "final_rank"] = range(1, len(ranked_representatives) + 1)
            selected = 0
            for index in [*ranked_representatives, *unranked_representatives]:
                result.at[index, "finalist"] = False
                if result.at[index, "prior_rejected"]:
                    result.at[index, "auto_status"] = "RESERVE"
                    result.at[index, "elimination_reason"] = "PRIOR_USER_REJECTED"
                elif index not in order_position:
                    result.at[index, "auto_status"] = "RESERVE"
                    result.at[index, "elimination_reason"] = "RANK_NOT_EVALUATED_INSUFFICIENT_DATA"
                elif selected < stage.top_n:
                    result.at[index, "auto_status"] = "FINALIST"
                    result.at[index, "finalist"] = True
                    result.at[index, "elimination_reason"] = None
                    selected += 1
                else:
                    result.at[index, "auto_status"] = "RESERVE"
                    result.at[index, "elimination_reason"] = stage.id.upper()
                result.at[index, column] = not result.at[index, "finalist"]
            eliminated_count = int(result.loc[survivors.index, column].sum())
            stage_counts[stage.id] = {"enabled": stage.enabled, "eliminated": eliminated_count, "remaining": int(result["finalist"].sum())}
            continue
        survivors = result.loc[result["finalist"]]
        for _, group in _scope_groups(survivors, stage.scope):
            if stage.id in {"filter_holding_outlier", "filter_low_trades"}:
                metric = "holding_p95_minutes" if stage.id == "filter_holding_outlier" else "trades_30d"
                if metric not in group:
                    continue
                values = pd.to_numeric(group[metric], errors="coerce").dropna()
                if values.empty:
                    continue
                q1, q3 = values.quantile(.25), values.quantile(.75)
                threshold = q3 + 1.5 * (q3 - q1) if stage.id == "filter_holding_outlier" else q1 - 1.5 * (q3 - q1)
                failed = group[metric] > threshold if stage.id == "filter_holding_outlier" else group[metric] < threshold
                eliminated = group.index[failed.fillna(False)]
            elif stage.id == "filter_min_shift":
                threshold_bp = stage.min_shift_pct * 100
                shift_columns = [f"order_{order}_shift_bp" for order in range(1, 5) if f"order_{order}_shift_bp" in group]
                failed = group[shift_columns].apply(
                    lambda shifts: any(_present(value) and Decimal(str(value)) < threshold_bp for value in shifts), axis=1
                ) if shift_columns else pd.Series(False, index=group.index)
                eliminated = group.index[failed]
            elif stage.id == "filter_best_trade_dependency":
                evaluable = group["best_trade_reliable"].fillna(False) & (
                    pd.to_numeric(group["completed_profitable_trade_count"], errors="coerce") >= config.best_trade_min_profitable_trades
                )
                failed = evaluable & (
                    (pd.to_numeric(group["pnl_without_best_trade"], errors="coerce") <= 0)
                    | (pd.to_numeric(group["best_trade_profit_share_pct"], errors="coerce") > float(config.best_trade_max_profit_share_pct))
                )
                eliminated = group.index[failed.fillna(False)]
            elif stage.id == "filter_time_consistency":
                status = group.get("positive_quarter_status")
                if status is not None:
                    failed = status.eq("FAIL")
                else:
                    if not {"positive_quarter_available_count", "positive_quarter_count"}.issubset(group.columns):
                        continue
                    available = pd.to_numeric(group["positive_quarter_available_count"], errors="coerce")
                    positive = pd.to_numeric(group["positive_quarter_count"], errors="coerce")
                    failed = available.eq(4) & positive.lt(3)
                eliminated = group.index[failed.fillna(False)]
            elif stage.id == "ab_deterioration":
                eliminated = []
                for index, row in group.iterrows():
                    decision = _ab_eliminates(row, config)
                    if decision is None:
                        result.at[index, "elimination_reason"] = "AB_NOT_EVALUATED_INSUFFICIENT_DATA"
                    elif decision:
                        eliminated.append(index)
            else:
                if stage.id == "pareto_conditional_close_ma" and len(group) <= 3:
                    stage_counts[stage.id] = {"enabled": True, "eliminated": 0, "remaining": int(result["finalist"].sum())}
                    continue
                eliminated = (
                    _plateau_pareto_eliminated(group, stage.id, config)
                    if stage.id.startswith("pareto_plateau_points")
                    else _shift_near_tie_eliminated(group, stage, config)
                    if stage.id == "pareto_shift_near_tie"
                    else _close_ma_near_tie_eliminated(group, stage)
                    if stage.id == "pareto_close_ma_near_tie"
                    else _pareto_eliminated(group, *_PARETO_OBJECTIVES[stage.id])
                )
            if len(eliminated):
                result.loc[eliminated, column] = True
                result.loc[eliminated, "finalist"] = False
                result.loc[eliminated, "elimination_reason"] = stage.id.upper()
        if stage.id != _LOT_VARIANT_STAGE_ID or not implicit_lot_variant_stage:
            stage_counts[stage.id] = {"enabled": True, "eliminated": int(result[column].sum()), "remaining": int(result["finalist"].sum())}
    result.loc[result["auto_status"].isna() & result["finalist"], "auto_status"] = "FINALIST"
    result.loc[result["auto_status"].isna() & ~result["finalist"], "auto_status"] = "FILTERED"
    result = result.drop(columns=["_source_order"])
    result.attrs["stage_counts"] = stage_counts
    return result


def write_selection_workbook(
    result: pd.DataFrame,
    path: Path,
    request: SelectionRequest,
    review_metadata: Mapping[str, str] | None = None,
) -> Path:
    """Write the one disposable selection workbook; internal A/B facts stay internal."""
    display = result.drop(columns=[
        column for column in result.columns
        if column.startswith("ab_") and column not in {
            "ab_pnl_change_30d_pct", "ab_return_a_30d_pct", "ab_calendar_days_a", "ab_return_b_30d_pct", "ab_calendar_days_b", "ab_stability_ratio",
        }
    ] + ["total_pnl", "total_pnl_pct", "max_drawdown", "total_fees", "risk_scale", "scaled_lot_sum", "daily_log_return"], errors="ignore").copy()
    if "ab_pnl_change_30d_pct" not in display:
        display["ab_pnl_change_30d_pct"] = None
    reason_aliases = {"PARETO_PLATEAU_POINTS_PER_ORDER": "PARETO_PL_PTS_PER_ORDER"}
    enabled_stages = [stage for stage in request.stages if stage.enabled]
    reason_positions = {stage.id.upper(): index for index, stage in enumerate(enabled_stages, start=1)}
    reason_colors = {
        stage.id.upper(): "".join(
            f"{round(value * 255):02X}"
            for value in hls_to_rgb(((220 + 140 * index / max(1, len(enabled_stages) - 1)) % 360) / 360, .96 if index == 1 else .92, .55)
        )
        for index, stage in enumerate(enabled_stages)
    }
    row_fills = [
        "D9EAD3" if finalist else reason_colors.get(str(reason))
        for finalist, reason in zip(display.get("finalist", []), display.get("elimination_reason", []))
    ]
    if "elimination_reason" in display:
        display["elimination_reason"] = display["elimination_reason"].map(
            lambda reason: (
                f"{reason_positions[reason]}. {reason_aliases.get(reason, reason)}"
                if reason in reason_positions else reason_aliases.get(reason, reason)
            ) if isinstance(reason, str) else reason
        )
    for column in ("pnl_30d_pct", "profit_factor", "win_rate_pct"):
        if column not in display:
            display[column] = None
    if "positive_quarter_status" in display:
        if "positive_quarter_count" not in display:
            display["positive_quarter_count"] = None
        if "positive_quarter_available_count" not in display:
            display["positive_quarter_available_count"] = None
    if {"positive_quarter_count", "positive_quarter_available_count"}.issubset(display.columns):
        display["positive_quarter_count"] = display.apply(
            lambda row: (
                "N/A" if row.get("positive_quarter_status") == "UNAVAILABLE" else
                f"{int(row['positive_quarter_count'])}/{int(row['positive_quarter_available_count'])}"
                if _present(row["positive_quarter_count"]) and _present(row["positive_quarter_available_count"])
                else None
            ),
            axis=1,
        )
    for column in display.columns:
        if column.endswith("_id") or "count" in column or column.endswith("_bp"):
            continue
        display[column] = display[column].map(
            lambda value: value.quantize(Decimal(".01")) if isinstance(value, Decimal) else value
        )
    for column in ("first_shift_bp", *(f"order_{order}_shift_bp" for order in range(1, 5))):
        if column in display:
            display[column] = display[column].map(
                lambda value: (Decimal(str(value)) / Decimal("100")).quantize(
                    Decimal(".1"), rounding=ROUND_HALF_UP
                ) if value is not None and not pd.isna(value) else value
            )
    for column in (
        "pnl_30d_pct", "dd5_proxy", "profit_factor", "ab_pnl_change_30d_pct",
        "ab_return_a_30d_pct", "ab_return_b_30d_pct", "pnl_without_best_trade_pct",
        "capital_efficiency", "win_rate_pct", "holding_p95_minutes",
        "holding_median_minutes", "total_plateau_point_count", "minimum_plateau_point_count", "final_rank",
        *(f"order_{order}_plateau_point_count" for order in range(1, 5)),
        *(f"order_{order}_open_ma_len" for order in range(1, 5)),
    ):
        if column in display:
            display[column] = display[column].map(
                lambda value: int(Decimal(str(value)).quantize(Decimal("1"), rounding=ROUND_HALF_UP))
                if value is not None and not pd.isna(value) else value
            )
    enabled_filter_columns = [f"eliminated_by_{stage.id}" for stage in request.stages if stage.enabled]
    implicit_lot_column = "eliminated_by_filter_lot_variant_redundancy"
    if implicit_lot_column in display and implicit_lot_column not in enabled_filter_columns and not any(
        stage.id == _LOT_VARIANT_STAGE_ID for stage in request.stages
    ):
        enabled_filter_columns.insert(0, implicit_lot_column)
    display = display.drop(columns=[
        column for column in display if column.startswith("eliminated_by_") and column not in enabled_filter_columns
    ], errors="ignore")
    for column in enabled_filter_columns:
        if column in display:
            if column == "eliminated_by_filter_time_consistency" and "positive_quarter_status" in display:
                display[column] = display.apply(
                    lambda row: "N/A" if row.get("positive_quarter_status") == "UNAVAILABLE"
                    else ("BLOCK" if bool(row[column]) else "PASS") if pd.notna(row[column]) else row[column],
                    axis=1,
                )
            else:
                display[column] = display[column].map(
                    lambda value: ("BLOCK" if bool(value) else "PASS") if pd.notna(value) else value
                )
    def order_values(columns: list[str], render: Callable[[object], str]) -> pd.Series:
        values = display.reindex(columns=columns)
        return values.apply(
            lambda row: " / ".join(
                "-" if value is None or pd.isna(value) else render(value)
                for value in row.iloc[:max(index for index, value in enumerate(row) if value is not None and not pd.isna(value)) + 1]
            ) if any(value is not None and not pd.isna(value) for value in row) else None,
            axis=1,
        )
    ma_columns = [f"order_{order}_open_ma_len" for order in range(1, 5)]
    display["open_ma"] = order_values(
        ma_columns, lambda value: str(int(Decimal(str(value)).quantize(Decimal("1"), rounding=ROUND_HALF_UP))),
    )
    lot_columns = [f"order_{order}_lot_x" for order in range(1, 5)]
    display["lots"] = order_values(
        lot_columns, lambda value: str(int((Decimal(str(value)) * 100).quantize(Decimal("1"), rounding=ROUND_HALF_UP))),
    )
    point_columns = [f"order_{order}_plateau_point_count" for order in range(1, 5)]
    display["points"] = order_values(
        point_columns, lambda value: str(int(Decimal(str(value)).quantize(Decimal("1"), rounding=ROUND_HALF_UP))),
    )
    if review_metadata is not None:
        display["user_status"] = display.apply(
            lambda row: "REJECTED" if bool(row.get("prior_rejected", False)) else row.get("auto_status"), axis=1
        )
        display["retest"] = display.apply(
            lambda row: "RETEST" if bool(row.get("prior_retest", False)) else None, axis=1
        )
        display["auto_rank"] = display.get("final_rank")
        display["user_rank"] = display.apply(
            lambda row: row.get("final_rank") if row.get("user_status") in {"FINALIST", "RESERVE"} else None,
            axis=1,
        )
        display["user_analog_of_strategy_id"] = display.apply(
            lambda row: row.get("auto_analog_of_strategy_id") if row.get("user_status") == "ANALOG" else None,
            axis=1,
        )
        display["comment"] = None
    review_identity_columns = ["result_id"] if review_metadata is not None else []
    review_columns = [
        "auto_status", "user_status", "retest", "auto_rank", "user_rank", "auto_analog_of_strategy_id",
        "user_analog_of_strategy_id", "comment",
    ] if review_metadata is not None else []
    column_order = [
        "strategy_id", *review_identity_columns, "strategy_name", "symbol", "side", "timeframe", "order_count", "close_ma_len",
        "pnl_30d_pct", "dd5_proxy", "ab_pnl_change_30d_pct", "ab_return_a_30d_pct", "ab_calendar_days_a", "ab_return_b_30d_pct", "ab_calendar_days_b", "positive_quarter_count",
        "capital_efficiency", "profit_factor", "max_drawdown_pct", "win_rate_pct", "total_trades", "trades_30d", "capital_proxy",
        "holding_p95_minutes", "holding_median_minutes", "total_plateau_point_count", "minimum_plateau_point_count",
        "best_trade_profit_share_pct", "pnl_without_best_trade_pct", "completed_profitable_trade_count",
        "robust_pnl_30d_pct", "worst_drawdown_pct", "worst_holding_p95_minutes", "ab_stability_ratio",
        "rank_quality_robust_pnl", "rank_quality_worst_drawdown", "rank_quality_ab_stability", "rank_quality_first_shift", "rank_quality_minimum_plateau_points", "rank_quality_close_ma",
        "rank_weight_coverage_pct", "rank_weight_robust_pnl", "rank_weight_worst_drawdown", "rank_weight_ab_stability", "rank_weight_first_shift", "rank_weight_minimum_plateau_points", "rank_weight_close_ma",
        "final_score",
        *(f"order_{order}_shift_bp" for order in range(1, 5)),
        "lots",
        "points",
        "open_ma",
        "final_rank", "finalist", "elimination_reason", "lot_variant_group_key", "lot_variant_representative_strategy_id",
        *enabled_filter_columns, *review_columns,
    ]
    display = display.reindex(columns=column_order)
    display = display.rename(columns={
        "strategy_id": "ID", "result_id": "Result ID", "strategy_name": "Стратегия", "symbol": "Пара", "side": "Side", "timeframe": "ТФ",
        "close_ma_len": "Close", "order_count": "ORD",
        "pnl_30d_pct": "PnL/30", "dd5_proxy": "PnL DD5/30", "profit_factor": "PF",
        "ab_pnl_change_30d_pct": "∆ PnL A/B", "ab_return_a_30d_pct": "PnL A/30д, %", "ab_calendar_days_a": "Дней A", "ab_return_b_30d_pct": "PnL B/30д, %", "ab_calendar_days_b": "Дней B", "capital_efficiency": "CE",
        "max_drawdown_pct": "DD", "win_rate_pct": "W/R", "total_trades": "Trades", "capital_proxy": "Lot DD5",
        "holding_p95_minutes": "Hold p95", "holding_median_minutes": "Hold M",
        "total_plateau_point_count": "PointsALL", "finalist": "Final", "elimination_reason": "Причина",
        "best_trade_profit_share_pct": "Best trade, %", "pnl_without_best_trade_pct": "PnL without best, %",
        "completed_profitable_trade_count": "Positive trades", "positive_quarter_count": "Positive windows", "trades_30d": "Trades/30",
        "robust_pnl_30d_pct": "Robust PnL/30", "worst_drawdown_pct": "Worst DD", "worst_holding_p95_minutes": "Worst Hold p95",
        "ab_stability_ratio": "A/B stability", "minimum_plateau_point_count": "PointsMin",
        "rank_quality_robust_pnl": "Rank q PnL", "rank_quality_worst_drawdown": "Rank q DD",
        "rank_quality_ab_stability": "Rank q A/B", "rank_quality_first_shift": "Rank q Shift", "rank_quality_minimum_plateau_points": "Rank q Points",
        "rank_quality_close_ma": "Rank q Close MA",
        "rank_weight_coverage_pct": "Rank coverage, %", "rank_weight_robust_pnl": "Rank w PnL",
        "rank_weight_worst_drawdown": "Rank w DD", "rank_weight_ab_stability": "Rank w A/B",
        "rank_weight_first_shift": "Rank w Shift", "rank_weight_minimum_plateau_points": "Rank w Points",
        "rank_weight_close_ma": "Rank w Close MA",
        "final_score": "Final score (Pair+Side)", "final_rank": "Final rank",
        "lot_variant_group_key": "Lot variant group key", "lot_variant_representative_strategy_id": "Lot variant representative ID",
        "open_ma": "MA",
        **{f"order_{order}_shift_bp": f"{order} Shift" for order in range(1, 5)},
        "lots": "Lots",
        "points": "Points",
        "auto_status": "Auto Status", "user_status": "User Status", "retest": "RETEST", "auto_rank": "Auto Rank",
        "user_rank": "User Rank", "auto_analog_of_strategy_id": "Auto Analog Of ID",
        "user_analog_of_strategy_id": "Analog Of ID", "comment": "Comment",
    })
    finalists = display.loc[display["Final"]].copy()
    finalist_fills = [color for color, finalist in zip(row_fills, display["Final"]) if finalist]
    workbook_path = write_audit_workbook(
        {"All candidates": display, "Finalists": finalists}, Path(path), data_widths_only=True,
        minimum_width=3, hidden_columns=frozenset({
            "Result ID", "Стратегия", "Auto Analog Of ID", "Positive trades", "Rank coverage, %", "Rank w PnL", "Rank w DD",
            "Rank w A/B", "Rank w Shift", "Rank w Points", "Rank w Close MA",
            "Robust PnL/30", "Worst Hold p95", "Rank q PnL", "Rank q DD", "Rank q A/B",
            "Rank q Shift", "Rank q Points", "Rank q Close MA", "Final score (Pair+Side)",
            "Worst DD", "A/B stability", "Best trade, %", "PnL without best, %",
        }), numeric_decimals=True, left_aligned_columns=frozenset({"Причина"}), row_fill_colors={
            "All candidates": row_fills,
            "Finalists": finalist_fills,
        },
        number_formats={
            **{f"{order} Shift": "0.0" for order in range(1, 5)},
            **{header: "0" for header in (
                "PnL/30", "PnL DD5/30", "PF", "PnL A/30д, %", "Дней A", "PnL B/30д, %", "Дней B", "PnL without best, %",
            )},
            "Final rank": "0",
        },
        center_from_column=5,
        font_colors={"Дней A": "FF0000FF", "Дней B": "FF0000FF"},
        bold_columns=frozenset({"Close", "DD", "Hold p95", "1 Shift", "Final rank"}),
        column_edge_borders={
                "Positive windows": ("right",),
            "Hold p95": ("left",),
            "Hold M": ("right",),
            "1 Shift": ("left",),
            "4 Shift": ("right",),
                "Points": ("left", "right"),
            "MA": ("left", "right"),
            "Final rank": ("left", "right"),
            "Close": ("left", "right"),
        },
    )
    if review_metadata is None:
        return workbook_path
    workbook = load_workbook(workbook_path)
    metadata_sheet = workbook.create_sheet("_MRS_SELECTION_META")
    for row in review_metadata.items():
        metadata_sheet.append(row)
    metadata_sheet.sheet_state = "veryHidden"
    for sheet_name in ("All candidates", "Finalists"):
        worksheet = workbook[sheet_name]
        headers = {cell.value: cell.column_letter for cell in worksheet[1]}
        if "User Status" in headers:
            validation = DataValidation(
                type="list", formula1='"FINALIST,RESERVE,ANALOG,FILTERED,REJECTED"', allow_blank=False
            )
            worksheet.add_data_validation(validation)
            validation.add(f'{headers["User Status"]}2:{headers["User Status"]}{max(2, worksheet.max_row)}')
        if "RETEST" in headers:
            validation = DataValidation(type="list", formula1='"RETEST"', allow_blank=True)
            worksheet.add_data_validation(validation)
            validation.add(f'{headers["RETEST"]}2:{headers["RETEST"]}{max(2, worksheet.max_row)}')
    workbook.save(workbook_path)
    _normalize_xlsx_archive(workbook_path)
    return workbook_path
