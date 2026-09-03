from __future__ import annotations

from copy import deepcopy
from dataclasses import dataclass
from decimal import Decimal
from hashlib import sha256
import json
import os
from pathlib import Path
import shutil
import tempfile
import threading
from typing import BinaryIO, Mapping
from uuid import uuid4

import duckdb
import pandas as pd
from openpyxl import load_workbook

from .config import AlgorithmConfig
from .lots import LotMethod
from .panel_strategy_batch import validate_strategy_manifest
from .performance_v2_store import PerformanceV2StoreError, require_performance_v2
from .pipeline import _publish_strategies, _write_json_atomic
from .strategy_json import generate_strategy
from .performance_v2_input import adapt_strategy_identity


_RETEST_PUBLICATION_LOCK = threading.Lock()


class _RetestPublicationLock:
    """Cross-process advisory lock for the atomic RETEST output swap."""

    def __init__(self, output_dir: Path) -> None:
        self.output_dir = Path(output_dir).resolve()
        self.path = self.output_dir.parent / f".{self.output_dir.name}.retest-publication.lock"
        self._handle: BinaryIO | None = None

    def __enter__(self) -> "_RetestPublicationLock":
        # The publisher creates the output directory after acquiring this
        # lock. A lock probe must not initialize a missing publication root.
        if not self.output_dir.parent.is_dir():
            raise PerformanceV2StoreError("RETEST publication output root does not exist")
        handle = self.path.open("a+b")
        try:
            if os.name == "nt":
                import msvcrt

                if self.path.stat().st_size == 0:
                    handle.write(b"\0")
                    handle.flush()
                handle.seek(0)
                msvcrt.locking(handle.fileno(), msvcrt.LK_NBLCK, 1)
            else:
                import fcntl

                fcntl.flock(handle.fileno(), fcntl.LOCK_EX | fcntl.LOCK_NB)
        except OSError as error:
            handle.close()
            raise PerformanceV2StoreError("RETEST publication output is busy") from error
        self._handle = handle
        return self

    def __exit__(self, exc_type: object, exc_value: object, traceback: object) -> None:
        handle = self._handle
        if handle is None:
            return
        try:
            if os.name == "nt":
                import msvcrt

                handle.seek(0)
                msvcrt.locking(handle.fileno(), msvcrt.LK_UNLCK, 1)
            else:
                import fcntl

                fcntl.flock(handle.fileno(), fcntl.LOCK_UN)
        finally:
            handle.close()
            self._handle = None


@dataclass(frozen=True)
class RetestStatus:
    active_count: int


@dataclass(frozen=True, slots=True)
class RetestBatch:
    """One atomically published strategy batch for the native tester."""

    run_id: str
    strategies_path: Path
    manifest_path: Path
    strategy_count: int

    @property
    def analysis_run_id(self) -> str:
        return self.run_id

    @property
    def batch_id(self) -> str:
        return self.run_id

    @property
    def strategy_source(self) -> Path:
        return self.strategies_path

    @property
    def output_dir(self) -> Path:
        return self.manifest_path.parent


def retest_status(connection: duckdb.DuckDBPyConnection) -> RetestStatus:
    require_performance_v2(connection)
    count = connection.execute(
        """
        select count(*)
        from strategy_tags tags
        join strategies strategies on strategies.strategy_id = tags.strategy_id
        where tags.tag = 'RETEST'
          and strategies.lifecycle_status = 'ACTIVE'
          and strategies.current_result_id is not null
        """
    ).fetchone()[0]
    return RetestStatus(active_count=int(count))


def _canonical(value: object) -> bytes:
    return json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":")).encode("utf-8")


def _strategy_digest(strategy: Mapping[str, object]) -> str:
    value = json.loads(_canonical(strategy).decode("utf-8"))
    provenance = value.get("provenance")
    if isinstance(provenance, dict):
        provenance.pop("strategy_json_sha256", None)
        provenance.pop("generation_manifest_sha256", None)
    return sha256(_canonical(value)).hexdigest()


def _template(templates: Mapping[object, object], side: str) -> Mapping[str, object]:
    try:
        raw = templates[side]
    except (KeyError, TypeError) as error:
        raise PerformanceV2StoreError(f"RETEST template is missing for {side}") from error
    if isinstance(raw, Mapping):
        return deepcopy(dict(raw))
    path = Path(raw) if isinstance(raw, (str, Path)) else None
    try:
        value = json.loads(path.read_text(encoding="utf-8")) if path is not None and path.is_file() else json.loads(str(raw))
    except (OSError, TypeError, UnicodeDecodeError, json.JSONDecodeError) as error:
        raise PerformanceV2StoreError(f"invalid RETEST template for {side}") from error
    if not isinstance(value, Mapping):
        raise PerformanceV2StoreError(f"invalid RETEST template for {side}")
    return deepcopy(dict(value))


def _safe_strategy_name(value: object) -> str:
    if not isinstance(value, str) or not value.strip() or Path(value).name != value:
        raise PerformanceV2StoreError("RETEST strategy name is invalid")
    return value


def _required_text(value: object, field: str, strategy_name: str) -> str:
    if not isinstance(value, str) or not value.strip():
        raise PerformanceV2StoreError(f"RETEST {field} is invalid: {strategy_name}")
    return value.strip()


def _valid_manifest_pair(manifest_path: Path) -> bool:
    if not manifest_path.is_file() or manifest_path.is_symlink():
        return False
    try:
        validate_strategy_manifest(manifest_path)
    except Exception:
        return False
    return True


def _reconcile_retest_publication(output_dir: Path) -> None:
    """Recover or remove interrupted publications before creating a new stage."""
    target = output_dir / "strategies"
    target_manifest = output_dir / "strategy_manifest.json"
    if target.is_symlink() or target_manifest.is_symlink():
        raise PerformanceV2StoreError("RETEST publication output cannot be a symlink")
    if target.exists() and not target.is_dir():
        raise PerformanceV2StoreError("RETEST publication strategies output is not a directory")
    if target_manifest.exists() and not target_manifest.is_file():
        raise PerformanceV2StoreError("RETEST publication manifest output is not a file")

    stages = sorted(
        path for path in output_dir.glob(".retest-stage-*")
        if path.is_dir() and not path.is_symlink()
    )
    backups = sorted(
        path for path in output_dir.glob(".retest-backup-*")
        if path.is_dir() and not path.is_symlink()
    )
    complete_backups = [
        path for path in backups
        if (path / "strategies").is_dir()
        and not (path / "strategies").is_symlink()
        and _valid_manifest_pair(path / "strategy_manifest.json")
    ]
    complete_backups.sort(key=lambda path: path.stat().st_mtime_ns)
    complete_stages = [
        path for path in stages
        if (path / "strategies").is_dir()
        and not (path / "strategies").is_symlink()
        and _valid_manifest_pair(path / "strategy_manifest.json")
    ]
    complete_stages.sort(key=lambda path: path.stat().st_mtime_ns)
    if len(complete_backups) + len(complete_stages) > 1:
        candidates = [*complete_backups, *complete_stages]
        raise PerformanceV2StoreError(
            "ambiguous RETEST publication recovery candidates: "
            + ", ".join(str(path) for path in candidates)
        )

    current_valid = target.is_dir() and _valid_manifest_pair(target_manifest)
    if current_valid:
        for path in (*stages, *backups):
            shutil.rmtree(path)
        return

    if complete_backups:
        backup = complete_backups[-1]
        if target.exists():
            shutil.rmtree(target)
        if target_manifest.exists():
            target_manifest.unlink()
        (backup / "strategies").rename(target)
        (backup / "strategy_manifest.json").rename(target_manifest)
        if not _valid_manifest_pair(target_manifest):
            raise PerformanceV2StoreError(f"RETEST publication backup is invalid: {backup}")
        for path in stages:
            if path.exists():
                shutil.rmtree(path)
        if backup.exists():
            shutil.rmtree(backup)
        return

    if complete_stages and not target.exists() and not target_manifest.exists():
        stage = complete_stages[-1]
        (stage / "strategies").rename(target)
        (stage / "strategy_manifest.json").rename(target_manifest)
        if not _valid_manifest_pair(target_manifest):
            raise PerformanceV2StoreError(f"RETEST publication stage is invalid: {stage}")

    for path in stages:
        if path.exists():
            shutil.rmtree(path)


def _read_retest_rows(connection: duckdb.DuckDBPyConnection) -> list[dict[str, object]]:
    tagged = connection.execute(
        """
        select s.strategy_id
        from strategy_tags tags
        join strategies s on s.strategy_id = tags.strategy_id
        where tags.tag = 'RETEST'
          and s.lifecycle_status = 'ACTIVE'
          and s.current_result_id is not null
        """
    ).fetchall()
    if not tagged:
        raise PerformanceV2StoreError("no active RETEST strategies with current results")
    selected = connection.execute(
        """
        select s.strategy_id, s.strategy_name, s.symbol, s.side, s.timeframe,
               s.close_ma_len, s.order_count, s.analysis_run_id,
               s.candidate_identity, s.current_result_id
        from strategy_tags tags
        join strategies s on s.strategy_id = tags.strategy_id
        join strategy_results r on r.result_id = s.current_result_id
                            and r.strategy_id = s.strategy_id
        where tags.tag = 'RETEST'
          and s.lifecycle_status = 'ACTIVE'
          and s.current_result_id is not null
        order by s.strategy_id
        """
    ).fetchall()
    if len(selected) != len(tagged):
        raise PerformanceV2StoreError("RETEST strategy current result is missing or mismatched")
    order_rows = connection.execute(
        """
        select o.strategy_id, o.order_id, o.open_ma_len, o.open_multiplier,
               o.shift_bp, o.lot_x, o.analysis_run_id, o.plateau_id,
               o.base_point_trades, p.plateau_point_count, p.plateau_total_trades
        from strategy_orders o
        left join analysis_plateaus p
          on p.analysis_run_id = o.analysis_run_id and p.plateau_id = o.plateau_id
        where o.strategy_id in (select strategy_id from strategy_tags where tag = 'RETEST')
        order by o.strategy_id, o.order_id
        """
    ).fetchall()
    by_strategy: dict[int, list[tuple[object, ...]]] = {}
    for row in order_rows:
        by_strategy.setdefault(int(row[0]), []).append(tuple(row))
    result: list[dict[str, object]] = []
    names_by_casefold: dict[str, str] = {}
    for row in selected:
        if type(row[0]) is not int:
            raise PerformanceV2StoreError("RETEST strategy ID is invalid")
        strategy_id = row[0]
        name = _safe_strategy_name(row[1])
        folded_name = name.casefold()
        if folded_name in names_by_casefold:
            raise PerformanceV2StoreError(
                f"RETEST strategy names collide case-insensitively: {names_by_casefold[folded_name]!r}, {name!r}"
            )
        names_by_casefold[folded_name] = name
        order_count = row[6]
        if type(order_count) is not int or not 1 <= order_count <= 4:
            raise PerformanceV2StoreError(f"RETEST strategy order count is invalid: {name}")
        orders = by_strategy.get(strategy_id, [])
        if any(type(item[1]) is not int for item in orders):
            raise PerformanceV2StoreError(f"RETEST order ID is invalid: {name}")
        if len(orders) != order_count or [item[1] for item in orders] != list(range(1, order_count + 1)):
            raise PerformanceV2StoreError(f"RETEST orders do not match strategy {name}")
        symbol = _required_text(row[2], "symbol", name)
        side = _required_text(row[3], "side", name)
        timeframe = _required_text(row[4], "timeframe", name)
        if side not in {"LONG", "SHORT"}:
            raise PerformanceV2StoreError(f"RETEST strategy identity is invalid: {name}")
        if type(row[5]) is not int or row[5] <= 0:
            raise PerformanceV2StoreError(f"RETEST close MA is invalid: {name}")
        strategy_run = _required_text(row[7], "analysis run ID", name)
        candidate = _required_text(row[8], "candidate identity", name)
        diagnostics: list[dict[str, object]] = []
        structure_orders: list[dict[str, object]] = []
        for order in orders:
            if (
                type(order[1]) is not int
                or type(order[2]) is not int
                or type(order[4]) is not int
                or type(order[8]) is not int
                or order[2] <= 0
                or order[4] < 0
                or order[8] < 0
                or not isinstance(order[3], Decimal)
                or not order[3].is_finite()
                or not isinstance(order[5], Decimal)
                or not order[5].is_finite()
                or order[5] <= 0
                or type(order[9]) is not int
                or type(order[10]) is not int
                or order[9] <= 0
                or order[10] < 0
                or not isinstance(order[6], str)
                or not order[6].strip()
                or not isinstance(order[7], str)
                or not order[7].strip()
                or order[6].strip() != strategy_run
            ):
                raise PerformanceV2StoreError(f"RETEST order lineage is invalid: {name}")
            multiplier = Decimal(str(order[3]))
            shift = int(order[4])
            expected_multiplier = Decimal(1) - Decimal(shift) / Decimal(10000) if side == "LONG" else Decimal(1) + Decimal(shift) / Decimal(10000)
            if multiplier != expected_multiplier:
                raise PerformanceV2StoreError(f"RETEST order multiplier does not match shift: {name}")
            diagnostic = {
                "order_id": int(order[1]),
                "plateau_id": str(order[7]),
                "plateau_point_count": int(order[9]),
                "base_point_trades": int(order[8]),
                "plateau_total_trades": int(order[10]),
            }
            diagnostics.append(diagnostic)
            structure_orders.append({
                "id": int(order[1]),
                "open_ma": int(order[2]),
                "shift_bp": shift,
            })
        result.append({
            "strategy_id": strategy_id,
            "name": name,
            "symbol": symbol,
            "side": side,
            "timeframe": timeframe,
            "close_ma": int(row[5]),
            "order_count": order_count,
            "analysis_run_id": strategy_run,
            "candidate_identity": candidate,
            "orders": structure_orders,
            "diagnostics": {"order_count": order_count, "orders": diagnostics},
            "lots": tuple(Decimal(str(item[5])) for item in orders),
        })
    return result


def _publish_retest(
    output_dir: Path,
    generated: list[dict[str, object]],
    manifest: Mapping[str, object],
) -> tuple[Path, Path]:
    """Publish with both process-local and cross-process writer guards."""
    with _RETEST_PUBLICATION_LOCK:
        with _RetestPublicationLock(output_dir):
            return _publish_retest_locked(output_dir, generated, manifest)


def _publish_retest_locked(
    output_dir: Path,
    generated: list[dict[str, object]],
    manifest: Mapping[str, object],
) -> tuple[Path, Path]:
    output_dir.mkdir(parents=True, exist_ok=True)
    _reconcile_retest_publication(output_dir)
    stage = Path(tempfile.mkdtemp(prefix=".retest-stage-", dir=output_dir))
    target = output_dir / "strategies"
    target_manifest = output_dir / "strategy_manifest.json"
    backup = output_dir / f".retest-backup-{uuid4().hex}"
    retain_backup = False
    try:
        variants = pd.DataFrame({"json_filename": [f"{item['name']}.json" for item in generated]})
        _publish_strategies(stage, variants, generated)
        strategy_hashes = manifest.get("strategy_json_sha256")
        if not isinstance(strategy_hashes, Mapping):
            raise PerformanceV2StoreError("RETEST strategy hash map is missing")
        try:
            staged_names = {path.name for path in (stage / "strategies").iterdir()}
        except OSError as error:
            raise PerformanceV2StoreError("RETEST staged strategy directory is unreadable") from error
        if staged_names != set(strategy_hashes):
            raise PerformanceV2StoreError("RETEST staged strategy files do not match manifest hash map")
        for filename, expected_hash in strategy_hashes.items():
            if not isinstance(filename, str) or Path(filename).name != filename or Path(filename).suffix.casefold() != ".json":
                raise PerformanceV2StoreError("RETEST strategy hash map contains an unsafe filename")
            strategy_path = stage / "strategies" / filename
            try:
                strategy_document = json.loads(strategy_path.read_text(encoding="utf-8"))
            except (OSError, UnicodeDecodeError, json.JSONDecodeError) as error:
                raise PerformanceV2StoreError(f"RETEST staged strategy is invalid: {filename}") from error
            if not isinstance(strategy_document, Mapping) or _strategy_digest(strategy_document) != expected_hash:
                raise PerformanceV2StoreError(f"RETEST staged strategy hash mismatch: {filename}")
        _write_json_atomic(stage / "strategy_manifest.json", manifest)
        validate_strategy_manifest(stage / "strategy_manifest.json")
        target_was_present = target.exists()
        manifest_was_present = target_manifest.exists()
        backup.mkdir()
        if target.is_symlink() or (target.exists() and not target.is_dir()):
            raise PerformanceV2StoreError("strategy output is not a directory")
        if target.exists():
            target.rename(backup / "strategies")
        if target_manifest.is_symlink() or (target_manifest.exists() and not target_manifest.is_file()):
            raise PerformanceV2StoreError("strategy manifest output is not a file")
        if target_manifest.exists():
            target_manifest.rename(backup / "strategy_manifest.json")
        (stage / "strategies").rename(target)
        (stage / "strategy_manifest.json").rename(target_manifest)
        shutil.rmtree(backup)
    except BaseException as error:
        if not backup.exists():
            raise
        try:
            # A rename can complete before an injected/OS error is raised, so
            # inspect destinations and the pre-publication state rather than
            # relying only on the flags.
            old_target_backup = backup / "strategies"
            if old_target_backup.exists():
                if target.exists():
                    shutil.rmtree(target)
                old_target_backup.rename(target)
            elif target.exists() and not target_was_present:
                shutil.rmtree(target)
            old_manifest_backup = backup / "strategy_manifest.json"
            if old_manifest_backup.exists():
                if target_manifest.exists():
                    target_manifest.unlink()
                old_manifest_backup.rename(target_manifest)
            elif target_manifest.exists() and not manifest_was_present:
                target_manifest.unlink()
            shutil.rmtree(backup)
        except BaseException as rollback_error:
            retain_backup = True
            raise PerformanceV2StoreError(
                f"RETEST publication failed: {error}; rollback failed: {rollback_error}; "
                f"backup retained at {backup}"
            ) from rollback_error
        raise error
    finally:
        if stage.exists():
            try:
                shutil.rmtree(stage)
            except OSError:
                if not retain_backup:
                    raise
        if backup.exists() and not retain_backup:
            shutil.rmtree(backup)
    return target, target_manifest


def build_retest_manifest(
    connection: duckdb.DuckDBPyConnection,
    templates: Mapping[object, object],
    output_dir: Path,
) -> RetestBatch:
    """Render active RETEST strategies from typed Performance v2 rows."""
    require_performance_v2(connection)
    if not isinstance(templates, Mapping):
        raise PerformanceV2StoreError("RETEST templates must be a side mapping")
    rows = _read_retest_rows(connection)
    batch_id = f"retest-{uuid4().hex}"
    generated: list[dict[str, object]] = []
    strategy_runs: dict[str, str] = {}
    candidate_names: dict[str, list[str]] = {}
    candidate_diagnostics: dict[str, Mapping[str, object]] = {}
    for row in rows:
        name = str(row["name"])
        # EQUAL only supplies generate_strategy's transient name; stored lots
        # below are authoritative and are passed through unchanged.
        strategy = generate_strategy(
            _template(templates, str(row["side"])),
            {
                "symbol": row["symbol"],
                "side": row["side"],
                "timeframe": row["timeframe"],
                "common_close_ma": row["close_ma"],
                "order_count": row["order_count"],
                "structure_id": f"RETEST_{row['strategy_id']}",
                "orders": row["orders"],
            },
            row["lots"],
            LotMethod.EQUAL,
            AlgorithmConfig.defaults(),
        )
        strategy["name"] = name
        identity = adapt_strategy_identity(
            strategy,
            strategy_name=name,
            order_plateau_diagnostics=row["diagnostics"],
        )
        expected_orders = tuple(row["orders"])
        if (
            identity.symbol != row["symbol"]
            or identity.side != row["side"]
            or identity.timeframe != row["timeframe"]
            or identity.close_ma_len != row["close_ma"]
            or identity.order_count != row["order_count"]
            or any(
                (actual.open_ma_len, actual.shift_bp, actual.lot_x, actual.plateau_id,
                 actual.plateau_point_count, actual.base_point_trades, actual.plateau_total_trades)
                != (int(expected["open_ma"]), int(expected["shift_bp"]), Decimal(str(lot)),
                    str(diagnostic["plateau_id"]), int(diagnostic["plateau_point_count"]),
                    int(diagnostic["base_point_trades"]), int(diagnostic["plateau_total_trades"]))
                for actual, expected, diagnostic, lot in zip(
                    identity.orders, expected_orders, row["diagnostics"]["orders"], row["lots"], strict=True
                )
            )
        ):
            raise PerformanceV2StoreError(f"generated RETEST strategy identity mismatch: {name}")
        filename = f"{name}.json"
        strategy_runs[filename] = str(row["analysis_run_id"])
        candidate = str(row["candidate_identity"])
        candidate_names.setdefault(candidate, []).append(name)
        previous = candidate_diagnostics.get(candidate)
        if previous is not None and previous != row["diagnostics"]:
            raise PerformanceV2StoreError("candidate has conflicting RETEST plateau diagnostics")
        candidate_diagnostics[candidate] = row["diagnostics"]
        generated.append(strategy)

    hashes = {f"{row['name']}.json": _strategy_digest(strategy) for row, strategy in zip(rows, generated, strict=True)}
    unsigned: dict[str, object] = {
        "format_version": 1,
        "analysis_run_id": batch_id,
        "event_mode": "real_independent_events",
        "strategy_count": len(generated),
        "strategy_json_sha256": hashes,
        "strategy_analysis_run_ids": strategy_runs,
        "candidate_identity_to_strategy_names": {
            key: sorted(names) for key, names in sorted(candidate_names.items())
        },
        "candidate_diagnostics": {
            key: candidate_diagnostics[key] for key in sorted(candidate_diagnostics)
        },
    }
    manifest = {**unsigned, "generation_manifest_sha256": sha256(_canonical(unsigned)).hexdigest()}
    strategies_path, manifest_path = _publish_retest(Path(output_dir).resolve(), generated, manifest)
    return RetestBatch(batch_id, strategies_path, manifest_path, len(generated))


def _audit_ids(path: Path) -> set[int]:
    workbook = None
    try:
        workbook = load_workbook(path, read_only=True, data_only=False)
        if not {"HIGH", "REVIEW"}.issubset(workbook.sheetnames):
            raise PerformanceV2StoreError("audit workbook requires HIGH and REVIEW sheets")
        ids: set[int] = set()
        for sheet_name in ("HIGH", "REVIEW"):
            sheet = workbook[sheet_name]
            sheet.reset_dimensions()
            header = next(sheet.iter_rows(min_row=1, max_row=1), ())
            matches = [index for index, cell in enumerate(header) if cell.value == "Strategy ID"]
            if len(matches) != 1:
                raise PerformanceV2StoreError("audit workbook requires one exact Strategy ID header")
            target = matches[0]
            for row in sheet.iter_rows(min_row=2):
                values = [cell.value for cell in row]
                if all(value is None or value == "" for value in values):
                    continue
                if target >= len(values) or values[target] is None or values[target] == "":
                    raise PerformanceV2StoreError("audit workbook has a missing Strategy ID")
                value = values[target]
                if not isinstance(value, int) or isinstance(value, bool):
                    raise PerformanceV2StoreError("audit workbook Strategy ID must be an integer")
                strategy_id = int(value)
                if strategy_id in ids:
                    raise PerformanceV2StoreError("audit workbook contains duplicate Strategy ID")
                ids.add(strategy_id)
        return ids
    except PerformanceV2StoreError:
        raise
    except Exception as error:
        raise PerformanceV2StoreError("could not read audit workbook") from error
    finally:
        if workbook is not None:
            workbook.close()


def mark_retest_from_audit(connection: duckdb.DuckDBPyConnection, path: Path) -> int:
    require_performance_v2(connection)
    ids = _audit_ids(Path(path))
    if not ids:
        return 0
    known = {
        int(row[0])
        for row in connection.execute(
            "select strategy_id from strategies where strategy_id in (select unnest(?::bigint[]))",
            [sorted(ids)],
        ).fetchall()
    }
    unknown = ids - known
    if unknown:
        raise PerformanceV2StoreError("audit workbook references unknown Strategy ID")

    source_ref = Path(path).name
    try:
        connection.execute("begin transaction")
        for strategy_id in sorted(ids):
            connection.execute(
                """
                insert into strategy_tags (strategy_id, tag, source, source_ref, updated_at_utc)
                values (?, 'RETEST', 'PERIOD_INTEGRITY_AUDIT', ?, now())
                on conflict (strategy_id, tag) do update set
                    source = excluded.source,
                    source_ref = excluded.source_ref,
                    updated_at_utc = excluded.updated_at_utc
                """,
                [strategy_id, source_ref],
            )
        connection.execute("commit")
    except Exception as error:
        try:
            connection.execute("rollback")
        except Exception:
            pass
        if isinstance(error, PerformanceV2StoreError):
            raise
        raise PerformanceV2StoreError("could not mark RETEST strategies") from error
    return len(ids)
